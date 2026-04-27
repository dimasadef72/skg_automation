import os
import csv
import numpy as np
import pandas as pd
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# === Impor dari Modul Terpisah ===
from kalman_module import process_kalman
from kuantisasi_module import process_kuantisasi
try:
    from bch_module import process_bch
    from hash_module import process_hash
    from nist_module import process_nist
except ImportError:
    pass  # Akan dibuat nanti

# =====================================================================
# GLOBAL PARAMETERS
# =====================================================================
KUANTISASI_NUM_BITS = 3
BENCHMARK_ITERATIONS = 10
CHANPROB_TIME_SECONDS = 120.0  # Hardcoded sementara sesuai arahan: waktu channel probing
BLOCK_SIZE_OPTIONS = [100, 200]
FULL_MODE_KEY = "full"
REPORT_PANEL_ORDER = [
    (FULL_MODE_KEY, "Data Full"),
    (100, "Blok 100"),
    (200, "Blok 200"),
]

# Variasi Parameter Pengujian Skenario
PARAM_VARIATIONS = [
    {"q": 0.01, "r": 0.5, "bb": 1},
    {"q": 0.01, "r": 0.5, "bb": 5},
    {"q": 0.01, "r": 0.5, "bb": 50},
    {"q": 0.01, "r": 0.5, "bb": 100},
    {"q": 0.01, "r": 0.5, "bb": 200},
    {"q": 0.5, "r": 0.01, "bb": 1},
    {"q": 0.5, "r": 0.01, "bb": 5},
    {"q": 0.5, "r": 0.01, "bb": 50},
    {"q": 0.5, "r": 0.01, "bb": 100},
    {"q": 0.5, "r": 0.01, "bb": 200},
]

# Skenario iterasi 1 - 4
SCENARIOS = [1, 2, 3, 4]

# =====================================================================
# UTILITY FUNCTIONS
# =====================================================================
def read_rssi_csv(path):
    data = []
    if not os.path.exists(path):
        print(f"Warning: File {path} tidak ditemukan!")
        return data
        
    with open(path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                data.append(int(row[0]))
            except:
                continue
    return data

def split_full_blocks(data, block_size):
    """Potong data menjadi blok utuh; sisa data di akhir dibuang."""
    if block_size <= 0:
        return [], 0
    total_full = (len(data) // block_size) * block_size
    if total_full == 0:
        return [], 0
    trimmed = data[:total_full]
    blocks = [trimmed[i:i + block_size] for i in range(0, total_full, block_size)]
    return blocks, total_full

def process_kalman_per_blocks(raw_data, q, r, block_size, kalman_bb, benchmark_iterations=BENCHMARK_ITERATIONS):
    """Jalankan Kalman per blok dan kembalikan list hasil per blok + waktu per blok."""
    blocks, _ = split_full_blocks(raw_data, block_size)
    if not blocks:
        return [], []

    kalman_blocks = []
    kalman_times = []
    for block in blocks:
        kal_block, _, time_block = process_kalman(block, q, r, kalman_bb, benchmark_iterations=benchmark_iterations)
        kalman_blocks.append(kal_block)
        kalman_times.append(float(time_block))
    return kalman_blocks, kalman_times

def process_kuantisasi_per_blocks(kalman_blocks, num_bits=KUANTISASI_NUM_BITS, benchmark_iterations=BENCHMARK_ITERATIONS):
    """Jalankan kuantisasi per blok hasil Kalman."""
    if not kalman_blocks:
        return [], [], "", 0.0, 0.0

    bitstreams = []
    block_kgr_locals = []
    block_times = []
    total_time = 0.0
    for block in kalman_blocks:
        bs_block, _, time_block = process_kuantisasi(block, num_bits, benchmark_iterations)
        bitstreams.append(bs_block)
        block_times.append(float(time_block))
        block_kgr_locals.append(calculate_local_kgr(len(bs_block), float(time_block)))
        total_time += float(time_block)

    merged_bitstream = "".join(bitstreams)
    kgr_local = calculate_local_kgr(len(merged_bitstream), total_time)
    return bitstreams, block_kgr_locals, merged_bitstream, kgr_local, total_time

def average_numeric(values):
    vals = [float(v) for v in values if v is not None]
    if not vals:
        return 0.0
    return float(np.mean(vals))

def average_corr(values):
    vals = [float(v) for v in values if isinstance(v, (int, float, np.floating))]
    if not vals:
        return "N/A"
    return float(np.mean(vals))

def calculate_kdr(a, b):
    if not a or not b: return 0.0
    n = min(len(a), len(b))
    if n == 0: return 0.0
    diff = sum(1 for i in range(n) if a[i] != b[i])
    return (diff / n) * 100.0

def calc_corr(a, b):
    ln = min(len(a), len(b))
    if ln < 2: return "N/A"
    c, _ = pearsonr(a[:ln], b[:ln])
    return c

def calculate_cumulative_kgr(total_bits, *time_parts):
    """Hitung KGR kumulatif: total_bits / total_waktu_akumulasi."""
    total_time = 0.0
    for t in time_parts:
        try:
            total_time += float(t)
        except (TypeError, ValueError):
            continue
    if total_time <= 0:
        return 0.0
    return float(total_bits) / total_time

def calculate_local_kgr(total_bits, elapsed):
    """Hitung KGR lokal: total_bits / waktu pada stage tersebut."""
    try:
        elapsed = float(elapsed)
    except (TypeError, ValueError):
        return 0.0
    if elapsed <= 0:
        return 0.0
    return float(total_bits) / elapsed

def calculate_kdr_from_matched_bits(matched_bits, compared_bits):
    """Hitung KDR final (%) dari mismatch bits terhadap total compared bits."""
    try:
        matched_bits = float(matched_bits)
        compared_bits = float(compared_bits)
    except (TypeError, ValueError):
        return 0.0
    if compared_bits <= 0:
        return 0.0
    mismatch = max(0.0, compared_bits - matched_bits)
    return (mismatch / compared_bits) * 100.0
# =====================================================================
# EXCEL FORMATTING FUNCTIONS
# =====================================================================
def save_data_list(output_dir, filename, data_list, header):
    os.makedirs(output_dir, exist_ok=True)
    df = pd.DataFrame({header: data_list})
    df.to_excel(os.path.join(output_dir, filename), index=False)

def save_workbook_safely(workbook, target_path):
    """Simpan workbook ke target_path; jika file terkunci, simpan ke nama alternatif."""
    os.makedirs(os.path.dirname(target_path), exist_ok=True)
    try:
        workbook.save(target_path)
        return target_path
    except PermissionError:
        base_name, ext = os.path.splitext(target_path)
        alt_path = f"{base_name}_baru{ext}"
        workbook.save(alt_path)
        return alt_path

def pair_records_by_param(records):
    pairs = {}
    for r in records:
        key = (r.get("skenario"), r.get("q"), r.get("r"), r.get("bb"))
        pairs.setdefault(key, {})[r.get("block_size")] = r
    sorted_keys = sorted(pairs.keys(), key=lambda x: (x[0], x[1], x[2], x[3]))
    return [(k, pairs[k]) for k in sorted_keys]

def iter_report_panels(by_block, start_cols):
    for (block_key, label), start_col in zip(REPORT_PANEL_ORDER, start_cols):
        rec = by_block.get(block_key)
        if rec is not None:
            yield start_col, rec, label

def build_kalman_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kalman"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_kalman_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+8)

        ws.cell(row=start_row, column=start_col, value="Parameter").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row+1, end_column=start_col)

        ws.cell(row=start_row, column=start_col+1, value="Sebelum Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col+1, end_row=start_row, end_column=start_col+4)

        ws.cell(row=start_row, column=start_col+5, value="Setelah Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col+5, end_row=start_row, end_column=start_col+8)

        cols_names = ["Alice", "Bob", "Eve-Alice", "Eve-Bob", "Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, cname in enumerate(cols_names):
            ws.cell(row=start_row+1, column=start_col+1+idx, value=cname).font = header_font

        ws.cell(row=start_row+2, column=start_col, value="Maksimum (dBm)")
        vals_max = [rec['orig_max_alice'], rec['orig_max_bob'], rec['orig_max_evealice'], rec['orig_max_evebob'],
                    rec['kalman_max_alice'], rec['kalman_max_bob'], rec['kalman_max_evealice'], rec['kalman_max_evebob']]
        for idx, val in enumerate(vals_max):
            ws.cell(row=start_row+2, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+3, column=start_col, value="Minimum (dBm)")
        vals_min = [rec['orig_min_alice'], rec['orig_min_bob'], rec['orig_min_evealice'], rec['orig_min_evebob'],
                    rec['kalman_min_alice'], rec['kalman_min_bob'], rec['kalman_min_evealice'], rec['kalman_min_evebob']]
        for idx, val in enumerate(vals_min):
            ws.cell(row=start_row+3, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+4, column=start_col, value="Koefisien Korelasi")
        c1 = ws.cell(row=start_row+4, column=start_col+1, value=rec['orig_corr_ab'])
        c1.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+1, end_row=start_row+4, end_column=start_col+2)
        c2 = ws.cell(row=start_row+4, column=start_col+3, value=rec['orig_corr_eve'])
        c2.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+3, end_row=start_row+4, end_column=start_col+4)
        c3 = ws.cell(row=start_row+4, column=start_col+5, value=rec['kalman_corr_ab'])
        c3.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+5, end_row=start_row+4, end_column=start_col+6)
        c4 = ws.cell(row=start_row+4, column=start_col+7, value=rec['kalman_corr_eve'])
        c4.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+7, end_row=start_row+4, end_column=start_col+8)

        ws.cell(row=start_row+5, column=start_col, value="Waktu Komputasi (s)")
        ws.merge_cells(start_row=start_row+5, start_column=start_col, end_row=start_row+5, end_column=start_col+4)
        for idx, val in enumerate([rec['time_alice'], rec['time_bob'], rec['time_evealice'], rec['time_evebob']]):
            ws.cell(row=start_row+5, column=start_col+5+idx, value=val)

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+5, min_col=start_col, max_col=start_col+8):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=29)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 11, 21]):
            write_kalman_table(start_row, start_col, rec, label)

        current_row = start_row + 9

    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['U'].width = 25
        
    wb.save(output_path)

def build_kuantisasi_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kuantisasi"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_kuan_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+8)

        ws.cell(row=start_row, column=start_col, value="Parameter Performansi (Rata-rata per Blok)").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row, column=start_col+6, value="Metrik Bitstream Gabungan").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col+6, end_row=start_row, end_column=start_col+8)
        ws.cell(row=start_row+1, column=start_col+6, value="Parameter").font = header_font
        ws.cell(row=start_row+1, column=start_col+7, value="A & B").font = header_font
        ws.cell(row=start_row+1, column=start_col+8, value="E-A & E-B").font = header_font

        ws.cell(row=start_row+1, column=start_col, value="KDR (%)")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['kdr_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=start_row+1, end_column=start_col+2)
        ws.cell(row=start_row+1, column=start_col+3, value=rec['kdr_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+3, end_row=start_row+1, end_column=start_col+4)

        ws.cell(row=start_row+2, column=start_col, value="KGR Lokal (bit/s)")
        for idx, val in enumerate([rec['kgr_local_alice'], rec['kgr_local_bob'], rec['kgr_local_evealice'], rec['kgr_local_evebob']]):
            ws.cell(row=start_row+2, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+3, column=start_col, value="KGR Kumulatif (bit/s)")
        for idx, val in enumerate([rec['kgr_cum_alice'], rec['kgr_cum_bob'], rec['kgr_cum_evealice'], rec['kgr_cum_evebob']]):
            ws.cell(row=start_row+3, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+4, column=start_col, value="Total Bit Dihasilkan")
        for idx, val in enumerate([rec['total_bits_alice'], rec['total_bits_bob'], rec['total_bits_ea'], rec['total_bits_eb']]):
            ws.cell(row=start_row+4, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+5, column=start_col, value="Waktu komputasi (s)")
        for idx, val in enumerate([rec['time_alice'], rec['time_bob'], rec['time_evealice'], rec['time_evebob']]):
            ws.cell(row=start_row+5, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+2, column=start_col+6, value="KDR Gabungan (%)")
        ws.cell(row=start_row+2, column=start_col+7, value=rec['kdr_merge_ab'])
        ws.cell(row=start_row+2, column=start_col+8, value=rec['kdr_merge_eve'])
        ws.cell(row=start_row+3, column=start_col+6, value="KGR Lokal Gabungan (bit/s)")
        ws.cell(row=start_row+3, column=start_col+7, value=rec['kgr_merge_local_ab'])
        ws.cell(row=start_row+3, column=start_col+8, value=rec['kgr_merge_local_eve'])
        ws.cell(row=start_row+4, column=start_col+6, value="KGR Kumulatif Gabungan (bit/s)")
        ws.cell(row=start_row+4, column=start_col+7, value=rec['kgr_merge_cum_ab'])
        ws.cell(row=start_row+4, column=start_col+8, value=rec['kgr_merge_cum_eve'])
        ws.cell(row=start_row+5, column=start_col+6, value="Total Bit Gabungan")
        ws.cell(row=start_row+5, column=start_col+7, value=rec['total_bits_merge_ab'])
        ws.cell(row=start_row+5, column=start_col+8, value=rec['total_bits_merge_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+5, min_col=start_col, max_col=start_col+8):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=29)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 11, 21]):
            write_kuan_table(start_row, start_col, rec, label)

        current_row = start_row + 9

    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 20
        
    wb.save(output_path)

def build_bch_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap BCH"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_bch_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+2)

        ws.cell(row=start_row, column=start_col, value="Parameter BCH").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row+1, column=start_col, value="KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['kdr_after_ab'])
        ws.cell(row=start_row+1, column=start_col+2, value=rec['kdr_after_eve'])

        ws.cell(row=start_row+2, column=start_col, value="KGR BCH Lokal (bit/s)")
        ws.cell(row=start_row+2, column=start_col+1, value=rec['kgr_bch_ab_local'])
        ws.cell(row=start_row+2, column=start_col+2, value=rec['kgr_bch_eve_local'])

        ws.cell(row=start_row+3, column=start_col, value="KGR BCH Kumulatif (bit/s)")
        ws.cell(row=start_row+3, column=start_col+1, value=rec['kgr_bch_ab'])
        ws.cell(row=start_row+3, column=start_col+2, value=rec['kgr_bch_eve'])

        ws.cell(row=start_row+4, column=start_col, value="Parity Bits Dikirim")
        ws.cell(row=start_row+4, column=start_col+1, value=rec['parity_bits_ab'])
        ws.cell(row=start_row+4, column=start_col+2, value=rec['parity_bits_eve'])

        ws.cell(row=start_row+5, column=start_col, value="Total Bit (A/B | E-A/E-B)")
        ws.cell(row=start_row+5, column=start_col+1, value=f"{rec['total_bits_alice']}/{rec['total_bits_bob']}")
        ws.cell(row=start_row+5, column=start_col+2, value=f"{rec['total_bits_ea']}/{rec['total_bits_eb']}")

        ws.cell(row=start_row+6, column=start_col, value="Error Bit Sebelum")
        ws.cell(row=start_row+6, column=start_col+1, value=rec['error_bits_ab_before'])
        ws.cell(row=start_row+6, column=start_col+2, value=rec['error_bits_eve_before'])

        ws.cell(row=start_row+7, column=start_col, value="Error Bit Setelah")
        ws.cell(row=start_row+7, column=start_col+1, value=rec['error_bits_ab_after'])
        ws.cell(row=start_row+7, column=start_col+2, value=rec['error_bits_eve_after'])

        ws.cell(row=start_row+8, column=start_col, value="Bit Terkoreksi/Disamakan")
        ws.cell(row=start_row+8, column=start_col+1, value=rec['corrected_bits_ab'])
        ws.cell(row=start_row+8, column=start_col+2, value=rec['corrected_bits_eve'])

        ws.cell(row=start_row+9, column=start_col, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+9, column=start_col+1, value=rec['time_bch_ab'])
        ws.cell(row=start_row+9, column=start_col+2, value=rec['time_bch_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+9, min_col=start_col, max_col=start_col+2):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 5, 9]):
            write_bch_table(start_row, start_col, rec, label)

        current_row = start_row + 12

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 20
    wb.save(output_path)
def build_hash_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Hash"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_hash_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+4)

        ws.cell(row=start_row, column=start_col, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row+1, column=start_col, value="Jumlah Kunci Cocok (Match)")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['aes_count_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=start_row+1, end_column=start_col+2)
        ws.cell(row=start_row+1, column=start_col+3, value=rec['aes_count_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+3, end_row=start_row+1, end_column=start_col+4)

        ws.cell(row=start_row+2, column=start_col, value="Jumlah Kandidat Key")
        ws.cell(row=start_row+2, column=start_col+1, value=rec['keys_count_alice'])
        ws.cell(row=start_row+2, column=start_col+2, value=rec['keys_count_bob'])
        ws.cell(row=start_row+2, column=start_col+3, value=rec['keys_count_ea'])
        ws.cell(row=start_row+2, column=start_col+4, value=rec['keys_count_eb'])

        ws.cell(row=start_row+3, column=start_col, value="Total Bit Key (128*jumlah_key)")
        ws.cell(row=start_row+3, column=start_col+1, value=rec['total_key_bits_alice'])
        ws.cell(row=start_row+3, column=start_col+2, value=rec['total_key_bits_bob'])
        ws.cell(row=start_row+3, column=start_col+3, value=rec['total_key_bits_ea'])
        ws.cell(row=start_row+3, column=start_col+4, value=rec.get('total_key_bits_eb', 'N/A'))

        ws.cell(row=start_row+4, column=start_col, value="Total Bit AES Match")
        ws.cell(row=start_row+4, column=start_col+1, value=rec['matched_key_bits_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=start_col+1, end_row=start_row+4, end_column=start_col+2)
        ws.cell(row=start_row+4, column=start_col+3, value=rec['matched_key_bits_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=start_col+3, end_row=start_row+4, end_column=start_col+4)

        ws.cell(row=start_row+5, column=start_col, value="Kunci Pertama (Hex)")
        ws.cell(row=start_row+5, column=start_col+1, value=rec['final_key_alice'])
        ws.cell(row=start_row+5, column=start_col+2, value=rec['final_key_bob'])
        ws.cell(row=start_row+5, column=start_col+3, value=rec['final_key_ea'])
        ws.cell(row=start_row+5, column=start_col+4, value=rec['final_key_eb'])

        ws.cell(row=start_row+6, column=start_col, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+6, column=start_col+1, value=rec['time_hash_ab'])
        ws.cell(row=start_row+6, column=start_col+2, value=rec['time_hash_ab'])
        ws.cell(row=start_row+6, column=start_col+3, value=rec['time_hash_eve'])
        ws.cell(row=start_row+6, column=start_col+4, value=rec['time_hash_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+6, min_col=start_col, max_col=start_col+4):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 7, 13]):
            write_hash_table(start_row, start_col, rec, label)

        current_row = start_row + 9

    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 28
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['M'].width = 25
    wb.save(output_path)

def build_nist_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap NIST"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_nist_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+2)

        ws.cell(row=start_row, column=start_col, value="Parameter NIST").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row+1, column=start_col, value="Jumlah Key Lulus")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['passed_keys_ab'])
        ws.cell(row=start_row+1, column=start_col+2, value=rec['passed_keys_eve'])

        ws.cell(row=start_row+2, column=start_col, value="Rata-rata p-value (ApEn)")
        ws.cell(row=start_row+2, column=start_col+1, value=rec['pval_ab'])
        ws.cell(row=start_row+2, column=start_col+2, value=rec['pval_eve'])

        ws.cell(row=start_row+3, column=start_col, value="Pass Rate (%)")
        ws.cell(row=start_row+3, column=start_col+1, value=rec['pass_rate_ab'])
        ws.cell(row=start_row+3, column=start_col+2, value=rec['pass_rate_eve'])

        ws.cell(row=start_row+4, column=start_col, value="Distribusi p-value")
        ws.cell(row=start_row+4, column=start_col+1, value=rec['pval_dist_ab'])
        ws.cell(row=start_row+4, column=start_col+2, value=rec['pval_dist_eve'])

        ws.cell(row=start_row+5, column=start_col, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+5, column=start_col+1, value=rec['time_nist_ab'])
        ws.cell(row=start_row+5, column=start_col+2, value=rec['time_nist_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+5, min_col=start_col, max_col=start_col+2):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 5, 9]):
            write_nist_table(start_row, start_col, rec, label)

        current_row = start_row + 8

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 25
    wb.save(output_path)

def build_kalman_sheet(wb, records):
    ws = wb.create_sheet(title="Kalman")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_kalman_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+8)

        ws.cell(row=start_row, column=start_col, value="Parameter").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row+1, end_column=start_col)

        ws.cell(row=start_row, column=start_col+1, value="Sebelum Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col+1, end_row=start_row, end_column=start_col+4)

        ws.cell(row=start_row, column=start_col+5, value="Setelah Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col+5, end_row=start_row, end_column=start_col+8)

        cols_names = ["Alice", "Bob", "Eve-Alice", "Eve-Bob", "Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, cname in enumerate(cols_names):
            ws.cell(row=start_row+1, column=start_col+1+idx, value=cname).font = header_font

        ws.cell(row=start_row+2, column=start_col, value="Maksimum (dBm)")
        vals_max = [rec['orig_max_alice'], rec['orig_max_bob'], rec['orig_max_evealice'], rec['orig_max_evebob'],
                    rec['kalman_max_alice'], rec['kalman_max_bob'], rec['kalman_max_evealice'], rec['kalman_max_evebob']]
        for idx, val in enumerate(vals_max):
            ws.cell(row=start_row+2, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+3, column=start_col, value="Minimum (dBm)")
        vals_min = [rec['orig_min_alice'], rec['orig_min_bob'], rec['orig_min_evealice'], rec['orig_min_evebob'],
                    rec['kalman_min_alice'], rec['kalman_min_bob'], rec['kalman_min_evealice'], rec['kalman_min_evebob']]
        for idx, val in enumerate(vals_min):
            ws.cell(row=start_row+3, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+4, column=start_col, value="Koefisien Korelasi")
        c1 = ws.cell(row=start_row+4, column=start_col+1, value=rec['orig_corr_ab'])
        c1.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+1, end_row=start_row+4, end_column=start_col+2)
        c2 = ws.cell(row=start_row+4, column=start_col+3, value=rec['orig_corr_eve'])
        c2.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+3, end_row=start_row+4, end_column=start_col+4)
        c3 = ws.cell(row=start_row+4, column=start_col+5, value=rec['kalman_corr_ab'])
        c3.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+5, end_row=start_row+4, end_column=start_col+6)
        c4 = ws.cell(row=start_row+4, column=start_col+7, value=rec['kalman_corr_eve'])
        c4.number_format = '0.0000000000'
        ws.merge_cells(start_row=start_row+4, start_column=start_col+7, end_row=start_row+4, end_column=start_col+8)

        ws.cell(row=start_row+5, column=start_col, value="Waktu Komputasi (s)")
        ws.merge_cells(start_row=start_row+5, start_column=start_col, end_row=start_row+5, end_column=start_col+4)
        for idx, val in enumerate([rec['time_alice'], rec['time_bob'], rec['time_evealice'], rec['time_evebob']]):
            ws.cell(row=start_row+5, column=start_col+5+idx, value=val)

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+5, min_col=start_col, max_col=start_col+8):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=29)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 11, 21]):
            write_kalman_table(start_row, start_col, rec, label)

        current_row = start_row + 9

    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['U'].width = 25

def build_kuantisasi_sheet(wb, records):
    ws = wb.create_sheet(title="Kuantisasi")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_kuan_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+8)

        ws.cell(row=start_row, column=start_col, value="Parameter Performansi (Rata-rata per Blok)").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row, column=start_col+6, value="Metrik Bitstream Gabungan").font = header_font
        ws.merge_cells(start_row=start_row, start_column=start_col+6, end_row=start_row, end_column=start_col+8)
        ws.cell(row=start_row+1, column=start_col+6, value="Parameter").font = header_font
        ws.cell(row=start_row+1, column=start_col+7, value="A & B").font = header_font
        ws.cell(row=start_row+1, column=start_col+8, value="E-A & E-B").font = header_font

        ws.cell(row=start_row+1, column=start_col, value="KDR (%)")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['kdr_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=start_row+1, end_column=start_col+2)
        ws.cell(row=start_row+1, column=start_col+3, value=rec['kdr_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+3, end_row=start_row+1, end_column=start_col+4)

        ws.cell(row=start_row+2, column=start_col, value="KGR Lokal (bit/s)")
        for idx, val in enumerate([rec['kgr_local_alice'], rec['kgr_local_bob'], rec['kgr_local_evealice'], rec['kgr_local_evebob']]):
            ws.cell(row=start_row+2, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+3, column=start_col, value="KGR Kumulatif (bit/s)")
        for idx, val in enumerate([rec['kgr_cum_alice'], rec['kgr_cum_bob'], rec['kgr_cum_evealice'], rec['kgr_cum_evebob']]):
            ws.cell(row=start_row+3, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+4, column=start_col, value="Total Bit Dihasilkan")
        for idx, val in enumerate([rec['total_bits_alice'], rec['total_bits_bob'], rec['total_bits_ea'], rec['total_bits_eb']]):
            ws.cell(row=start_row+4, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+5, column=start_col, value="Waktu komputasi (s)")
        for idx, val in enumerate([rec['time_alice'], rec['time_bob'], rec['time_evealice'], rec['time_evebob']]):
            ws.cell(row=start_row+5, column=start_col+1+idx, value=val)

        ws.cell(row=start_row+2, column=start_col+6, value="KDR Gabungan (%)")
        ws.cell(row=start_row+2, column=start_col+7, value=rec['kdr_merge_ab'])
        ws.cell(row=start_row+2, column=start_col+8, value=rec['kdr_merge_eve'])

        ws.cell(row=start_row+3, column=start_col+6, value="KGR Lokal Gabungan (bit/s)")
        ws.cell(row=start_row+3, column=start_col+7, value=rec['kgr_merge_local_ab'])
        ws.cell(row=start_row+3, column=start_col+8, value=rec['kgr_merge_local_eve'])

        ws.cell(row=start_row+4, column=start_col+6, value="KGR Kumulatif Gabungan (bit/s)")
        ws.cell(row=start_row+4, column=start_col+7, value=rec['kgr_merge_cum_ab'])
        ws.cell(row=start_row+4, column=start_col+8, value=rec['kgr_merge_cum_eve'])

        ws.cell(row=start_row+5, column=start_col+6, value="Total Bit Gabungan")
        ws.cell(row=start_row+5, column=start_col+7, value=rec['total_bits_merge_ab'])
        ws.cell(row=start_row+5, column=start_col+8, value=rec['total_bits_merge_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+5, min_col=start_col, max_col=start_col+8):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=29)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 11, 21]):
            write_kuan_table(start_row, start_col, rec, label)

        current_row = start_row + 9

    for col in range(1, 30):
        ws.column_dimensions[get_column_letter(col)].width = 20

def build_bch_sheet(wb, records):
    ws = wb.create_sheet(title="BCH")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_bch_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+2)

        ws.cell(row=start_row, column=start_col, value="Parameter BCH").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row+1, column=start_col, value="KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['kdr_after_ab'])
        ws.cell(row=start_row+1, column=start_col+2, value=rec['kdr_after_eve'])

        ws.cell(row=start_row+2, column=start_col, value="KGR BCH Lokal (bit/s)")
        ws.cell(row=start_row+2, column=start_col+1, value=rec['kgr_bch_ab_local'])
        ws.cell(row=start_row+2, column=start_col+2, value=rec['kgr_bch_eve_local'])

        ws.cell(row=start_row+3, column=start_col, value="KGR BCH Kumulatif (bit/s)")
        ws.cell(row=start_row+3, column=start_col+1, value=rec['kgr_bch_ab'])
        ws.cell(row=start_row+3, column=start_col+2, value=rec['kgr_bch_eve'])

        ws.cell(row=start_row+4, column=start_col, value="Parity Bits Dikirim")
        ws.cell(row=start_row+4, column=start_col+1, value=rec['parity_bits_ab'])
        ws.cell(row=start_row+4, column=start_col+2, value=rec['parity_bits_eve'])

        ws.cell(row=start_row+5, column=start_col, value="Total Bit (A/B | E-A/E-B)")
        ws.cell(row=start_row+5, column=start_col+1, value=f"{rec['total_bits_alice']}/{rec['total_bits_bob']}")
        ws.cell(row=start_row+5, column=start_col+2, value=f"{rec['total_bits_ea']}/{rec['total_bits_eb']}")

        ws.cell(row=start_row+6, column=start_col, value="Error Bit Sebelum")
        ws.cell(row=start_row+6, column=start_col+1, value=rec['error_bits_ab_before'])
        ws.cell(row=start_row+6, column=start_col+2, value=rec['error_bits_eve_before'])

        ws.cell(row=start_row+7, column=start_col, value="Error Bit Setelah")
        ws.cell(row=start_row+7, column=start_col+1, value=rec['error_bits_ab_after'])
        ws.cell(row=start_row+7, column=start_col+2, value=rec['error_bits_eve_after'])

        ws.cell(row=start_row+8, column=start_col, value="Bit Terkoreksi/Disamakan")
        ws.cell(row=start_row+8, column=start_col+1, value=rec['corrected_bits_ab'])
        ws.cell(row=start_row+8, column=start_col+2, value=rec['corrected_bits_eve'])

        ws.cell(row=start_row+9, column=start_col, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+9, column=start_col+1, value=rec['time_bch_ab'])
        ws.cell(row=start_row+9, column=start_col+2, value=rec['time_bch_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+9, min_col=start_col, max_col=start_col+2):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 5, 9]):
            write_bch_table(start_row, start_col, rec, label)

        current_row = start_row + 12

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 25

def build_hash_sheet(wb, records):
    ws = wb.create_sheet(title="Hash_SHA_AES")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_hash_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+4)

        ws.cell(row=start_row, column=start_col, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row+1, column=start_col, value="Jumlah Kunci Cocok (Match)")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['aes_count_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+1, end_row=start_row+1, end_column=start_col+2)
        ws.cell(row=start_row+1, column=start_col+3, value=rec['aes_count_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=start_col+3, end_row=start_row+1, end_column=start_col+4)

        ws.cell(row=start_row+2, column=start_col, value="Jumlah Kandidat Key")
        ws.cell(row=start_row+2, column=start_col+1, value=rec['keys_count_alice'])
        ws.cell(row=start_row+2, column=start_col+2, value=rec['keys_count_bob'])
        ws.cell(row=start_row+2, column=start_col+3, value=rec['keys_count_ea'])
        ws.cell(row=start_row+2, column=start_col+4, value=rec['keys_count_eb'])

        ws.cell(row=start_row+3, column=start_col, value="Total Bit Key (128*jumlah_key)")
        ws.cell(row=start_row+3, column=start_col+1, value=rec['total_key_bits_alice'])
        ws.cell(row=start_row+3, column=start_col+2, value=rec['total_key_bits_bob'])
        ws.cell(row=start_row+3, column=start_col+3, value=rec['total_key_bits_ea'])
        ws.cell(row=start_row+3, column=start_col+4, value=rec['total_key_bits_eb'])

        ws.cell(row=start_row+4, column=start_col, value="Total Bit AES Match")
        ws.cell(row=start_row+4, column=start_col+1, value=rec['matched_key_bits_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=start_col+1, end_row=start_row+4, end_column=start_col+2)
        ws.cell(row=start_row+4, column=start_col+3, value=rec['matched_key_bits_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=start_col+3, end_row=start_row+4, end_column=start_col+4)

        ws.cell(row=start_row+5, column=start_col, value="Kunci Pertama (Hex)")
        ws.cell(row=start_row+5, column=start_col+1, value=rec['final_key_alice'])
        ws.cell(row=start_row+5, column=start_col+2, value=rec['final_key_bob'])
        ws.cell(row=start_row+5, column=start_col+3, value=rec['final_key_ea'])
        ws.cell(row=start_row+5, column=start_col+4, value=rec['final_key_eb'])

        ws.cell(row=start_row+6, column=start_col, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+6, column=start_col+1, value=rec['time_hash_ab'])
        ws.cell(row=start_row+6, column=start_col+2, value=rec['time_hash_ab'])
        ws.cell(row=start_row+6, column=start_col+3, value=rec['time_hash_eve'])
        ws.cell(row=start_row+6, column=start_col+4, value=rec['time_hash_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+6, min_col=start_col, max_col=start_col+4):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 7, 13]):
            write_hash_table(start_row, start_col, rec, label)

        current_row = start_row + 9

    for col in range(1, 18):
        ws.column_dimensions[get_column_letter(col)].width = 32
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['M'].width = 25

def build_nist_sheet(wb, records):
    ws = wb.create_sheet(title="NIST")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def write_nist_table(start_row, start_col, rec, label):
        ws.cell(row=start_row-1, column=start_col, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=start_row-1, start_column=start_col, end_row=start_row-1, end_column=start_col+2)

        ws.cell(row=start_row, column=start_col, value="Parameter NIST").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=start_col+1+idx, value=val).font = header_font

        ws.cell(row=start_row+1, column=start_col, value="Jumlah Key Lulus")
        ws.cell(row=start_row+1, column=start_col+1, value=rec['passed_keys_ab'])
        ws.cell(row=start_row+1, column=start_col+2, value=rec['passed_keys_eve'])

        ws.cell(row=start_row+2, column=start_col, value="Rata-rata p-value (ApEn)")
        ws.cell(row=start_row+2, column=start_col+1, value=rec['pval_ab'])
        ws.cell(row=start_row+2, column=start_col+2, value=rec['pval_eve'])

        ws.cell(row=start_row+3, column=start_col, value="Pass Rate (%)")
        ws.cell(row=start_row+3, column=start_col+1, value=rec['pass_rate_ab'])
        ws.cell(row=start_row+3, column=start_col+2, value=rec['pass_rate_eve'])

        ws.cell(row=start_row+4, column=start_col, value="Distribusi p-value")
        ws.cell(row=start_row+4, column=start_col+1, value=rec['pval_dist_ab'])
        ws.cell(row=start_row+4, column=start_col+2, value=rec['pval_dist_eve'])

        ws.cell(row=start_row+5, column=start_col, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+5, column=start_col+1, value=rec['time_nist_ab'])
        ws.cell(row=start_row+5, column=start_col+2, value=rec['time_nist_eve'])

        for row in ws.iter_rows(min_row=start_row-1, max_row=start_row+5, min_col=start_col, max_col=start_col+2):
            for cell in row:
                cell.alignment = center_align

    current_row = 1
    paired_records = pair_records_by_param(records)
    for (skenario, q, r, bb), by_block in paired_records:
        ws.cell(row=current_row, column=1, value=f"Skenario {skenario} - Q={q}, R={r}, BB={bb}").font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        current_row += 2

        start_row = current_row
        for start_col, rec, label in iter_report_panels(by_block, [1, 5, 9]):
            write_nist_table(start_row, start_col, rec, label)

        current_row = start_row + 8

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 25

def build_endtoend_sheet(wb, records):
    ws = wb.create_sheet(title="EndToEnd")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        "Skenario", "Block Size", "Q", "R", "BB",
        "KGR Kumulatif Akhir AB (bit/s)", "KGR Kumulatif Akhir Eve (bit/s)",
        "KDR Kumulatif Akhir AB (%)", "KDR Kumulatif Akhir Eve (%)",
        "t_total AB (s)", "t_total Eve (s)",
    ]
    for col_idx, val in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=val).font = header_font

    for row_idx, r in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=r['skenario'])
        ws.cell(row=row_idx, column=2, value=r.get('block_size', 'N/A'))
        ws.cell(row=row_idx, column=3, value=r['q'])
        ws.cell(row=row_idx, column=4, value=r['r'])
        ws.cell(row=row_idx, column=5, value=r['bb'])
        ws.cell(row=row_idx, column=6, value=r['kgr_final_ab'])
        ws.cell(row=row_idx, column=7, value=r['kgr_final_eve'])
        ws.cell(row=row_idx, column=8, value=r['kdr_final_ab'])
        ws.cell(row=row_idx, column=9, value=r['kdr_final_eve'])
        ws.cell(row=row_idx, column=10, value=r['t_total_ab'])
        ws.cell(row=row_idx, column=11, value=r['t_total_eve'])

    for row in ws.iter_rows(min_row=1, max_row=max(2, len(records) + 1), min_col=1, max_col=11):
        for cell in row:
            cell.alignment = center_align

    widths = [10, 12, 8, 8, 8, 30, 30, 28, 30, 14, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

# =====================================================================
# MAIN ENTRY POINT
# =====================================================================
def main():
    print("=== FULL SECRET KEY GENERATION (SKG) AUTOMATION ===")
    base_data = "data_mita"
    output_base = "Output_mita_copy"
    
    # === Inisialisasi list untuk rekap global semua skenario ===
    global_kalman_records = []
    global_kuan_records = []
    global_bch_records = []
    global_hash_records = []
    global_nist_records = []
    global_endtoend_records = []
    
    for skenario in SCENARIOS:
        print(f"\n>>>> Memproses Skenario {skenario} <<<<")

        path_alice = os.path.join(base_data, "alice", f"skenario{skenario}_mita_alice.csv")
        path_bob = os.path.join(base_data, "bob", f"skenario{skenario}_mita_bob.csv")
        path_eve_a = os.path.join(base_data, "eve alice", f"skenario{skenario}_mita_evealice.csv")
        path_eve_b = os.path.join(base_data, "eve bob", f"skenario{skenario}_mita_evebob.csv")

        raw_alice = read_rssi_csv(path_alice)
        raw_bob = read_rssi_csv(path_bob)
        raw_eve_a = read_rssi_csv(path_eve_a)
        raw_eve_b = read_rssi_csv(path_eve_b)

        if not (raw_alice and raw_bob and raw_eve_a and raw_eve_b):
            print(f"Melewati skenario {skenario} karena data file tidak lengkap di direktori.")
            continue

        total_len = min(len(raw_alice), len(raw_bob), len(raw_eve_a), len(raw_eve_b))

        skenario_out_dir = os.path.join(output_base, f"skenario_{skenario}")
        os.makedirs(skenario_out_dir, exist_ok=True)

        kalman_records = []
        kuan_records = []
        bch_records = []
        hash_records = []
        nist_records = []

        for block_size in [FULL_MODE_KEY] + BLOCK_SIZE_OPTIONS:
            if block_size == FULL_MODE_KEY:
                synchronized_len = total_len
            else:
                synchronized_len = (total_len // block_size) * block_size

            if synchronized_len == 0:
                print(f"   Lewati pembagian {block_size}: data kurang dari ukuran blok.")
                continue

            raw_alice_sync = raw_alice[:synchronized_len]
            raw_bob_sync = raw_bob[:synchronized_len]
            raw_eve_a_sync = raw_eve_a[:synchronized_len]
            raw_eve_b_sync = raw_eve_b[:synchronized_len]

            if block_size == FULL_MODE_KEY:
                block_count = 1
                dropped = 0
                print(f"   Mode data full: pakai {synchronized_len} sampel (tanpa pembagian blok)")
                block_suffix = "full"
            else:
                block_count = synchronized_len // block_size
                dropped = total_len - synchronized_len
                print(
                    f"   Mode blok {block_size}: pakai {synchronized_len} sampel ({block_count} blok), buang {dropped} sampel"
                )
                block_suffix = str(block_size)

            block_out_dir = os.path.join(skenario_out_dir, f"blok_{block_suffix}")
            excel_kalman_dir = os.path.join(block_out_dir, "data_excel_kalman")
            excel_kuan_dir = os.path.join(block_out_dir, "data_excel_kuantisasi")
            os.makedirs(excel_kalman_dir, exist_ok=True)
            os.makedirs(excel_kuan_dir, exist_ok=True)

            for param in PARAM_VARIATIONS:
                q, r, bb = param['q'], param['r'], param['bb']
                if bb > synchronized_len:
                    print(f" -> Variasi BLK={block_size}, Q={q}, R={r}, BB={bb} dilewati (BB lebih besar dari data).")
                    continue
                if block_size != FULL_MODE_KEY and ((bb > block_size) or ((block_size % bb) != 0)):
                    print(f" -> Variasi BLK={block_size}, Q={q}, R={r}, BB={bb} dilewati (BB tidak membagi ukuran blok).")
                    continue
                print(f" -> Variasi dijalankan: BLK={block_size}, Q={q}, R={r}, BB={bb}")

                ra_full = raw_alice_sync
                rb_full = raw_bob_sync
                rea_full = raw_eve_a_sync
                reb_full = raw_eve_b_sync

                orig_max_alice = np.max(ra_full) if synchronized_len > 0 else 0
                orig_max_bob = np.max(rb_full) if synchronized_len > 0 else 0
                orig_max_evea = np.max(rea_full) if synchronized_len > 0 else 0
                orig_max_eveb = np.max(reb_full) if synchronized_len > 0 else 0

                orig_min_alice = np.min(ra_full) if synchronized_len > 0 else 0
                orig_min_bob = np.min(rb_full) if synchronized_len > 0 else 0
                orig_min_evea = np.min(rea_full) if synchronized_len > 0 else 0
                orig_min_eveb = np.min(reb_full) if synchronized_len > 0 else 0

                orig_corr_ab = calc_corr(ra_full, rb_full)
                orig_corr_eve = calc_corr(rea_full, reb_full)

                if block_size == FULL_MODE_KEY:
                    kal_a, _, t_a = process_kalman(raw_alice_sync, q, r, bb, BENCHMARK_ITERATIONS)
                    kal_b, _, t_b = process_kalman(raw_bob_sync, q, r, bb, BENCHMARK_ITERATIONS)
                    kal_ea, _, t_ea = process_kalman(raw_eve_a_sync, q, r, bb, BENCHMARK_ITERATIONS)
                    kal_eb, _, t_eb = process_kalman(raw_eve_b_sync, q, r, bb, BENCHMARK_ITERATIONS)
                    kal_blocks_a, kal_times_a = ([kal_a] if kal_a else []), ([float(t_a)] if kal_a else [])
                    kal_blocks_b, kal_times_b = ([kal_b] if kal_b else []), ([float(t_b)] if kal_b else [])
                    kal_blocks_ea, kal_times_ea = ([kal_ea] if kal_ea else []), ([float(t_ea)] if kal_ea else [])
                    kal_blocks_eb, kal_times_eb = ([kal_eb] if kal_eb else []), ([float(t_eb)] if kal_eb else [])
                else:
                    kal_blocks_a, kal_times_a = process_kalman_per_blocks(raw_alice_sync, q, r, block_size, bb, BENCHMARK_ITERATIONS)
                    kal_blocks_b, kal_times_b = process_kalman_per_blocks(raw_bob_sync, q, r, block_size, bb, BENCHMARK_ITERATIONS)
                    kal_blocks_ea, kal_times_ea = process_kalman_per_blocks(raw_eve_a_sync, q, r, block_size, bb, BENCHMARK_ITERATIONS)
                    kal_blocks_eb, kal_times_eb = process_kalman_per_blocks(raw_eve_b_sync, q, r, block_size, bb, BENCHMARK_ITERATIONS)

                if not (kal_blocks_a and kal_blocks_b and kal_blocks_ea and kal_blocks_eb):
                    print("    Kalman per-blok gagal menghasilkan data, variasi dilewati.")
                    continue

                time_kal_a = sum(kal_times_a)
                time_kal_b = sum(kal_times_b)
                time_kal_ea = sum(kal_times_ea)
                time_kal_eb = sum(kal_times_eb)

                block_tag = "FULL" if block_size == FULL_MODE_KEY else str(block_size)
                v_name = f"BLK{block_tag}_Q{q}_R{r}_BB{bb}"

                for idx, block in enumerate(kal_blocks_a, start=1):
                    save_data_list(excel_kalman_dir, f"{v_name}_kalman_alice_block{idx}.xlsx", block, "alice_kalman")
                for idx, block in enumerate(kal_blocks_b, start=1):
                    save_data_list(excel_kalman_dir, f"{v_name}_kalman_bob_block{idx}.xlsx", block, "bob_kalman")
                for idx, block in enumerate(kal_blocks_ea, start=1):
                    save_data_list(excel_kalman_dir, f"{v_name}_kalman_evealice_block{idx}.xlsx", block, "evealice_kalman")
                for idx, block in enumerate(kal_blocks_eb, start=1):
                    save_data_list(excel_kalman_dir, f"{v_name}_kalman_evebob_block{idx}.xlsx", block, "evebob_kalman")

                kal_max_alice = average_numeric([np.max(b) if len(b) > 0 else 0 for b in kal_blocks_a])
                kal_max_bob = average_numeric([np.max(b) if len(b) > 0 else 0 for b in kal_blocks_b])
                kal_max_evea = average_numeric([np.max(b) if len(b) > 0 else 0 for b in kal_blocks_ea])
                kal_max_eveb = average_numeric([np.max(b) if len(b) > 0 else 0 for b in kal_blocks_eb])

                kal_min_alice = average_numeric([np.min(b) if len(b) > 0 else 0 for b in kal_blocks_a])
                kal_min_bob = average_numeric([np.min(b) if len(b) > 0 else 0 for b in kal_blocks_b])
                kal_min_evea = average_numeric([np.min(b) if len(b) > 0 else 0 for b in kal_blocks_ea])
                kal_min_eveb = average_numeric([np.min(b) if len(b) > 0 else 0 for b in kal_blocks_eb])

                kal_corr_ab = average_corr([calc_corr(a, b) for a, b in zip(kal_blocks_a, kal_blocks_b)])
                kal_corr_eve = average_corr([calc_corr(a, b) for a, b in zip(kal_blocks_ea, kal_blocks_eb)])

                kalman_records.append({
                    "skenario": skenario,
                    "block_size": block_size,
                    "q": q,
                    "r": r,
                    "bb": bb,
                    "orig_max_alice": orig_max_alice,
                    "orig_max_bob": orig_max_bob,
                    "orig_max_evealice": orig_max_evea,
                    "orig_max_evebob": orig_max_eveb,
                    "orig_min_alice": orig_min_alice,
                    "orig_min_bob": orig_min_bob,
                    "orig_min_evealice": orig_min_evea,
                    "orig_min_evebob": orig_min_eveb,
                    "orig_corr_ab": orig_corr_ab,
                    "orig_corr_eve": orig_corr_eve,
                    "kalman_max_alice": kal_max_alice,
                    "kalman_max_bob": kal_max_bob,
                    "kalman_max_evealice": kal_max_evea,
                    "kalman_max_evebob": kal_max_eveb,
                    "kalman_min_alice": kal_min_alice,
                    "kalman_min_bob": kal_min_bob,
                    "kalman_min_evealice": kal_min_evea,
                    "kalman_min_evebob": kal_min_eveb,
                    "kalman_corr_ab": kal_corr_ab,
                    "kalman_corr_eve": kal_corr_eve,
                    "time_alice": time_kal_a,
                    "time_bob": time_kal_b,
                    "time_evealice": time_kal_ea,
                    "time_evebob": time_kal_eb,
                })

                bs_blocks_a, kgr_blocks_a, bs_a, kgr_merge_a_local, time_kuan_a = process_kuantisasi_per_blocks(kal_blocks_a, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                bs_blocks_b, kgr_blocks_b, bs_b, kgr_merge_b_local, time_kuan_b = process_kuantisasi_per_blocks(kal_blocks_b, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                bs_blocks_ea, kgr_blocks_ea, bs_ea, kgr_merge_ea_local, time_kuan_ea = process_kuantisasi_per_blocks(kal_blocks_ea, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                bs_blocks_eb, kgr_blocks_eb, bs_eb, kgr_merge_eb_local, time_kuan_eb = process_kuantisasi_per_blocks(kal_blocks_eb, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)

                for idx, block_bits in enumerate(bs_blocks_a, start=1):
                    save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_alice_block{idx}.xlsx", [block_bits], "bitstream")
                for idx, block_bits in enumerate(bs_blocks_b, start=1):
                    save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_bob_block{idx}.xlsx", [block_bits], "bitstream")
                for idx, block_bits in enumerate(bs_blocks_ea, start=1):
                    save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evealice_block{idx}.xlsx", [block_bits], "bitstream")
                for idx, block_bits in enumerate(bs_blocks_eb, start=1):
                    save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evebob_block{idx}.xlsx", [block_bits], "bitstream")

                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_alice_merged.xlsx", [bs_a], "bitstream_merged")
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_bob_merged.xlsx", [bs_b], "bitstream_merged")
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evealice_merged.xlsx", [bs_ea], "bitstream_merged")
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evebob_merged.xlsx", [bs_eb], "bitstream_merged")

                kdr_blocks_ab = [calculate_kdr(a, b) for a, b in zip(bs_blocks_a, bs_blocks_b)]
                kdr_blocks_eve = [calculate_kdr(a, b) for a, b in zip(bs_blocks_ea, bs_blocks_eb)]

                block_metrics_df = pd.DataFrame({
                    "block_index": list(range(1, len(kdr_blocks_ab) + 1)),
                    "kdr_ab_block": kdr_blocks_ab,
                    "kdr_eve_block": kdr_blocks_eve,
                    "kgr_local_alice_block": kgr_blocks_a,
                    "kgr_local_bob_block": kgr_blocks_b,
                    "kgr_local_evealice_block": kgr_blocks_ea,
                    "kgr_local_evebob_block": kgr_blocks_eb,
                    "bits_alice_block": [len(v) for v in bs_blocks_a],
                    "bits_bob_block": [len(v) for v in bs_blocks_b],
                })
                block_metrics_df.to_excel(
                    os.path.join(excel_kuan_dir, f"{v_name}_rekap_per_block.xlsx"),
                    index=False,
                )

                kdr_ab = average_numeric(kdr_blocks_ab)
                kdr_eve = average_numeric(kdr_blocks_eve)

                kgr_kuan_a_local = average_numeric(kgr_blocks_a)
                kgr_kuan_b_local = average_numeric(kgr_blocks_b)
                kgr_kuan_ea_local = average_numeric(kgr_blocks_ea)
                kgr_kuan_eb_local = average_numeric(kgr_blocks_eb)

                kgr_kuan_a = calculate_cumulative_kgr(len(bs_a), CHANPROB_TIME_SECONDS, time_kal_a, time_kuan_a)
                kgr_kuan_b = calculate_cumulative_kgr(len(bs_b), CHANPROB_TIME_SECONDS, time_kal_b, time_kuan_b)
                kgr_kuan_ea = calculate_cumulative_kgr(len(bs_ea), CHANPROB_TIME_SECONDS, time_kal_ea, time_kuan_ea)
                kgr_kuan_eb = calculate_cumulative_kgr(len(bs_eb), CHANPROB_TIME_SECONDS, time_kal_eb, time_kuan_eb)

                compared_bits_ab = min(len(bs_a), len(bs_b))
                compared_bits_eve = min(len(bs_ea), len(bs_eb))

                kdr_merge_ab = calculate_kdr(bs_a, bs_b)
                kdr_merge_eve = calculate_kdr(bs_ea, bs_eb)

                avg_time_kuan_ab = average_numeric([time_kuan_a, time_kuan_b])
                avg_time_kuan_eve = average_numeric([time_kuan_ea, time_kuan_eb])
                avg_time_kal_ab = average_numeric([time_kal_a, time_kal_b])
                avg_time_kal_eve = average_numeric([time_kal_ea, time_kal_eb])

                kgr_merge_local_ab = calculate_local_kgr(compared_bits_ab, avg_time_kuan_ab)
                kgr_merge_local_eve = calculate_local_kgr(compared_bits_eve, avg_time_kuan_eve)

                kgr_merge_cum_ab = calculate_cumulative_kgr(compared_bits_ab, CHANPROB_TIME_SECONDS, avg_time_kal_ab, avg_time_kuan_ab)
                kgr_merge_cum_eve = calculate_cumulative_kgr(compared_bits_eve, CHANPROB_TIME_SECONDS, avg_time_kal_eve, avg_time_kuan_eve)

                kuan_records.append({
                    "skenario": skenario,
                    "block_size": block_size,
                    "q": q,
                    "r": r,
                    "bb": bb,
                    "kdr_ab": kdr_ab,
                    "kdr_eve": kdr_eve,
                    "time_alice": time_kuan_a,
                    "time_bob": time_kuan_b,
                    "time_evealice": time_kuan_ea,
                    "time_evebob": time_kuan_eb,
                    "kgr_local_alice": kgr_kuan_a_local,
                    "kgr_local_bob": kgr_kuan_b_local,
                    "kgr_local_evealice": kgr_kuan_ea_local,
                    "kgr_local_evebob": kgr_kuan_eb_local,
                    "kgr_cum_alice": kgr_kuan_a,
                    "kgr_cum_bob": kgr_kuan_b,
                    "kgr_cum_evealice": kgr_kuan_ea,
                    "kgr_cum_evebob": kgr_kuan_eb,
                    "total_bits_alice": len(bs_a),
                    "total_bits_bob": len(bs_b),
                    "total_bits_ea": len(bs_ea),
                    "total_bits_eb": len(bs_eb),
                    "kdr_merge_ab": kdr_merge_ab,
                    "kdr_merge_eve": kdr_merge_eve,
                    "kgr_merge_local_ab": kgr_merge_local_ab,
                    "kgr_merge_local_eve": kgr_merge_local_eve,
                    "kgr_merge_cum_ab": kgr_merge_cum_ab,
                    "kgr_merge_cum_eve": kgr_merge_cum_eve,
                    "total_bits_merge_ab": compared_bits_ab,
                    "total_bits_merge_eve": compared_bits_eve,
                })

                try:
                    from bch_module import process_bch

                    b_alice, b_bob, stats_ab = process_bch(bs_a, bs_b, apply_correction=True)
                    b_ea, b_eb, stats_eve = process_bch(bs_ea, bs_eb, apply_correction=False)

                    kgr_bch_ab_local = calculate_local_kgr(len(b_alice), stats_ab["time_bch"])
                    kgr_bch_eve_local = calculate_local_kgr(len(b_ea), stats_eve["time_bch"])

                    kgr_bch_ab = calculate_cumulative_kgr(
                        len(b_alice),
                        CHANPROB_TIME_SECONDS,
                        time_kal_a,
                        time_kuan_a,
                        stats_ab["time_bch"],
                    )
                    kgr_bch_eve = calculate_cumulative_kgr(
                        len(b_ea),
                        CHANPROB_TIME_SECONDS,
                        time_kal_ea,
                        time_kuan_ea,
                        stats_eve["time_bch"],
                    )

                    bch_dir = os.path.join(block_out_dir, "data_excel_bch")
                    os.makedirs(bch_dir, exist_ok=True)
                    save_data_list(bch_dir, f"{v_name}_bch_alice.xlsx", b_alice, "alice_bch_bits")
                    save_data_list(bch_dir, f"{v_name}_bch_bob.xlsx", b_bob, "bob_bch_bits")

                    bch_records.append({
                        "skenario": skenario,
                        "block_size": block_size,
                        "q": q,
                        "r": r,
                        "bb": bb,
                        "kdr_after_ab": stats_ab["kdr_after"],
                        "kdr_after_eve": stats_eve["kdr_after"],
                        "kgr_bch_ab_local": kgr_bch_ab_local,
                        "kgr_bch_eve_local": kgr_bch_eve_local,
                        "kgr_bch_ab": kgr_bch_ab,
                        "kgr_bch_eve": kgr_bch_eve,
                        "parity_bits_ab": stats_ab["parity_bits_sent"],
                        "parity_bits_eve": stats_eve["parity_bits_sent"],
                        "total_bits_alice": stats_ab["total_bits_alice"],
                        "total_bits_bob": stats_ab["total_bits_bob"],
                        "total_bits_ea": stats_eve["total_bits_alice"],
                        "total_bits_eb": stats_eve["total_bits_bob"],
                        "error_bits_ab_before": stats_ab["error_bits_before"],
                        "error_bits_eve_before": stats_eve["error_bits_before"],
                        "error_bits_ab_after": stats_ab["error_bits_after"],
                        "error_bits_eve_after": stats_eve["error_bits_after"],
                        "corrected_bits_ab": stats_ab["corrected_bits"],
                        "corrected_bits_eve": stats_eve["corrected_bits"],
                        "time_bch_ab": stats_ab["time_bch"],
                        "time_bch_eve": stats_eve["time_bch"],
                    })

                    try:
                        from hash_module import process_hash

                        h_alice, h_bob, aes_ab, time_hash_ab, hash_metrics_ab = process_hash(b_alice, b_bob)
                        h_ea, h_eb, aes_eve, time_hash_eve, hash_metrics_eve = process_hash(b_ea, b_eb)

                        hash_dir = os.path.join(block_out_dir, "data_excel_hash")
                        os.makedirs(hash_dir, exist_ok=True)
                        save_data_list(hash_dir, f"{v_name}_hash_alice.xlsx", h_alice, "AES_keys")
                        save_data_list(hash_dir, f"{v_name}_hash_bob.xlsx", h_bob, "AES_keys")
                        save_data_list(hash_dir, f"{v_name}_hash_evealice.xlsx", h_ea, "AES_keys")
                        save_data_list(hash_dir, f"{v_name}_hash_evebob.xlsx", h_eb, "AES_keys")

                        hash_records.append({
                            "skenario": skenario,
                            "block_size": block_size,
                            "q": q,
                            "r": r,
                            "bb": bb,
                            "aes_count_ab": len(aes_ab),
                            "aes_count_eve": len(aes_eve),
                            "keys_count_alice": hash_metrics_ab["keys_count_alice"],
                            "keys_count_bob": hash_metrics_ab["keys_count_bob"],
                            "keys_count_ea": hash_metrics_eve["keys_count_alice"],
                            "keys_count_eb": hash_metrics_eve["keys_count_bob"],
                            "total_key_bits_alice": hash_metrics_ab["total_key_bits_alice"],
                            "total_key_bits_bob": hash_metrics_ab["total_key_bits_bob"],
                            "total_key_bits_ea": hash_metrics_eve["total_key_bits_alice"],
                            "total_key_bits_eb": hash_metrics_eve["total_key_bits_bob"],
                            "matched_key_bits_ab": hash_metrics_ab["matched_key_bits"],
                            "matched_key_bits_eve": hash_metrics_eve["matched_key_bits"],
                            "final_key_alice": h_alice[0] if len(h_alice) > 0 else "N/A",
                            "final_key_bob": h_bob[0] if len(h_bob) > 0 else "N/A",
                            "final_key_ea": h_ea[0] if len(h_ea) > 0 else "N/A",
                            "final_key_eb": h_eb[0] if len(h_eb) > 0 else "N/A",
                            "time_hash_ab": time_hash_ab,
                            "time_hash_eve": time_hash_eve,
                        })

                        total_compared_bits_ab = min(
                            hash_metrics_ab["total_key_bits_alice"],
                            hash_metrics_ab["total_key_bits_bob"],
                        )
                        total_compared_bits_eve = min(
                            hash_metrics_eve["total_key_bits_alice"],
                            hash_metrics_eve["total_key_bits_bob"],
                        )

                        t_total_ab = (
                            CHANPROB_TIME_SECONDS
                            + float(time_kal_a)
                            + float(time_kuan_a)
                            + float(stats_ab["time_bch"])
                            + float(time_hash_ab)
                        )
                        t_total_eve = (
                            CHANPROB_TIME_SECONDS
                            + float(time_kal_ea)
                            + float(time_kuan_ea)
                            + float(stats_eve["time_bch"])
                            + float(time_hash_eve)
                        )

                        kgr_final_ab = calculate_cumulative_kgr(
                            hash_metrics_ab["matched_key_bits"],
                            CHANPROB_TIME_SECONDS,
                            time_kal_a,
                            time_kuan_a,
                            stats_ab["time_bch"],
                            time_hash_ab,
                        )
                        kgr_final_eve = calculate_cumulative_kgr(
                            hash_metrics_eve["matched_key_bits"],
                            CHANPROB_TIME_SECONDS,
                            time_kal_ea,
                            time_kuan_ea,
                            stats_eve["time_bch"],
                            time_hash_eve,
                        )

                        kdr_final_ab = calculate_kdr_from_matched_bits(
                            hash_metrics_ab["matched_key_bits"],
                            total_compared_bits_ab,
                        )
                        kdr_final_eve = calculate_kdr_from_matched_bits(
                            hash_metrics_eve["matched_key_bits"],
                            total_compared_bits_eve,
                        )

                        global_endtoend_records.append({
                            "skenario": skenario,
                            "block_size": block_size,
                            "q": q,
                            "r": r,
                            "bb": bb,
                            "kgr_final_ab": kgr_final_ab,
                            "kgr_final_eve": kgr_final_eve,
                            "kdr_final_ab": kdr_final_ab,
                            "kdr_final_eve": kdr_final_eve,
                            "t_total_ab": t_total_ab,
                            "t_total_eve": t_total_eve,
                        })

                        try:
                            from nist_module import process_nist

                            pass_ab, pval_ab, pass_rate_ab, pdist_ab, time_nist_ab = process_nist(aes_ab)
                            pass_eve, pval_eve, pass_rate_eve, pdist_eve, time_nist_eve = process_nist(aes_eve)

                            pdist_ab_str = ", ".join([f"{k}:{v}" for k, v in pdist_ab.items()])
                            pdist_eve_str = ", ".join([f"{k}:{v}" for k, v in pdist_eve.items()])

                            nist_records.append({
                                "skenario": skenario,
                                "block_size": block_size,
                                "q": q,
                                "r": r,
                                "bb": bb,
                                "passed_keys_ab": pass_ab,
                                "passed_keys_eve": pass_eve,
                                "pval_ab": pval_ab,
                                "pval_eve": pval_eve,
                                "pass_rate_ab": pass_rate_ab,
                                "pass_rate_eve": pass_rate_eve,
                                "pval_dist_ab": pdist_ab_str,
                                "pval_dist_eve": pdist_eve_str,
                                "time_nist_ab": time_nist_ab,
                                "time_nist_eve": time_nist_eve,
                            })
                        except Exception as e:
                            print("NIST Modul Error:", e)

                    except Exception as e:
                        print("Hash Modul Error:", e)
                except ImportError:
                    print("Modul BCH Belum ditaruh di root folder!")

            print(f"Selesai diproses untuk Skenario {skenario} mode blok {block_size}")

        print(" == Menyusun File Table Laporan ==")
        build_kalman_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Kalman.xlsx"), kalman_records)
        build_kuantisasi_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Kuantisasi.xlsx"), kuan_records)

        if bch_records:
            build_bch_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_BCH.xlsx"), bch_records)
        if hash_records:
            build_hash_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Hash.xlsx"), hash_records)
        if nist_records:
            build_nist_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_NIST.xlsx"), nist_records)

        global_kalman_records.extend(kalman_records)
        global_kuan_records.extend(kuan_records)
        if bch_records:
            global_bch_records.extend(bch_records)
        if hash_records:
            global_hash_records.extend(hash_records)
        if nist_records:
            global_nist_records.extend(nist_records)

    # === GENERATE GLOBAL REKAP ===
    print("\n==== MERANGKUM SELURUH SKENARIO KE DALAM SATU FILE EXCEL ====")
    rekap_excel_path = os.path.join(output_base, "Rekap_Evaluasi_SKG_Semua_Skenario.xlsx")
    
    # Buat workbook kosong
    rekap_wb = Workbook()
    
    # Hapus sheet default bawaan ('Sheet')
    if 'Sheet' in rekap_wb.sheetnames:
        rekap_wb.remove(rekap_wb['Sheet'])
        
    print("Menyusun sheet Kalman...")
    build_kalman_sheet(rekap_wb, global_kalman_records)
    
    print("Menyusun sheet Kuantisasi...")
    build_kuantisasi_sheet(rekap_wb, global_kuan_records)
    
    if global_bch_records:
        print("Menyusun sheet BCH...")
        build_bch_sheet(rekap_wb, global_bch_records)
        
    if global_hash_records:
        print("Menyusun sheet Hash...")
        build_hash_sheet(rekap_wb, global_hash_records)
        
    if global_nist_records:
        print("Menyusun sheet NIST...")
        build_nist_sheet(rekap_wb, global_nist_records)

    if global_endtoend_records:
        print("Menyusun sheet EndToEnd...")
        build_endtoend_sheet(rekap_wb, global_endtoend_records)
        
    saved_path = save_workbook_safely(rekap_wb, rekap_excel_path)
    print(f"Selesai! File rekap global berhasil disimpan di: {saved_path}")

if __name__ == "__main__":
    main()
