import os
import csv
import numpy as np
import pandas as pd
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# === Impor dari Modul Terpisah ===
NIST_TEST_ROWS = [
    ("Approximate Entropy", "Approximate Entropy"),
    ("Frequency", "Frequency"),
    ("Block Frequency", "Block Frequency"),
    ("Cumulative Sums (Forward)", "Cumulative Sums (Forward)"),
    ("Cumulative Sums (Reverse)", "Cumulative Sums (Reverse)"),
    ("Runs", "Runs"),
    ("Longest Runs of Ones", "Longest Runs of Ones"),
]
from kalman_module import process_kalman
from kuantisasi_module import process_kuantisasi
try:
    from bch_module import process_bch
    from hash_module import process_hash
    from nist_module import process_nist
except ImportError:
    pass  # Akan dibuat nanti

# =====================================================================
# GLOBAL PARAMETERS
# =====================================================================
KUANTISASI_NUM_BITS = 3
BENCHMARK_ITERATIONS = 10
CHANPROB_TIME_SECONDS = 120.0  # Hardcoded sementara sesuai arahan: waktu channel probing

# Variasi Parameter Pengujian Skenario
PARAM_VARIATIONS = [
    {"q": 0.01, "r": 0.5, "bb": 1},
    {"q": 0.01, "r": 0.5, "bb": 2},
    {"q": 0.01, "r": 0.5, "bb": 5},
    {"q": 0.01, "r": 0.5, "bb": 10},
    {"q": 0.5, "r": 0.01, "bb": 1},
    {"q": 0.5, "r": 0.01, "bb": 2},
    {"q": 0.5, "r": 0.01, "bb": 5},
    {"q": 0.5, "r": 0.01, "bb": 10},
]

# Skenario iterasi 1 - 4
SCENARIOS = [1, 2, 3, 4]

# =====================================================================
# UTILITY FUNCTIONS
# =====================================================================
def read_rssi_csv(path):
    data = []
    if not os.path.exists(path):
        print(f"Warning: File {path} tidak ditemukan!")
        return data
        
    with open(path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                data.append(int(row[0]))
            except:
                continue
    return data

def calculate_kdr(a, b):
    if not a or not b: return 0.0
    n = min(len(a), len(b))
    if n == 0: return 0.0
    diff = sum(1 for i in range(n) if a[i] != b[i])
    return (diff / n) * 100.0

def calc_corr(a, b):
    ln = min(len(a), len(b))
    if ln < 2: return "N/A"
    c, _ = pearsonr(a[:ln], b[:ln])
    return c

def calculate_cumulative_kgr(total_bits, *time_parts):
    """Hitung KGR kumulatif: total_bits / total_waktu_akumulasi."""
    total_time = 0.0
    for t in time_parts:
        try:
            total_time += float(t)
        except (TypeError, ValueError):
            continue
    if total_time <= 0:
        return 0.0
    return float(total_bits) / total_time

def calculate_local_kgr(total_bits, elapsed):
    """Hitung KGR lokal: total_bits / waktu pada stage tersebut."""
    try:
        elapsed = float(elapsed)
    except (TypeError, ValueError):
        return 0.0
    if elapsed <= 0:
        return 0.0
    return float(total_bits) / elapsed

def calculate_kdr_from_matched_bits(matched_bits, compared_bits):
    """Hitung KDR final (%) dari mismatch bits terhadap total compared bits."""
    try:
        matched_bits = float(matched_bits)
        compared_bits = float(compared_bits)
    except (TypeError, ValueError):
        return 0.0
    if compared_bits <= 0:
        return 0.0
    mismatch = max(0.0, compared_bits - matched_bits)
    return (mismatch / compared_bits) * 100.0

def split_full_blocks(data, block_size):
    """Potong data menjadi blok utuh; sisa data di akhir dibuang."""
    if block_size <= 0:
        return [], 0

    total_full = (len(data) // block_size) * block_size
    if total_full == 0:
        return [], 0

    trimmed = data[:total_full]
    blocks = [trimmed[i:i + block_size] for i in range(0, total_full, block_size)]
    return blocks, total_full

def save_block_series(output_dir, filename, blocks, header):
    """Simpan list per blok sebagai satu kolom Excel."""
    formatted_blocks = []
    for block in blocks:
        if isinstance(block, str):
            formatted_blocks.append(block)
        else:
            formatted_blocks.append(",".join(str(item) for item in block))
    save_data_list(output_dir, filename, formatted_blocks, header)

def process_kalman_per_blocks(raw_data, q, r, block_size, benchmark_iterations=10):
    """Jalankan Kalman per blok dan kembalikan hasil blok + hasil gabungan."""
    blocks, _ = split_full_blocks(raw_data, block_size)
    if not blocks:
        return [], [], [], 0.0, 0.0

    kalman_blocks = []
    kalman_times = []
    for block in blocks:
        kal_block, _, time_block = process_kalman(block, q, r, block_size, benchmark_iterations=benchmark_iterations)
        kalman_blocks.append(kal_block)
        kalman_times.append(float(time_block))

    merged_kalman = [value for block in kalman_blocks for value in block]
    total_time = float(sum(kalman_times))
    kgr_local = calculate_local_kgr(len(merged_kalman), total_time)
    return kalman_blocks, merged_kalman, kalman_times, kgr_local, total_time

def process_kuantisasi_per_blocks(kalman_blocks, num_bits=3, benchmark_iterations=10):
    """Jalankan kuantisasi per blok hasil Kalman lalu gabungkan bitstream."""
    if not kalman_blocks:
        return [], [], [], 0.0, 0.0

    merged_source = [value for block in kalman_blocks for value in block]
    if merged_source:
        shared_min = float(np.min(merged_source))
        shared_max = float(np.max(merged_source))
    else:
        shared_min = None
        shared_max = None

    bitstream_blocks = []
    block_times = []
    block_kgr_locals = []

    for block in kalman_blocks:
        bs_block, _, time_block = process_kuantisasi(
            block,
            num_bits,
            benchmark_iterations,
            reference_min=shared_min,
            reference_max=shared_max,
        )
        bitstream_blocks.append(bs_block)
        block_times.append(float(time_block))
        block_kgr_locals.append(calculate_local_kgr(len(bs_block), time_block))

    merged_bitstream = "".join(bitstream_blocks)
    total_time = float(sum(block_times))
    kgr_local = calculate_local_kgr(len(merged_bitstream), total_time)
    return bitstream_blocks, merged_bitstream, block_kgr_locals, kgr_local, total_time
# =====================================================================
# EXCEL FORMATTING FUNCTIONS
# =====================================================================
def save_data_list(output_dir, filename, data_list, header):
    os.makedirs(output_dir, exist_ok=True)
    df = pd.DataFrame({header: data_list})
    df.to_excel(os.path.join(output_dir, filename), index=False)

def build_kalman_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kalman"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    
    for r in records:
        # Title Data Block
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2
        
        start_row = current_row
        # Headers Top 
        ws.cell(row=start_row, column=1, value="Parameter").font = header_font
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
        
        ws.cell(row=start_row, column=2, value="Sebelum Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=5)
        
        ws.cell(row=start_row, column=6, value="Setelah Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=9)
        
        # Headers Low
        cols_names = ["Alice", "Bob", "Eve-Alice", "Eve-Bob", "Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, cname in enumerate(cols_names):
            ws.cell(row=start_row+1, column=2+idx, value=cname).font = header_font
            
        # Data Maksimum
        ws.cell(row=start_row+2, column=1, value="Maksimum (dBm)")
        vals_max = [r['orig_max_alice'], r['orig_max_bob'], r['orig_max_evealice'], r['orig_max_evebob'],
                    r['kalman_max_alice'], r['kalman_max_bob'], r['kalman_max_evealice'], r['kalman_max_evebob']]
        for idx, val in enumerate(vals_max): ws.cell(row=start_row+2, column=2+idx, value=val)
            
        # Data Minimum
        ws.cell(row=start_row+3, column=1, value="Minimum (dBm)")
        vals_min = [r['orig_min_alice'], r['orig_min_bob'], r['orig_min_evealice'], r['orig_min_evebob'],
                    r['kalman_min_alice'], r['kalman_min_bob'], r['kalman_min_evealice'], r['kalman_min_evebob']]
        for idx, val in enumerate(vals_min): ws.cell(row=start_row+3, column=2+idx, value=val)
            
        # Korelasi
        ws.cell(row=start_row+4, column=1, value="Koefisien Korelasi")
        
        # Sebelum Korelasi A&B + E&E (Merge 2 cell)
        c1 = ws.cell(row=start_row+4, column=2, value=r['orig_corr_ab'])
        c1.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=2, end_row=start_row+4, end_column=3)
        c2 = ws.cell(row=start_row+4, column=4, value=r['orig_corr_eve'])
        c2.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=4, end_row=start_row+4, end_column=5)
        
        # Sesudah Korelasi A&B + E&E
        c3 = ws.cell(row=start_row+4, column=6, value=r['kalman_corr_ab'])
        c3.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=6, end_row=start_row+4, end_column=7)
        c4 = ws.cell(row=start_row+4, column=8, value=r['kalman_corr_eve'])
        c4.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=8, end_row=start_row+4, end_column=9)
        
        # Waktu Komputasi
        ws.cell(row=start_row+5, column=1, value="Waktu Komputasi (s)")
        ws.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=5)
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]):
            ws.cell(row=start_row+5, column=6+idx, value=val)

        # Style alignment applying for all cells in block
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+5, min_col=1, max_col=9):
            for cell in row: cell.alignment = center_align

        # Additional 3 spaces for the next table
        current_row = start_row + 9 
        
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 17
    ws.column_dimensions['A'].width = 25
        
    wb.save(output_path)

def build_kuantisasi_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kuantisasi"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    
    for r in records:
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2
        
        start_row = current_row
        
        # Headers
        ws.cell(row=start_row, column=1, value="Parameter Performansi").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        # KDR
        ws.cell(row=start_row+1, column=1, value="KDR Merge (%)")
        ws.cell(row=start_row+1, column=2, value=r['kdr_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+1, end_column=3)
        ws.cell(row=start_row+1, column=4, value=r['kdr_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=4, end_row=start_row+1, end_column=5)

        ws.cell(row=start_row+2, column=1, value="KDR Rata-rata Per Blok (%)")
        ws.cell(row=start_row+2, column=2, value=r['kdr_block_ab'])
        ws.merge_cells(start_row=start_row+2, start_column=2, end_row=start_row+2, end_column=3)
        ws.cell(row=start_row+2, column=4, value=r['kdr_block_eve'])
        ws.merge_cells(start_row=start_row+2, start_column=4, end_row=start_row+2, end_column=5)
        
        # KGR Kumulatif
        ws.cell(row=start_row+3, column=1, value="KGR Kumulatif Per Blok (bit/s)")
        for idx, val in enumerate([r['kgr_cum_alice'], r['kgr_cum_bob'], r['kgr_cum_evealice'], r['kgr_cum_evebob']]):
            ws.cell(row=start_row+3, column=2+idx, value=val)

        ws.cell(row=start_row+4, column=1, value="KGR Kumulatif Merge (bit/s)")
        ws.cell(row=start_row+4, column=2, value=r['kgr_merge_cum_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=2, end_row=start_row+4, end_column=3)
        ws.cell(row=start_row+4, column=4, value=r['kgr_merge_cum_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=4, end_row=start_row+4, end_column=5)

        # Total bit yang dihasilkan
        ws.cell(row=start_row+5, column=1, value="Total Bit Dihasilkan")
        for idx, val in enumerate([r['total_bits_alice'], r['total_bits_bob'], r['total_bits_ea'], r['total_bits_eb']]):
            ws.cell(row=start_row+5, column=2+idx, value=val)
            
        # Waktu Komputasi
        ws.cell(row=start_row+6, column=1, value="Waktu komputasi (s)")
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]):
            ws.cell(row=start_row+6, column=2+idx, value=val)
            
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+6, min_col=1, max_col=5):
            for cell in row: cell.alignment = center_align

        current_row = start_row + 10
        
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
        
    wb.save(output_path)

def build_bch_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap BCH"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row
        
        ws.cell(row=start_row, column=1, value="Parameter BCH").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        ws.cell(row=start_row+1, column=1, value="KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row+1, column=2, value=r['kdr_after_ab'])
        ws.cell(row=start_row+1, column=3, value=r['kdr_after_eve'])
        
        ws.cell(row=start_row+2, column=1, value="KGR BCH Kumulatif (bit/s)")
        ws.cell(row=start_row+2, column=2, value=r['kgr_bch_ab'])
        ws.cell(row=start_row+2, column=3, value=r['kgr_bch_eve'])

        ws.cell(row=start_row+3, column=1, value="Parity Bits Dikirim")
        ws.cell(row=start_row+3, column=2, value=r['parity_bits_ab'])
        ws.cell(row=start_row+3, column=3, value=r['parity_bits_eve'])

        ws.cell(row=start_row+4, column=1, value="Total Bit (A/B | E-A/E-B)")
        ws.cell(row=start_row+4, column=2, value=f"{r['total_bits_alice']}/{r['total_bits_bob']}")
        ws.cell(row=start_row+4, column=3, value=f"{r['total_bits_ea']}/{r['total_bits_eb']}")

        ws.cell(row=start_row+5, column=1, value="Error Bit Sebelum")
        ws.cell(row=start_row+5, column=2, value=r['error_bits_ab_before'])
        ws.cell(row=start_row+5, column=3, value=r['error_bits_eve_before'])

        ws.cell(row=start_row+6, column=1, value="Error Bit Setelah")
        ws.cell(row=start_row+6, column=2, value=r['error_bits_ab_after'])
        ws.cell(row=start_row+6, column=3, value=r['error_bits_eve_after'])

        ws.cell(row=start_row+7, column=1, value="Bit Terkoreksi/Disamakan")
        ws.cell(row=start_row+7, column=2, value=r['corrected_bits_ab'])
        ws.cell(row=start_row+7, column=3, value=r['corrected_bits_eve'])

        ws.cell(row=start_row+8, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+8, column=2, value=r['time_bch_ab'])
        ws.cell(row=start_row+8, column=3, value=r['time_bch_eve'])
        
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+8, min_col=1, max_col=3):
            for cell in row: cell.alignment = center_align
            
        current_row = start_row + 11
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 20
    wb.save(output_path)
def build_hash_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Hash"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row
        
        ws.cell(row=start_row, column=1, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        ws.cell(row=start_row+1, column=1, value="Jumlah Kunci Cocok (Match)")
        ws.cell(row=start_row+1, column=2, value=r['aes_count_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+1, end_column=3)
        ws.cell(row=start_row+1, column=4, value=r['aes_count_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=4, end_row=start_row+1, end_column=5)

        ws.cell(row=start_row+2, column=1, value="Jumlah Kandidat Key")
        ws.cell(row=start_row+2, column=2, value=r['keys_count_alice'])
        ws.cell(row=start_row+2, column=3, value=r['keys_count_bob'])
        ws.cell(row=start_row+2, column=4, value=r['keys_count_ea'])
        ws.cell(row=start_row+2, column=5, value=r['keys_count_eb'])

        ws.cell(row=start_row+3, column=1, value="Total Bit Key (128*jumlah_key)")
        ws.cell(row=start_row+3, column=2, value=r['total_key_bits_alice'])
        ws.cell(row=start_row+3, column=3, value=r['total_key_bits_bob'])
        ws.cell(row=start_row+3, column=4, value=r['total_key_bits_ea'])
        ws.cell(row=start_row+4, column=2, value=r['matched_key_bits_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=2, end_row=start_row+4, end_column=3)
        ws.cell(row=start_row+4, column=4, value=r['matched_key_bits_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=4, end_row=start_row+4, end_column=5)
        
        ws.cell(row=start_row+5, column=1, value="Kunci Pertama (Hex)")
        ws.cell(row=start_row+5, column=2, value=r['final_key_alice'])
        ws.cell(row=start_row+5, column=3, value=r['final_key_bob'])
        ws.cell(row=start_row+5, column=4, value=r['final_key_ea'])
        ws.cell(row=start_row+5, column=5, value=r['final_key_eb'])
        
        ws.cell(row=start_row+6, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+6, column=2, value=r['time_hash_ab'])
        ws.cell(row=start_row+6, column=3, value=r['time_hash_ab'])
        ws.cell(row=start_row+6, column=4, value=r['time_hash_eve'])
        ws.cell(row=start_row+6, column=5, value=r['time_hash_eve'])
        
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+6, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = center_align
        current_row = start_row + 9
        
    for col in range(1, 6): ws.column_dimensions[get_column_letter(col)].width = 38
    ws.column_dimensions['A'].width = 25
    wb.save(output_path)

def build_hash_detailed_excel(output_path, records):
    """Build Excel dengan 3 tabel side-by-side: Hex Keys | SHA-1 | Matched Keys"""
    from openpyxl.styles import PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Hash Detail"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    current_row = 1
    for r in records:
        # Title
        title = f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}"
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, italic=True)
        current_row += 2
        
        # === TABEL ALICE & BOB (AB) ===
        start_row = current_row
        
        # Header tabel 1: HEX Keys
        ws.cell(row=start_row, column=1, value="HASH (Hex Keys)").font = header_font
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
        
        # Header tabel 2: SHA-1
        ws.cell(row=start_row, column=3, value="SHA-1 Hash").font = header_font
        ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=4)
        
        # Header tabel 3: Matched
        ws.cell(row=start_row, column=5, value="Matched Final Keys").font = header_font
        ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=6)
        
        # Sub-header untuk tabel 1
        ws.cell(row=start_row+1, column=1, value="Alice").font = header_font
        ws.cell(row=start_row+1, column=2, value="Bob").font = header_font
        # Sub-header untuk tabel 2
        ws.cell(row=start_row+1, column=3, value="Alice").font = header_font
        ws.cell(row=start_row+1, column=4, value="Bob").font = header_font
        # Sub-header untuk tabel 3
        ws.cell(row=start_row+1, column=5, value="Index").font = header_font
        ws.cell(row=start_row+1, column=6, value="Hex Key").font = header_font
        
        # Data
        hex_alice_ab = r["hex_alice_ab"]
        hex_bob_ab = r["hex_bob_ab"]
        sha_alice_ab = r["sha_alice_ab"]
        sha_bob_ab = r["sha_bob_ab"]
        matched_idx_ab = r["matched_idx_ab"]
        best_key_ab = r.get("best_key_alice", "N/A")
        
        # Tabel 1 & 2: Semua keys
        max_keys_ab = max(len(hex_alice_ab), len(hex_bob_ab))
        for key_idx in range(max_keys_ab):
            row_num = start_row + 2 + key_idx
            
            # Tabel 1: Hex keys
            if key_idx < len(hex_alice_ab):
                cell_a = ws.cell(row=row_num, column=1, value=hex_alice_ab[key_idx])
                cell_a.font = Font(size=8)
            if key_idx < len(hex_bob_ab):
                cell_b = ws.cell(row=row_num, column=2, value=hex_bob_ab[key_idx])
                cell_b.font = Font(size=8)
            
            # Tabel 2: SHA-1
            if key_idx < len(sha_alice_ab):
                cell_sha_a = ws.cell(row=row_num, column=3, value=sha_alice_ab[key_idx])
                cell_sha_a.font = Font(size=8)
            if key_idx < len(sha_bob_ab):
                cell_sha_b = ws.cell(row=row_num, column=4, value=sha_bob_ab[key_idx])
                cell_sha_b.font = Font(size=8)
        
        # Tabel 3: Only matched keys (dan highlight best key)
        for match_idx, idx in enumerate(matched_idx_ab):
            row_num = start_row + 2 + match_idx
            cell_idx = ws.cell(row=row_num, column=5, value=idx)
            if idx < len(hex_alice_ab):
                key_hex = hex_alice_ab[idx]
                cell_key = ws.cell(row=row_num, column=6, value=key_hex)
                # Highlight best key in yellow
                if key_hex == best_key_ab:
                    cell_idx.fill = fill_yellow
                    cell_key.fill = fill_yellow
                else:
                    cell_idx.fill = fill_green
                    cell_key.fill = fill_green
                cell_idx.font = Font(size=8)
                cell_key.font = Font(size=8)
        
        # Apply alignment
        for row_obj in ws.iter_rows(min_row=start_row, max_row=start_row+1+max(max_keys_ab, len(matched_idx_ab)), min_col=1, max_col=6):
            for cell in row_obj:
                cell.alignment = center_align
        
        current_row = start_row + 2 + max(max_keys_ab, len(matched_idx_ab)) + 3
        
        # === TABEL EVE (Eve-Alice & Eve-Bob) ===
        ws.cell(row=current_row, column=1, value="EVE - HASH & SHA-1 & MATCHED").font = Font(bold=True, italic=True)
        current_row += 2
        start_row_eve = current_row
        
        # Header tabel Eve
        ws.cell(row=start_row_eve, column=1, value="HASH (Hex Keys)").font = header_font
        ws.merge_cells(start_row=start_row_eve, start_column=1, end_row=start_row_eve, end_column=2)
        ws.cell(row=start_row_eve, column=3, value="SHA-1 Hash").font = header_font
        ws.merge_cells(start_row=start_row_eve, start_column=3, end_row=start_row_eve, end_column=4)
        ws.cell(row=start_row_eve, column=5, value="Matched Final Keys").font = header_font
        ws.merge_cells(start_row=start_row_eve, start_column=5, end_row=start_row_eve, end_column=6)
        
        # Sub-header Eve
        ws.cell(row=start_row_eve+1, column=1, value="Eve-Alice").font = header_font
        ws.cell(row=start_row_eve+1, column=2, value="Eve-Bob").font = header_font
        ws.cell(row=start_row_eve+1, column=3, value="Eve-Alice").font = header_font
        ws.cell(row=start_row_eve+1, column=4, value="Eve-Bob").font = header_font
        ws.cell(row=start_row_eve+1, column=5, value="Index").font = header_font
        ws.cell(row=start_row_eve+1, column=6, value="Hex Key").font = header_font
        
        hex_alice_eve = r["hex_alice_eve"]
        hex_bob_eve = r["hex_bob_eve"]
        sha_alice_eve = r["sha_alice_eve"]
        sha_bob_eve = r["sha_bob_eve"]
        matched_idx_eve = r["matched_idx_eve"]
        best_key_eve = r.get("best_key_ea", "N/A")
        
        # Tabel 1 & 2: Semua keys
        max_keys_eve = max(len(hex_alice_eve), len(hex_bob_eve))
        for key_idx in range(max_keys_eve):
            row_num = start_row_eve + 2 + key_idx
            
            if key_idx < len(hex_alice_eve):
                ws.cell(row=row_num, column=1, value=hex_alice_eve[key_idx]).font = Font(size=8)
            if key_idx < len(hex_bob_eve):
                ws.cell(row=row_num, column=2, value=hex_bob_eve[key_idx]).font = Font(size=8)
            if key_idx < len(sha_alice_eve):
                ws.cell(row=row_num, column=3, value=sha_alice_eve[key_idx]).font = Font(size=8)
            if key_idx < len(sha_bob_eve):
                ws.cell(row=row_num, column=4, value=sha_bob_eve[key_idx]).font = Font(size=8)
        
        # Tabel 3: Only matched keys (highlight best key)
        for match_idx, idx in enumerate(matched_idx_eve):
            row_num = start_row_eve + 2 + match_idx
            cell_idx = ws.cell(row=row_num, column=5, value=idx)
            if idx < len(hex_alice_eve):
                key_hex = hex_alice_eve[idx]
                cell_key = ws.cell(row=row_num, column=6, value=key_hex)
                # Highlight best key in yellow
                if key_hex == best_key_eve:
                    cell_idx.fill = fill_yellow
                    cell_key.fill = fill_yellow
                else:
                    cell_idx.fill = fill_green
                    cell_key.fill = fill_green
                cell_idx.font = Font(size=8)
                cell_key.font = Font(size=8)
        
        for row_obj in ws.iter_rows(min_row=start_row_eve, max_row=start_row_eve+1+max(max_keys_eve, len(matched_idx_eve)), min_col=1, max_col=6):
            for cell in row_obj:
                cell.alignment = center_align
        
        current_row = start_row_eve + 2 + max(max_keys_eve, len(matched_idx_eve)) + 3
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 35
    
    wb.save(output_path)

def build_nist_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap NIST"
    _build_nist_sheet_content(ws, records)
    wb.save(output_path)

def build_kalman_sheet(wb, records):
    ws = wb.create_sheet(title="Kalman")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    current_row = 1
    for r in records:
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2
        
        start_row = current_row
        ws.cell(row=start_row, column=1, value="Parameter").font = header_font
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
        
        ws.cell(row=start_row, column=2, value="Sebelum Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=5)
        
        ws.cell(row=start_row, column=6, value="Setelah Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=9)
        
        cols_names = ["Alice", "Bob", "Eve-Alice", "Eve-Bob", "Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, cname in enumerate(cols_names):
            ws.cell(row=start_row+1, column=2+idx, value=cname).font = header_font
            
        ws.cell(row=start_row+2, column=1, value="Maksimum (dBm)")
        vals_max = [r['orig_max_alice'], r['orig_max_bob'], r['orig_max_evealice'], r['orig_max_evebob'], r['kalman_max_alice'], r['kalman_max_bob'], r['kalman_max_evealice'], r['kalman_max_evebob']]
        for idx, val in enumerate(vals_max): ws.cell(row=start_row+2, column=2+idx, value=val)
            
        ws.cell(row=start_row+3, column=1, value="Minimum (dBm)")
        vals_min = [r['orig_min_alice'], r['orig_min_bob'], r['orig_min_evealice'], r['orig_min_evebob'], r['kalman_min_alice'], r['kalman_min_bob'], r['kalman_min_evealice'], r['kalman_min_evebob']]
        for idx, val in enumerate(vals_min): ws.cell(row=start_row+3, column=2+idx, value=val)
            
        ws.cell(row=start_row+4, column=1, value="Koefisien Korelasi")
        c1 = ws.cell(row=start_row+4, column=2, value=r['orig_corr_ab'])
        c1.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=2, end_row=start_row+4, end_column=3)
        c2 = ws.cell(row=start_row+4, column=4, value=r['orig_corr_eve'])
        c2.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=4, end_row=start_row+4, end_column=5)
        c3 = ws.cell(row=start_row+4, column=6, value=r['kalman_corr_ab'])
        c3.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=6, end_row=start_row+4, end_column=7)
        c4 = ws.cell(row=start_row+4, column=8, value=r['kalman_corr_eve'])
        c4.number_format = '0.00000'
        ws.merge_cells(start_row=start_row+4, start_column=8, end_row=start_row+4, end_column=9)
        
        ws.cell(row=start_row+5, column=1, value="Waktu Komputasi (s)")
        ws.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=5)
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]): ws.cell(row=start_row+5, column=6+idx, value=val)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row+5, min_col=1, max_col=9):
            for cell in row: cell.alignment = center_align

        current_row = start_row + 9 
        
    for col in range(1, 10): ws.column_dimensions[get_column_letter(col)].width = 17
    ws.column_dimensions['A'].width = 25

def build_kuantisasi_sheet(wb, records):
    ws = wb.create_sheet(title="Kuantisasi")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2
        
        start_row = current_row
        ws.cell(row=start_row, column=1, value="Parameter Performansi").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols): ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        ws.cell(row=start_row+1, column=1, value="KDR (%)")
        ws.cell(row=start_row+1, column=2, value=r['kdr_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+1, end_column=3)
        ws.cell(row=start_row+1, column=4, value=r['kdr_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=4, end_row=start_row+1, end_column=5)
        
        ws.cell(row=start_row+2, column=1, value="KGR Kumulatif (bit/s)")
        for idx, val in enumerate([r['kgr_cum_alice'], r['kgr_cum_bob'], r['kgr_cum_evealice'], r['kgr_cum_evebob']]): ws.cell(row=start_row+2, column=2+idx, value=val)

        ws.cell(row=start_row+3, column=1, value="Total Bit Dihasilkan")
        for idx, val in enumerate([r['total_bits_alice'], r['total_bits_bob'], r['total_bits_ea'], r['total_bits_eb']]): ws.cell(row=start_row+3, column=2+idx, value=val)
            
        ws.cell(row=start_row+4, column=1, value="Waktu komputasi (s)")
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]): ws.cell(row=start_row+4, column=2+idx, value=val)
            
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+4, min_col=1, max_col=5):
            for cell in row: cell.alignment = center_align

        current_row = start_row + 7
        
    for col in range(1, 6): ws.column_dimensions[get_column_letter(col)].width = 20

def build_bch_sheet(wb, records):
    ws = wb.create_sheet(title="BCH")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row
        
        ws.cell(row=start_row, column=1, value="Parameter BCH").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]): ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        ws.cell(row=start_row+1, column=1, value="KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row+1, column=2, value=r['kdr_after_ab'])
        ws.cell(row=start_row+1, column=3, value=r['kdr_after_eve'])
        
        ws.cell(row=start_row+2, column=1, value="KGR BCH Kumulatif (bit/s)")
        ws.cell(row=start_row+2, column=2, value=r['kgr_bch_ab'])
        ws.cell(row=start_row+2, column=3, value=r['kgr_bch_eve'])

        ws.cell(row=start_row+3, column=1, value="Parity Bits Dikirim")
        ws.cell(row=start_row+3, column=2, value=r['parity_bits_ab'])
        ws.cell(row=start_row+3, column=3, value=r['parity_bits_eve'])

        ws.cell(row=start_row+4, column=1, value="Total Bit (A/B | E-A/E-B)")
        ws.cell(row=start_row+4, column=2, value=f"{r['total_bits_alice']}/{r['total_bits_bob']}")
        ws.cell(row=start_row+4, column=3, value=f"{r['total_bits_ea']}/{r['total_bits_eb']}")

        ws.cell(row=start_row+5, column=1, value="Error Bit Sebelum")
        ws.cell(row=start_row+5, column=2, value=r['error_bits_ab_before'])
        ws.cell(row=start_row+5, column=3, value=r['error_bits_eve_before'])

        ws.cell(row=start_row+6, column=1, value="Error Bit Setelah")
        ws.cell(row=start_row+6, column=2, value=r['error_bits_ab_after'])
        ws.cell(row=start_row+6, column=3, value=r['error_bits_eve_after'])

        ws.cell(row=start_row+7, column=1, value="Bit Terkoreksi/Disamakan")
        ws.cell(row=start_row+7, column=2, value=r['corrected_bits_ab'])
        ws.cell(row=start_row+7, column=3, value=r['corrected_bits_eve'])

        ws.cell(row=start_row+8, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+8, column=2, value=r['time_bch_ab'])
        ws.cell(row=start_row+8, column=3, value=r['time_bch_eve'])
        
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+8, min_col=1, max_col=3):
            for cell in row: cell.alignment = center_align
            
        current_row = start_row + 11
        
    for col in range(1, 4): ws.column_dimensions[get_column_letter(col)].width = 25

def build_hash_sheet(wb, records):
    ws = wb.create_sheet(title="Hash_SHA_AES")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row
        
        ws.cell(row=start_row, column=1, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols): ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        ws.cell(row=start_row+1, column=1, value="Jumlah Kunci Cocok (Match)")
        ws.cell(row=start_row+1, column=2, value=r['aes_count_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+1, end_column=3)
        ws.cell(row=start_row+1, column=4, value=r['aes_count_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=4, end_row=start_row+1, end_column=5)

        ws.cell(row=start_row+2, column=1, value="Jumlah Kandidat Key")
        ws.cell(row=start_row+2, column=2, value=r['keys_count_alice'])
        ws.cell(row=start_row+2, column=3, value=r['keys_count_bob'])
        ws.cell(row=start_row+2, column=4, value=r['keys_count_ea'])
        ws.cell(row=start_row+2, column=5, value=r['keys_count_eb'])

        ws.cell(row=start_row+3, column=1, value="Total Bit Key (128*jumlah_key)")
        ws.cell(row=start_row+3, column=2, value=r['total_key_bits_alice'])
        ws.cell(row=start_row+3, column=3, value=r['total_key_bits_bob'])
        ws.cell(row=start_row+3, column=4, value=r['total_key_bits_ea'])
        ws.cell(row=start_row+3, column=5, value=r['total_key_bits_eb'])

        ws.cell(row=start_row+4, column=1, value="Total Bit AES Match")
        ws.cell(row=start_row+4, column=2, value=r['matched_key_bits_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=2, end_row=start_row+4, end_column=3)
        ws.cell(row=start_row+4, column=4, value=r['matched_key_bits_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=4, end_row=start_row+4, end_column=5)
        
        ws.cell(row=start_row+5, column=1, value="Kunci Terbaik (Best Key by NIST)")
        ws.cell(row=start_row+5, column=2, value=r.get('best_key_alice', 'N/A'))
        ws.cell(row=start_row+5, column=3, value=r.get('best_key_bob', 'N/A'))
        ws.cell(row=start_row+5, column=4, value=r.get('best_key_ea', 'N/A'))
        ws.cell(row=start_row+5, column=5, value=r.get('best_key_eb', 'N/A'))
        
        ws.cell(row=start_row+6, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row+6, column=2, value=r['time_hash_ab'])
        ws.cell(row=start_row+6, column=3, value=r['time_hash_ab'])
        ws.cell(row=start_row+6, column=4, value=r['time_hash_eve'])
        ws.cell(row=start_row+6, column=5, value=r['time_hash_eve'])
        
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+6, min_col=1, max_col=5):
            for cell in row: cell.alignment = center_align
        current_row = start_row + 9
        
    for col in range(1, 6): ws.column_dimensions[get_column_letter(col)].width = 38
    ws.column_dimensions['A'].width = 25

def build_nist_sheet(wb, records):
    ws = wb.create_sheet(title="NIST")
    _build_nist_sheet_content(ws, records)


def _build_nist_sheet_content(ws, records):
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1

    for record in records:
        ws.cell(
            row=current_row,
            column=1,
            value=f"Skenario {record['skenario']} - Q={record['q']}, R={record['r']}, BB={record['bb']}"
        ).font = Font(bold=True, italic=True)
        current_row += 2

        start_row = current_row
        ws.cell(row=start_row, column=1, value="Parameter NIST").font = header_font
        header_values = ["A & B Best p-value", "A & B Pass Rate (%)", "E-A & E-B Best p-value", "E-A & E-B Pass Rate (%)"]
        for idx, value in enumerate(header_values, start=2):
            ws.cell(row=start_row, column=idx, value=value).font = header_font

        ab_nist = record["nist_ab"]
        eve_nist = record["nist_eve"]
        row = start_row + 1

        ws.cell(row=row, column=1, value="Jumlah Key Diuji")
        ws.cell(row=row, column=2, value=ab_nist["num_keys"])
        ws.cell(row=row, column=4, value=eve_nist["num_keys"])
        row += 1

        ws.cell(row=row, column=1, value="Key Lolos Semua Tes")
        ws.cell(row, column=2, value=ab_nist["passed_all_keys_count"])
        ws.cell(row, column=3, value=(ab_nist["passed_all_keys_count"] / ab_nist["num_keys"] * 100.0) if ab_nist["num_keys"] else 0.0)
        ws.cell(row, column=4, value=eve_nist["passed_all_keys_count"])
        ws.cell(row, column=5, value=(eve_nist["passed_all_keys_count"] / eve_nist["num_keys"] * 100.0) if eve_nist["num_keys"] else 0.0)
        row += 1

        for label, _ in NIST_TEST_ROWS:
            ws.cell(row=row, column=1, value=label)
            ab_stats = ab_nist["tests"][label]
            eve_stats = eve_nist["tests"][label]

            cell = ws.cell(row=row, column=2, value=ab_stats.get("best_pvalue", ab_stats.get("avg_pvalue", 0.0)))
            cell.number_format = '0.000000'
            cell = ws.cell(row=row, column=3, value=ab_stats["pass_rate"])
            cell.number_format = '0.00'
            cell = ws.cell(row=row, column=4, value=eve_stats.get("best_pvalue", eve_stats.get("avg_pvalue", 0.0)))
            cell.number_format = '0.000000'
            cell = ws.cell(row=row, column=5, value=eve_stats["pass_rate"])
            cell.number_format = '0.00'
            row += 1

        ws.cell(row=row, column=1, value="Waktu Komputasi (s)")
        cell = ws.cell(row=row, column=2, value=ab_nist["time_nist"])
        cell.number_format = '0.000'
        cell = ws.cell(row=row, column=4, value=eve_nist["time_nist"])
        cell.number_format = '0.000'

        for row_cells in ws.iter_rows(min_row=start_row, max_row=row, min_col=1, max_col=5):
            for cell in row_cells:
                cell.alignment = center_align

        current_row = row + 3

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 24
    ws.column_dimensions['A'].width = 30

def build_endtoend_sheet(wb, records):
    ws = wb.create_sheet(title="EndToEnd")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        "Skenario", "Q", "R", "BB",
        "Key Kandidat AB", "Key Kandidat Eve",
        "Matched Key Bits AB", "Matched Key Bits Eve",
        "KGR Kumulatif Akhir AB (bit/s)", "KGR Kumulatif Akhir Eve (bit/s)",
        "KDR Kumulatif Akhir AB (%)", "KDR Kumulatif Akhir Eve (%)",
        "t_total AB (s)", "t_total Eve (s)",
    ]
    for col_idx, val in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=val).font = header_font

    for row_idx, r in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=r['skenario'])
        ws.cell(row=row_idx, column=2, value=r['q'])
        ws.cell(row=row_idx, column=3, value=r['r'])
        ws.cell(row=row_idx, column=4, value=r['bb'])
        ws.cell(row=row_idx, column=5, value=r['keys_count_alice'])
        ws.cell(row=row_idx, column=6, value=r['keys_count_ea'])
        ws.cell(row=row_idx, column=7, value=r['matched_key_bits_ab'])
        ws.cell(row=row_idx, column=8, value=r['matched_key_bits_eve'])
        ws.cell(row=row_idx, column=9, value=r['kgr_final_ab'])
        ws.cell(row=row_idx, column=10, value=r['kgr_final_eve'])
        ws.cell(row=row_idx, column=11, value=r['kdr_final_ab'])
        ws.cell(row=row_idx, column=12, value=r['kdr_final_eve'])
        ws.cell(row=row_idx, column=13, value=r['t_total_ab'])
        ws.cell(row=row_idx, column=14, value=r['t_total_eve'])

    for row in ws.iter_rows(min_row=1, max_row=max(2, len(records) + 1), min_col=1, max_col=14):
        for cell in row:
            cell.alignment = center_align

    widths = [10, 8, 8, 8, 16, 15, 18, 18, 30, 30, 28, 30, 14, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

# =====================================================================
# MAIN ENTRY POINT
# =====================================================================
def main():
    print("=== FULL SECRET KEY GENERATION (SKG) AUTOMATION ===")
    base_data = "data"
    output_base = "Output"
    
    # === Inisialisasi list untuk rekap global semua skenario ===
    global_kalman_records = []
    global_kuan_records = []
    global_bch_records = []
    global_hash_records = []
    global_nist_records = []
    global_endtoend_records = []
    
    for skenario in SCENARIOS:
        print(f"\n>>>> Memproses Skenario {skenario} <<<<")
        
        path_alice = os.path.join(base_data, "alice", f"skenario{skenario}_mita_alice.csv")
        path_bob   = os.path.join(base_data, "bob", f"skenario{skenario}_mita_bob.csv")
        path_eve_a = os.path.join(base_data, "eve alice", f"skenario{skenario}_mita_evealice.csv")
        path_eve_b = os.path.join(base_data, "eve bob", f"skenario{skenario}_mita_evebob.csv")
        
        raw_alice = read_rssi_csv(path_alice)
        raw_bob = read_rssi_csv(path_bob)
        raw_eve_a = read_rssi_csv(path_eve_a)
        raw_eve_b = read_rssi_csv(path_eve_b)
        
        if not (raw_alice and raw_bob and raw_eve_a and raw_eve_b):
            print(f"Melewati skenario {skenario} karena data file tidak lengkap di direktori.")
            continue
        skenario_out_dir = os.path.join(output_base, f"skenario_{skenario}")
        excel_kalman_dir = os.path.join(skenario_out_dir, "data_excel_kalman")
        excel_kuan_dir = os.path.join(skenario_out_dir, "data_excel_kuantisasi")
        os.makedirs(excel_kalman_dir, exist_ok=True)
        os.makedirs(excel_kuan_dir, exist_ok=True)
        
        kalman_records = []
        kuan_records = []
        bch_records = []
        hash_records = []
        nist_records = []
        
        for param in PARAM_VARIATIONS:
            q, r, bb = param['q'], param['r'], param['bb']
            print(f" -> Variasi dijalankan: Q={q}, R={r}, BB={bb}")
            
            total_len = min(len(raw_alice), len(raw_bob), len(raw_eve_a), len(raw_eve_b))
            usable_len = (total_len // bb) * bb
            if usable_len == 0:
                print(f" -> Variasi dijalankan: Q={q}, R={r}, BB={bb} dilewati karena data tidak cukup untuk satu blok utuh.")
                continue

            # Pre-processing metrics should describe the full synchronized raw part.
            ra_full = raw_alice[:total_len]
            rb_full = raw_bob[:total_len]
            rea_full = raw_eve_a[:total_len]
            reb_full = raw_eve_b[:total_len]

            # Processing data tetap harus habis dibagi blok.
            ra_proc = raw_alice[:usable_len]
            rb_proc = raw_bob[:usable_len]
            rea_proc = raw_eve_a[:usable_len]
            reb_proc = raw_eve_b[:usable_len]

            # --- 1. Evaluasi Sebelum Praproses ---
            orig_max_alice = np.max(ra_full) if total_len > 0 else 0
            orig_max_bob = np.max(rb_full) if total_len > 0 else 0
            orig_max_evea = np.max(rea_full) if total_len > 0 else 0
            orig_max_eveb = np.max(reb_full) if total_len > 0 else 0

            orig_min_alice = np.min(ra_full) if total_len > 0 else 0
            orig_min_bob = np.min(rb_full) if total_len > 0 else 0
            orig_min_evea = np.min(rea_full) if total_len > 0 else 0
            orig_min_eveb = np.min(reb_full) if total_len > 0 else 0

            orig_corr_ab = calc_corr(ra_full, rb_full)
            orig_corr_eve = calc_corr(rea_full, reb_full)
            
            # --- 2. Filter Kalman Praproses per blok ---
            kal_blocks_a, kal_a, kal_times_a, kgr_kal_a_local, time_kal_a = process_kalman_per_blocks(ra_proc, q, r, bb, benchmark_iterations=BENCHMARK_ITERATIONS)
            kal_blocks_b, kal_b, kal_times_b, kgr_kal_b_local, time_kal_b = process_kalman_per_blocks(rb_proc, q, r, bb, benchmark_iterations=BENCHMARK_ITERATIONS)
            kal_blocks_ea, kal_ea, kal_times_ea, kgr_kal_ea_local, time_kal_ea = process_kalman_per_blocks(rea_proc, q, r, bb, benchmark_iterations=BENCHMARK_ITERATIONS)
            kal_blocks_eb, kal_eb, kal_times_eb, kgr_kal_eb_local, time_kal_eb = process_kalman_per_blocks(reb_proc, q, r, bb, benchmark_iterations=BENCHMARK_ITERATIONS)

            if not kal_a or not kal_b or not kal_ea or not kal_eb:
                print(f" -> Variasi dijalankan: Q={q}, R={r}, BB={bb} dilewati karena Kalman per blok gagal menghasilkan data.")
                continue

            # Simpan Sinyal array ke excel (.xlsx)
            v_name = f"Q{q}_R{r}_BB{bb}"
            save_block_series(excel_kalman_dir, f"{v_name}_kalman_alice_blocks.xlsx", kal_blocks_a, "alice_kalman_block")
            save_block_series(excel_kalman_dir, f"{v_name}_kalman_bob_blocks.xlsx", kal_blocks_b, "bob_kalman_block")
            save_block_series(excel_kalman_dir, f"{v_name}_kalman_evealice_blocks.xlsx", kal_blocks_ea, "evealice_kalman_block")
            save_block_series(excel_kalman_dir, f"{v_name}_kalman_evebob_blocks.xlsx", kal_blocks_eb, "evebob_kalman_block")
            save_data_list(excel_kalman_dir, f"{v_name}_kalman_alice.xlsx", kal_a, "alice_kalman")
            save_data_list(excel_kalman_dir, f"{v_name}_kalman_bob.xlsx", kal_b, "bob_kalman")
            save_data_list(excel_kalman_dir, f"{v_name}_kalman_evealice.xlsx", kal_ea, "evealice_kalman")
            save_data_list(excel_kalman_dir, f"{v_name}_kalman_evebob.xlsx", kal_eb, "evebob_kalman")
            
            kal_max_alice = np.max(kal_a) if len(kal_a)>0 else 0
            kal_max_bob = np.max(kal_b) if len(kal_b)>0 else 0
            kal_max_evea = np.max(kal_ea) if len(kal_ea)>0 else 0
            kal_max_eveb = np.max(kal_eb) if len(kal_eb)>0 else 0
            
            kal_min_alice = np.min(kal_a) if len(kal_a)>0 else 0
            kal_min_bob = np.min(kal_b) if len(kal_b)>0 else 0
            kal_min_evea = np.min(kal_ea) if len(kal_ea)>0 else 0
            kal_min_eveb = np.min(kal_eb) if len(kal_eb)>0 else 0
            
            kal_corr_ab = calc_corr(kal_a, kal_b)
            kal_corr_eve = calc_corr(kal_ea, kal_eb)
            
            kalman_records.append({
                "skenario": skenario, "q": q, "r": r, "bb": bb,
                "orig_max_alice": orig_max_alice, "orig_max_bob": orig_max_bob, "orig_max_evealice": orig_max_evea, "orig_max_evebob": orig_max_eveb,
                "orig_min_alice": orig_min_alice, "orig_min_bob": orig_min_bob, "orig_min_evealice": orig_min_evea, "orig_min_evebob": orig_min_eveb,
                "orig_corr_ab": orig_corr_ab, "orig_corr_eve": orig_corr_eve,
                "kalman_max_alice": kal_max_alice, "kalman_max_bob": kal_max_bob, "kalman_max_evealice": kal_max_evea, "kalman_max_evebob": kal_max_eveb,
                "kalman_min_alice": kal_min_alice, "kalman_min_bob": kal_min_bob, "kalman_min_evealice": kal_min_evea, "kalman_min_evebob": kal_min_eveb,
                "kalman_corr_ab": kal_corr_ab, "kalman_corr_eve": kal_corr_eve,
                "time_alice": time_kal_a, "time_bob": time_kal_b, "time_evealice": time_kal_ea, "time_evebob": time_kal_eb,
                "block_count": len(kal_blocks_a)
            })
            
            # --- 3. Kuantisasi Multibit per blok hasil Kalman ---
            bs_blocks_a, bs_a, kgr_blocks_a, kgr_kuan_a_local, time_kuan_a = process_kuantisasi_per_blocks(kal_blocks_a, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
            bs_blocks_b, bs_b, kgr_blocks_b, kgr_kuan_b_local, time_kuan_b = process_kuantisasi_per_blocks(kal_blocks_b, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
            bs_blocks_ea, bs_ea, kgr_blocks_ea, kgr_kuan_ea_local, time_kuan_ea = process_kuantisasi_per_blocks(kal_blocks_ea, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
            bs_blocks_eb, bs_eb, kgr_blocks_eb, kgr_kuan_eb_local, time_kuan_eb = process_kuantisasi_per_blocks(kal_blocks_eb, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)

            if not bs_a or not bs_b or not bs_ea or not bs_eb:
                print(f" -> Variasi dijalankan: Q={q}, R={r}, BB={bb} dilewati karena kuantisasi per blok gagal menghasilkan bitstream.")
                continue

            # KGR Kuantisasi kumulatif = panjang bitstream / (t_chanprob + t_kalman + t_kuantisasi)
            kgr_kuan_a = calculate_cumulative_kgr(len(bs_a), CHANPROB_TIME_SECONDS, time_kal_a, time_kuan_a)
            kgr_kuan_b = calculate_cumulative_kgr(len(bs_b), CHANPROB_TIME_SECONDS, time_kal_b, time_kuan_b)
            kgr_kuan_ea = calculate_cumulative_kgr(len(bs_ea), CHANPROB_TIME_SECONDS, time_kal_ea, time_kuan_ea)
            kgr_kuan_eb = calculate_cumulative_kgr(len(bs_eb), CHANPROB_TIME_SECONDS, time_kal_eb, time_kuan_eb)
            
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_alice.xlsx", [bs_a], "bitstream")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_bob.xlsx", [bs_b], "bitstream")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evealice.xlsx", [bs_ea], "bitstream")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evebob.xlsx", [bs_eb], "bitstream")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_alice_blocks.xlsx", bs_blocks_a, "bitstream_block")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_bob_blocks.xlsx", bs_blocks_b, "bitstream_block")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evealice_blocks.xlsx", bs_blocks_ea, "bitstream_block")
            save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evebob_blocks.xlsx", bs_blocks_eb, "bitstream_block")
            
            kdr_blocks_ab = [calculate_kdr(a, b) for a, b in zip(bs_blocks_a, bs_blocks_b)]
            kdr_blocks_eve = [calculate_kdr(a, b) for a, b in zip(bs_blocks_ea, bs_blocks_eb)]

            kdr_ab = calculate_kdr(bs_a, bs_b)
            kdr_eve = calculate_kdr(bs_ea, bs_eb)
            kdr_block_ab = float(np.mean(kdr_blocks_ab)) if kdr_blocks_ab else 0.0
            kdr_block_eve = float(np.mean(kdr_blocks_eve)) if kdr_blocks_eve else 0.0

            compared_bits_ab = min(len(bs_a), len(bs_b))
            compared_bits_eve = min(len(bs_ea), len(bs_eb))
            avg_time_kuan_ab = float(np.mean([time_kuan_a, time_kuan_b]))
            avg_time_kuan_eve = float(np.mean([time_kuan_ea, time_kuan_eb]))
            avg_time_kal_ab = float(np.mean([time_kal_a, time_kal_b]))
            avg_time_kal_eve = float(np.mean([time_kal_ea, time_kal_eb]))
            kgr_merge_local_ab = calculate_local_kgr(compared_bits_ab, avg_time_kuan_ab)
            kgr_merge_local_eve = calculate_local_kgr(compared_bits_eve, avg_time_kuan_eve)
            kgr_merge_cum_ab = calculate_cumulative_kgr(compared_bits_ab, CHANPROB_TIME_SECONDS, avg_time_kal_ab, avg_time_kuan_ab)
            kgr_merge_cum_eve = calculate_cumulative_kgr(compared_bits_eve, CHANPROB_TIME_SECONDS, avg_time_kal_eve, avg_time_kuan_eve)
            
            kuan_records.append({
                "skenario": skenario, "q": q, "r": r, "bb": bb,
                "kdr_ab": kdr_ab, "kdr_eve": kdr_eve,
                "kdr_block_ab": kdr_block_ab, "kdr_block_eve": kdr_block_eve,
                "time_alice": time_kuan_a, "time_bob": time_kuan_b, "time_evealice": time_kuan_ea, "time_evebob": time_kuan_eb,
                "kgr_local_alice": kgr_kuan_a_local, "kgr_local_bob": kgr_kuan_b_local, "kgr_local_evealice": kgr_kuan_ea_local, "kgr_local_evebob": kgr_kuan_eb_local,
                "kgr_cum_alice": kgr_kuan_a, "kgr_cum_bob": kgr_kuan_b, "kgr_cum_evealice": kgr_kuan_ea, "kgr_cum_evebob": kgr_kuan_eb,
                "kgr_merge_local_ab": kgr_merge_local_ab, "kgr_merge_local_eve": kgr_merge_local_eve,
                "kgr_merge_cum_ab": kgr_merge_cum_ab, "kgr_merge_cum_eve": kgr_merge_cum_eve,
                "total_bits_alice": len(bs_a), "total_bits_bob": len(bs_b),
                "total_bits_ea": len(bs_ea), "total_bits_eb": len(bs_eb),
                "block_count": len(bs_blocks_a)
            })
            
            # --- 4. BCH, Hash dan NIST Test (Simulasi modul) ---
            try:
                from bch_module import process_bch
                b_alice, b_bob, stats_ab = process_bch(bs_a, bs_b, apply_correction=True)
                b_ea, b_eb, stats_eve = process_bch(bs_ea, bs_eb, apply_correction=False)
                # KGR BCH kumulatif dihitung dari bitstream hasil BCH terhadap total waktu sampai BCH.
                kgr_bch_ab = calculate_cumulative_kgr(
                    len(b_alice),
                    CHANPROB_TIME_SECONDS,
                    time_kal_a,
                    time_kuan_a,
                    stats_ab["time_bch"],
                )
                kgr_bch_eve = calculate_cumulative_kgr(
                    len(b_ea),
                    CHANPROB_TIME_SECONDS,
                    time_kal_ea,
                    time_kuan_ea,
                    stats_eve["time_bch"],
                )
                
                # Biar user bisa lihat list data array bits nya
                os.makedirs(os.path.join(skenario_out_dir, "data_excel_bch"), exist_ok=True)
                save_data_list(os.path.join(skenario_out_dir, "data_excel_bch"), f"{v_name}_bch_alice.xlsx", b_alice, "alice_bch_bits")
                save_data_list(os.path.join(skenario_out_dir, "data_excel_bch"), f"{v_name}_bch_bob.xlsx", b_bob, "bob_bch_bits")
                
                bch_records.append({
                    "skenario": skenario, "q": q, "r": r, "bb": bb,
                    "kdr_after_ab": stats_ab["kdr_after"], "kdr_after_eve": stats_eve["kdr_after"],
                    "kgr_bch_ab": kgr_bch_ab, "kgr_bch_eve": kgr_bch_eve,
                    "parity_bits_ab": stats_ab["parity_bits_sent"], "parity_bits_eve": stats_eve["parity_bits_sent"],
                    "total_bits_alice": stats_ab["total_bits_alice"], "total_bits_bob": stats_ab["total_bits_bob"],
                    "total_bits_ea": stats_eve["total_bits_alice"], "total_bits_eb": stats_eve["total_bits_bob"],
                    "error_bits_ab_before": stats_ab["error_bits_before"], "error_bits_eve_before": stats_eve["error_bits_before"],
                    "error_bits_ab_after": stats_ab["error_bits_after"], "error_bits_eve_after": stats_eve["error_bits_after"],
                    "corrected_bits_ab": stats_ab["corrected_bits"], "corrected_bits_eve": stats_eve["corrected_bits"],
                    "time_bch_ab": stats_ab["time_bch"], "time_bch_eve": stats_eve["time_bch"]
                })
                
                # --- 5. Hash & SHA & AES ---
                try: 
                    from hash_module import process_hash
                    # Proses Alice dan Bob
                    h_alice, h_bob, aes_ab, time_hash_ab, hash_metrics_ab = process_hash(b_alice, b_bob)
                    # Proses Eve-Alice dan Eve-Bob
                    h_ea, h_eb, aes_eve, time_hash_eve, hash_metrics_eve = process_hash(b_ea, b_eb)
                    
                    # Simpan Smua Keys (termasuk yg salah)
                    os.makedirs(os.path.join(skenario_out_dir, "data_excel_hash"), exist_ok=True)
                    save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_alice.xlsx", h_alice, "AES_keys")
                    save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_bob.xlsx", h_bob, "AES_keys")
                    save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_evealice.xlsx", h_ea, "AES_keys")
                    save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_evebob.xlsx", h_eb, "AES_keys")
                    
                    hash_records.append({
                        "skenario": skenario, "q": q, "r": r, "bb": bb,
                        "aes_count_ab": len(aes_ab), "aes_count_eve": len(aes_eve),
                        "keys_count_alice": hash_metrics_ab["keys_count_alice"], "keys_count_bob": hash_metrics_ab["keys_count_bob"],
                        "keys_count_ea": hash_metrics_eve["keys_count_alice"], "keys_count_eb": hash_metrics_eve["keys_count_bob"],
                        "total_key_bits_alice": hash_metrics_ab["total_key_bits_alice"],
                        "total_key_bits_bob": hash_metrics_ab["total_key_bits_bob"],
                        "total_key_bits_ea": hash_metrics_eve["total_key_bits_alice"],
                        "total_key_bits_eb": hash_metrics_eve["total_key_bits_bob"],
                        "matched_key_bits_ab": hash_metrics_ab["matched_key_bits"],
                        "matched_key_bits_eve": hash_metrics_eve["matched_key_bits"],
                        "final_key_alice": h_alice[0] if len(h_alice) > 0 else "N/A",
                        "final_key_bob": h_bob[0] if len(h_bob) > 0 else "N/A",
                        "final_key_ea": h_ea[0] if len(h_ea) > 0 else "N/A",
                        "final_key_eb": h_eb[0] if len(h_eb) > 0 else "N/A",
                        "best_key_alice": "TBD_NIST",  # Will be updated after NIST
                        "best_key_bob": "TBD_NIST",
                        "best_key_ea": "TBD_NIST",
                        "best_key_eb": "TBD_NIST",
                        "time_hash_ab": time_hash_ab, "time_hash_eve": time_hash_eve,
                        # Detail untuk 3 tabel (Hash | SHA-1 | Final)
                        "hex_alice_ab": hash_metrics_ab["hex_alice"],
                        "hex_bob_ab": hash_metrics_ab["hex_bob"],
                        "sha_alice_ab": hash_metrics_ab["sha_keys_alice"],
                        "sha_bob_ab": hash_metrics_ab["sha_keys_bob"],
                        "matched_idx_ab": hash_metrics_ab["matched_indices"],
                        "hex_alice_eve": hash_metrics_eve["hex_alice"],
                        "hex_bob_eve": hash_metrics_eve["hex_bob"],
                        "sha_alice_eve": hash_metrics_eve["sha_keys_alice"],
                        "sha_bob_eve": hash_metrics_eve["sha_keys_bob"],
                        "matched_idx_eve": hash_metrics_eve["matched_indices"],
                    })

                    # KGR BCH sudah dihitung langsung dari corrected_bits sebelumnya
                    # (tidak perlu update dari Hash)

                    total_compared_bits_ab = min(
                        hash_metrics_ab["total_key_bits_alice"],
                        hash_metrics_ab["total_key_bits_bob"],
                    )
                    total_compared_bits_eve = min(
                        hash_metrics_eve["total_key_bits_alice"],
                        hash_metrics_eve["total_key_bits_bob"],
                    )

                    t_total_ab = (
                        CHANPROB_TIME_SECONDS
                        + float(time_kal_a)
                        + float(time_kuan_a)
                        + float(stats_ab["time_bch"])
                        + float(time_hash_ab)
                    )
                    t_total_eve = (
                        CHANPROB_TIME_SECONDS
                        + float(time_kal_ea)
                        + float(time_kuan_ea)
                        + float(stats_eve["time_bch"])
                        + float(time_hash_eve)
                    )

                    kgr_final_ab = calculate_cumulative_kgr(
                        hash_metrics_ab["matched_key_bits"],
                        CHANPROB_TIME_SECONDS,
                        time_kal_a,
                        time_kuan_a,
                        stats_ab["time_bch"],
                        time_hash_ab,
                    )
                    kgr_final_eve = calculate_cumulative_kgr(
                        hash_metrics_eve["matched_key_bits"],
                        CHANPROB_TIME_SECONDS,
                        time_kal_ea,
                        time_kuan_ea,
                        stats_eve["time_bch"],
                        time_hash_eve,
                    )

                    kdr_final_ab = calculate_kdr_from_matched_bits(
                        hash_metrics_ab["matched_key_bits"],
                        total_compared_bits_ab,
                    )
                    kdr_final_eve = calculate_kdr_from_matched_bits(
                        hash_metrics_eve["matched_key_bits"],
                        total_compared_bits_eve,
                    )

                    global_endtoend_records.append({
                        "skenario": skenario,
                        "q": q,
                        "r": r,
                        "bb": bb,
                        "keys_count_alice": hash_metrics_ab["keys_count_alice"],
                        "keys_count_bob": hash_metrics_ab["keys_count_bob"],
                        "keys_count_ea": hash_metrics_eve["keys_count_alice"],
                        "keys_count_eb": hash_metrics_eve["keys_count_bob"],
                        "matched_key_bits_ab": hash_metrics_ab["matched_key_bits"],
                        "matched_key_bits_eve": hash_metrics_eve["matched_key_bits"],
                        "kgr_final_ab": kgr_final_ab,
                        "kgr_final_eve": kgr_final_eve,
                        "kdr_final_ab": kdr_final_ab,
                        "kdr_final_eve": kdr_final_eve,
                        "t_total_ab": t_total_ab,
                        "t_total_eve": t_total_eve,
                    })
                    
                    # --- 6. Uji NIST ---
                    try:
                        from nist_module import process_nist
                        nist_ab = process_nist(aes_ab)
                        nist_eve = process_nist(aes_eve)

                        nist_records.append({
                            "skenario": skenario, "q": q, "r": r, "bb": bb,
                            "nist_ab": nist_ab, "nist_eve": nist_eve,
                        })
                        # Update last hash_records entry to show best key by Approximate Entropy
                        # Fallback ke first key jika NIST tidak memberikan best key
                        try:
                            if hash_records:
                                hr = hash_records[-1]
                                best_ab = nist_ab.get('best_key_by_apen') if isinstance(nist_ab, dict) else None
                                if best_ab:
                                    hr['best_key_alice'] = best_ab
                                    hr['best_key_bob'] = best_ab
                                else:
                                    # Fallback ke first key jika NIST tidak ada best key
                                    hr['best_key_alice'] = hr['final_key_alice']
                                    hr['best_key_bob'] = hr['final_key_bob']
                                    
                                best_eve = nist_eve.get('best_key_by_apen') if isinstance(nist_eve, dict) else None
                                if best_eve:
                                    hr['best_key_ea'] = best_eve
                                    hr['best_key_eb'] = best_eve
                                else:
                                    # Fallback ke first key jika NIST tidak ada best key
                                    hr['best_key_ea'] = hr['final_key_ea']
                                    hr['best_key_eb'] = hr['final_key_eb']
                        except Exception:
                            pass
                    except Exception as e:
                        print("NIST Modul Error:", e)
                        
                except Exception as e:
                    print("Hash Modul Error:", e)
            except ImportError:
                print("Modul BCH Belum ditaruh di root folder!")
            
        # Mengukir Tabel Excel untuk Skenario 
        print(" == Menyusun File Table Laporan ==")
        build_kalman_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Kalman.xlsx"), kalman_records)
        build_kuantisasi_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Kuantisasi.xlsx"), kuan_records)
        
        if bch_records:
            build_bch_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_BCH.xlsx"), bch_records)
        if hash_records:
            build_hash_detailed_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Hash.xlsx"), hash_records)
        if nist_records:
            build_nist_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_NIST.xlsx"), nist_records)
            
        print("Selesai diproses untuk Skenario", skenario)
        
        # Tambahkan ke dictionary global untuk rekap semua skenario nanti
        global_kalman_records.extend(kalman_records)
        global_kuan_records.extend(kuan_records)
        if bch_records: global_bch_records.extend(bch_records)
        if hash_records: global_hash_records.extend(hash_records)
        if nist_records: global_nist_records.extend(nist_records)

    # === GENERATE GLOBAL REKAP ===
    print("\n==== MERANGKUM SELURUH SKENARIO KE DALAM SATU FILE EXCEL ====")
    rekap_excel_path = os.path.join(output_base, "Rekap_Evaluasi_SKG_Semua_Skenario.xlsx")
    
    # Buat workbook kosong
    rekap_wb = Workbook()
    
    # Hapus sheet default bawaan ('Sheet')
    if 'Sheet' in rekap_wb.sheetnames:
        rekap_wb.remove(rekap_wb['Sheet'])
        
    print("Menyusun sheet Kalman...")
    build_kalman_sheet(rekap_wb, global_kalman_records)
    
    print("Menyusun sheet Kuantisasi...")
    build_kuantisasi_sheet(rekap_wb, global_kuan_records)
    
    if global_bch_records:
        print("Menyusun sheet BCH...")
        build_bch_sheet(rekap_wb, global_bch_records)
        
    if global_hash_records:
        print("Menyusun sheet Hash...")
        build_hash_sheet(rekap_wb, global_hash_records)
        
    if global_nist_records:
        print("Menyusun sheet NIST...")
        build_nist_sheet(rekap_wb, global_nist_records)

    if global_endtoend_records:
        print("Menyusun sheet EndToEnd...")
        build_endtoend_sheet(rekap_wb, global_endtoend_records)
        
    rekap_wb.save(rekap_excel_path)
    print(f"Selesai! File rekap global berhasil disimpan di: {rekap_excel_path}")

if __name__ == "__main__":
    main()
