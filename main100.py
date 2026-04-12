import os
import csv
import numpy as np
import pandas as pd
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Impor dari Modul Terpisah ===
from kalman_module import process_kalman
from kuantisasi_module import process_kuantisasi
try:
    from bch_module import process_bch
    from hash_module import process_hash
    from nist_module import process_nist
except ImportError:
    pass  # Akan dibuat nanti

# =====================================================================
# GLOBAL PARAMETERS
# =====================================================================
KUANTISASI_NUM_BITS = 3
BENCHMARK_ITERATIONS = 10

# Variasi Parameter Pengujian Skenario
PARAM_VARIATIONS = [
    {"q": 0.01, "r": 0.5, "bb": 1},
    {"q": 0.01, "r": 0.5, "bb": 5},
    {"q": 0.01, "r": 0.5, "bb": 50},
    {"q": 0.01, "r": 0.5, "bb": 100},
    {"q": 0.5, "r": 0.01, "bb": 1},
    {"q": 0.5, "r": 0.01, "bb": 5},
    {"q": 0.5, "r": 0.01, "bb": 50},
    {"q": 0.5, "r": 0.01, "bb": 100},
]

# Skenario iterasi 1 - 4
SCENARIOS = [1, 2, 3, 4]
CHUNK_SIZE = 100

# =====================================================================
# UTILITY FUNCTIONS
# =====================================================================
def read_rssi_csv(path):
    data = []
    if not os.path.exists(path):
        print(f"Warning: File {path} tidak ditemukan!")
        return data
        
    with open(path, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            try:
                data.append(int(row[0]))
            except:
                continue
    return data

def calculate_kdr(a, b):
    if not a or not b: return 0.0
    n = min(len(a), len(b))
    if n == 0: return 0.0
    diff = sum(1 for i in range(n) if a[i] != b[i])
    return (diff / n) * 100.0

def calc_corr(a, b):
    ln = min(len(a), len(b))
    if ln < 2: return "N/A"
    c, _ = pearsonr(a[:ln], b[:ln])
    return c

# =====================================================================
# EXCEL FORMATTING FUNCTIONS
# =====================================================================
def save_data_list(output_dir, filename, data_list, header):
    os.makedirs(output_dir, exist_ok=True)
    df = pd.DataFrame({header: data_list})
    df.to_excel(os.path.join(output_dir, filename), index=False)

def build_kalman_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kalman"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    
    for r in records:
        # Title Data Block
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2
        
        start_row = current_row
        # Headers Top 
        ws.cell(row=start_row, column=1, value="Parameter").font = header_font
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
        
        ws.cell(row=start_row, column=2, value="Sebelum Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=5)
        
        ws.cell(row=start_row, column=6, value="Setelah Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=9)
        
        # Headers Low
        cols_names = ["Alice", "Bob", "Eve-Alice", "Eve-Bob", "Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, cname in enumerate(cols_names):
            ws.cell(row=start_row+1, column=2+idx, value=cname).font = header_font
            
        # Data Maksimum
        ws.cell(row=start_row+2, column=1, value="Maksimum (dBm)")
        vals_max = [r['orig_max_alice'], r['orig_max_bob'], r['orig_max_evealice'], r['orig_max_evebob'],
                    r['kalman_max_alice'], r['kalman_max_bob'], r['kalman_max_evealice'], r['kalman_max_evebob']]
        for idx, val in enumerate(vals_max): ws.cell(row=start_row+2, column=2+idx, value=val)
            
        # Data Minimum
        ws.cell(row=start_row+3, column=1, value="Minimum (dBm)")
        vals_min = [r['orig_min_alice'], r['orig_min_bob'], r['orig_min_evealice'], r['orig_min_evebob'],
                    r['kalman_min_alice'], r['kalman_min_bob'], r['kalman_min_evealice'], r['kalman_min_evebob']]
        for idx, val in enumerate(vals_min): ws.cell(row=start_row+3, column=2+idx, value=val)
            
        # Korelasi
        ws.cell(row=start_row+4, column=1, value="Koefisien Korelasi")
        
        # Sebelum Korelasi A&B + E&E (Merge 2 cell)
        ws.cell(row=start_row+4, column=2, value=r['orig_corr_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=2, end_row=start_row+4, end_column=3)
        ws.cell(row=start_row+4, column=4, value=r['orig_corr_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=4, end_row=start_row+4, end_column=5)
        
        # Sesudah Korelasi A&B + E&E
        ws.cell(row=start_row+4, column=6, value=r['kalman_corr_ab'])
        ws.merge_cells(start_row=start_row+4, start_column=6, end_row=start_row+4, end_column=7)
        ws.cell(row=start_row+4, column=8, value=r['kalman_corr_eve'])
        ws.merge_cells(start_row=start_row+4, start_column=8, end_row=start_row+4, end_column=9)
        
        # Waktu Komputasi
        ws.cell(row=start_row+5, column=1, value="Waktu Komputasi (s)")
        ws.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=5)
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]):
            ws.cell(row=start_row+5, column=6+idx, value=val)
            
        # KGR
        ws.cell(row=start_row+6, column=1, value="KGR (bit/s)")
        ws.merge_cells(start_row=start_row+6, start_column=1, end_row=start_row+6, end_column=5)
        for idx, val in enumerate([r['kgr_alice'], r['kgr_bob'], r['kgr_evealice'], r['kgr_evebob']]):
            ws.cell(row=start_row+6, column=6+idx, value=val)

        # Style alignment applying for all cells in block
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+6, min_col=1, max_col=9):
            for cell in row: cell.alignment = center_align

        # Additional 3 spaces for the next table
        current_row = start_row + 10 
        
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 17
    ws.column_dimensions['A'].width = 25
        
    wb.save(output_path)

def build_kuantisasi_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Kuantisasi"
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    
    for r in records:
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2
        
        start_row = current_row
        
        # Headers
        ws.cell(row=start_row, column=1, value="Parameter Performansi").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=2+idx, value=val).font = header_font
            
        # KDR
        ws.cell(row=start_row+1, column=1, value="KDR (%)")
        ws.cell(row=start_row+1, column=2, value=r['kdr_ab'])
        ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+1, end_column=3)
        ws.cell(row=start_row+1, column=4, value=r['kdr_eve'])
        ws.merge_cells(start_row=start_row+1, start_column=4, end_row=start_row+1, end_column=5)
        
        # KGR
        ws.cell(row=start_row+2, column=1, value="KGR (bit/s)")
        for idx, val in enumerate([r['kgr_alice'], r['kgr_bob'], r['kgr_evealice'], r['kgr_evebob']]):
            ws.cell(row=start_row+2, column=2+idx, value=val)
            
        # Waktu Komputasi
        ws.cell(row=start_row+3, column=1, value="Waktu komputasi (s)")
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]):
            ws.cell(row=start_row+3, column=2+idx, value=val)
            
        for row in ws.iter_rows(min_row=start_row, max_row=start_row+3, min_col=1, max_col=5):
            for cell in row: cell.alignment = center_align

        current_row = start_row + 7
        
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
        
    wb.save(output_path)

def build_bch_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap BCH"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row

        ws.cell(row=start_row, column=1, value="Parameter BCH").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row + 1, column=2, value=r['kdr_after_ab'])
        ws.cell(row=start_row + 1, column=3, value=r['kdr_after_eve'])

        ws.cell(row=start_row + 2, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row + 2, column=2, value=r['time_bch_ab'])
        ws.cell(row=start_row + 2, column=3, value=r['time_bch_eve'])

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 2, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = center_align

        current_row = start_row + 5

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25
    wb.save(output_path)

def build_hash_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Hash"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row

        ws.cell(row=start_row, column=1, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="Jumlah Kunci Cocok (Match)")
        ws.cell(row=start_row + 1, column=2, value=r['aes_count_ab'])
        ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=3)
        ws.cell(row=start_row + 1, column=4, value=r['aes_count_eve'])
        ws.merge_cells(start_row=start_row + 1, start_column=4, end_row=start_row + 1, end_column=5)

        ws.cell(row=start_row + 2, column=1, value="Kunci Pertama (Hex)")
        ws.cell(row=start_row + 2, column=2, value=r['final_key_alice'])
        ws.cell(row=start_row + 2, column=3, value=r['final_key_bob'])
        ws.cell(row=start_row + 2, column=4, value=r['final_key_ea'])
        ws.cell(row=start_row + 2, column=5, value=r['final_key_eb'])

        ws.cell(row=start_row + 3, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row + 3, column=2, value=r['time_hash_ab'])
        ws.cell(row=start_row + 3, column=3, value=r['time_hash_ab'])
        ws.cell(row=start_row + 3, column=4, value=r['time_hash_eve'])
        ws.cell(row=start_row + 3, column=5, value=r['time_hash_eve'])

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 3, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = center_align
        current_row = start_row + 6

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 38
    ws.column_dimensions['A'].width = 25
    wb.save(output_path)

def build_nist_excel(output_path, records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap NIST"
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row

        ws.cell(row=start_row, column=1, value="Parameter NIST").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="Jumlah Key Lulus")
        ws.cell(row=start_row + 1, column=2, value=r['passed_keys_ab'])
        ws.cell(row=start_row + 1, column=3, value=r['passed_keys_eve'])

        ws.cell(row=start_row + 2, column=1, value="Rata-rata p-value (ApEn)")
        ws.cell(row=start_row + 2, column=2, value=r['pval_ab'])
        ws.cell(row=start_row + 2, column=3, value=r['pval_eve'])

        ws.cell(row=start_row + 3, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row + 3, column=2, value=r['time_nist_ab'])
        ws.cell(row=start_row + 3, column=3, value=r['time_nist_eve'])

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 3, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = center_align
        current_row = start_row + 6

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25
    wb.save(output_path)

def build_bch_sheet(wb, records):
    ws = wb.create_sheet(title="BCH")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Pengujian Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row

        ws.cell(row=start_row, column=1, value="Parameter BCH").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row + 1, column=2, value=r['kdr_after_ab'])
        ws.cell(row=start_row + 1, column=3, value=r['kdr_after_eve'])

        ws.cell(row=start_row + 2, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row + 2, column=2, value=r['time_bch_ab'])
        ws.cell(row=start_row + 2, column=3, value=r['time_bch_eve'])

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 2, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = center_align

        current_row = start_row + 5

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # ===============================================================
    # TABEL RATA-RATA BCH (kanan tabel existing)
    # ===============================================================
    def _base_skenario_label(val):
        txt = str(val)
        if "(Part" in txt:
            return txt.split(" (Part")[0].strip()
        return txt

    def _mean_numeric(values):
        nums = []
        for v in values:
            try:
                nums.append(float(v))
            except:
                continue
        if not nums:
            return "N/A"
        return float(np.mean(nums))

    avg_start_col = 6
    avg_row = 1
    seen = []
    grouped = {}

    for rec in records:
        key = (_base_skenario_label(rec['skenario']), rec['q'], rec['r'], rec['bb'])
        if key not in grouped:
            grouped[key] = []
            seen.append(key)
        grouped[key].append(rec)

    for skenario_label, q, r, bb in seen:
        recs = grouped[(skenario_label, q, r, bb)]
        avg_kdr_ab = _mean_numeric([r['kdr_after_ab'] for r in recs])
        avg_kdr_eve = _mean_numeric([r['kdr_after_eve'] for r in recs])

        title_val = f"Pengujian Skenario {skenario_label} - Q={q}, R={r}, BB={bb}"
        ws.cell(row=avg_row, column=avg_start_col, value=title_val).font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=avg_row, start_column=avg_start_col, end_row=avg_row, end_column=avg_start_col + 2)
        avg_row += 2

        start_row = avg_row
        ws.cell(row=start_row, column=avg_start_col, value="Parameter BCH").font = header_font
        ws.cell(row=start_row, column=avg_start_col + 1, value="A & B").font = header_font
        ws.cell(row=start_row, column=avg_start_col + 2, value="E-A & E-B").font = header_font

        ws.cell(row=start_row + 1, column=avg_start_col, value="Rata-rata KDR Setelah koreksi BCH (%)")
        ws.cell(row=start_row + 1, column=avg_start_col + 1, value=avg_kdr_ab)
        ws.cell(row=start_row + 1, column=avg_start_col + 2, value=avg_kdr_eve)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 1, min_col=avg_start_col, max_col=avg_start_col + 2):
            for cell in row:
                cell.alignment = center_align

        avg_row = start_row + 4

    for col in range(avg_start_col, avg_start_col + 3):
        ws.column_dimensions[get_column_letter(col)].width = 28

def build_hash_sheet(wb, records):
    ws = wb.create_sheet(title="Hash_SHA_AES")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row

        ws.cell(row=start_row, column=1, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="Jumlah Kunci Cocok (Match)")
        ws.cell(row=start_row + 1, column=2, value=r['aes_count_ab'])
        ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=3)
        ws.cell(row=start_row + 1, column=4, value=r['aes_count_eve'])
        ws.merge_cells(start_row=start_row + 1, start_column=4, end_row=start_row + 1, end_column=5)

        ws.cell(row=start_row + 2, column=1, value="Kunci Pertama (Hex)")
        ws.cell(row=start_row + 2, column=2, value=r['final_key_alice'])
        ws.cell(row=start_row + 2, column=3, value=r['final_key_bob'])
        ws.cell(row=start_row + 2, column=4, value=r['final_key_ea'])
        ws.cell(row=start_row + 2, column=5, value=r['final_key_eb'])

        ws.cell(row=start_row + 3, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row + 3, column=2, value=r['time_hash_ab'])
        ws.cell(row=start_row + 3, column=3, value=r['time_hash_ab'])
        ws.cell(row=start_row + 3, column=4, value=r['time_hash_eve'])
        ws.cell(row=start_row + 3, column=5, value=r['time_hash_eve'])

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 3, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = center_align
        current_row = start_row + 6

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 38
    ws.column_dimensions['A'].width = 25

    # ===============================================================
    # TABEL RATA-RATA/REPRESENTATIF HASH (kanan tabel existing)
    # ===============================================================
    def _base_skenario_label(val):
        txt = str(val)
        if "(Part" in txt:
            return txt.split(" (Part")[0].strip()
        return txt

    def _mode_text(values):
        freq = {}
        order = {}
        for idx, v in enumerate(values):
            s = str(v).strip() if v is not None else ""
            if not s or s == "N/A":
                continue
            if s not in freq:
                freq[s] = 0
                order[s] = idx
            freq[s] += 1
        if not freq:
            return "N/A"
        return sorted(freq.keys(), key=lambda k: (-freq[k], order[k]))[0]

    avg_start_col = 8
    avg_row = 1
    seen = []
    grouped = {}

    for rec in records:
        key = (_base_skenario_label(rec['skenario']), rec['q'], rec['r'], rec['bb'])
        if key not in grouped:
            grouped[key] = []
            seen.append(key)
        grouped[key].append(rec)

    for skenario_label, q, r, bb in seen:
        recs = grouped[(skenario_label, q, r, bb)]

        key_alice = _mode_text([r['final_key_alice'] for r in recs])
        key_bob = _mode_text([r['final_key_bob'] for r in recs])
        key_ea = _mode_text([r['final_key_ea'] for r in recs])
        key_eb = _mode_text([r['final_key_eb'] for r in recs])

        title_val = f"Skenario {skenario_label} - Q={q}, R={r}, BB={bb}"
        ws.cell(row=avg_row, column=avg_start_col, value=title_val).font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=avg_row, start_column=avg_start_col, end_row=avg_row, end_column=avg_start_col + 4)
        avg_row += 2

        start_row = avg_row
        ws.cell(row=start_row, column=avg_start_col, value="Parameter Hash").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=avg_start_col + 1 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=avg_start_col, value="Kunci Representatif (Hex)")
        ws.cell(row=start_row + 1, column=avg_start_col + 1, value=key_alice)
        ws.cell(row=start_row + 1, column=avg_start_col + 2, value=key_bob)
        ws.cell(row=start_row + 1, column=avg_start_col + 3, value=key_ea)
        ws.cell(row=start_row + 1, column=avg_start_col + 4, value=key_eb)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 1, min_col=avg_start_col, max_col=avg_start_col + 4):
            for cell in row:
                cell.alignment = center_align

        avg_row = start_row + 4

    ws.column_dimensions[get_column_letter(avg_start_col)].width = 28
    for col in range(avg_start_col + 1, avg_start_col + 5):
        ws.column_dimensions[get_column_letter(col)].width = 38

def build_nist_sheet(wb, records):
    ws = wb.create_sheet(title="NIST")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row = 1
    for r in records:
        ws.cell(row=current_row, column=1, value=f"Skenario {r['skenario']} - Q={r['q']}, R={r['r']}, BB={r['bb']}").font = Font(bold=True, italic=True)
        current_row += 2
        start_row = current_row

        ws.cell(row=start_row, column=1, value="Parameter NIST").font = header_font
        for idx, val in enumerate(["A & B", "E-A & E-B"]):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="Jumlah Key Lulus")
        ws.cell(row=start_row + 1, column=2, value=r['passed_keys_ab'])
        ws.cell(row=start_row + 1, column=3, value=r['passed_keys_eve'])

        ws.cell(row=start_row + 2, column=1, value="Rata-rata p-value (ApEn)")
        ws.cell(row=start_row + 2, column=2, value=r['pval_ab'])
        ws.cell(row=start_row + 2, column=3, value=r['pval_eve'])

        ws.cell(row=start_row + 3, column=1, value="Waktu Komputasi (s)")
        ws.cell(row=start_row + 3, column=2, value=r['time_nist_ab'])
        ws.cell(row=start_row + 3, column=3, value=r['time_nist_eve'])

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 3, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = center_align
        current_row = start_row + 6

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # ===============================================================
    # TABEL RATA-RATA NIST (kanan tabel existing)
    # ===============================================================
    def _base_skenario_label(val):
        txt = str(val)
        if "(Part" in txt:
            return txt.split(" (Part")[0].strip()
        return txt

    def _mean_numeric(values):
        nums = []
        for v in values:
            try:
                nums.append(float(v))
            except:
                continue
        if not nums:
            return "N/A"
        return float(np.mean(nums))

    def _sum_numeric(values):
        total = 0.0
        found = False
        for v in values:
            try:
                total += float(v)
                found = True
            except:
                continue
        if not found:
            return "N/A"
        return int(round(total))

    avg_start_col = 6
    avg_row = 1
    seen = []
    grouped = {}

    for rec in records:
        key = (_base_skenario_label(rec['skenario']), rec['q'], rec['r'], rec['bb'])
        if key not in grouped:
            grouped[key] = []
            seen.append(key)
        grouped[key].append(rec)

    for skenario_label, q, r, bb in seen:
        recs = grouped[(skenario_label, q, r, bb)]

        total_pass_ab = _sum_numeric([r['passed_keys_ab'] for r in recs])
        total_pass_eve = _sum_numeric([r['passed_keys_eve'] for r in recs])
        avg_pval_ab = _mean_numeric([r['pval_ab'] for r in recs])
        avg_pval_eve = _mean_numeric([r['pval_eve'] for r in recs])

        title_val = f"Skenario {skenario_label} - Q={q}, R={r}, BB={bb}"
        ws.cell(row=avg_row, column=avg_start_col, value=title_val).font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=avg_row, start_column=avg_start_col, end_row=avg_row, end_column=avg_start_col + 2)
        avg_row += 2

        start_row = avg_row
        ws.cell(row=start_row, column=avg_start_col, value="Parameter NIST").font = header_font
        ws.cell(row=start_row, column=avg_start_col + 1, value="A & B").font = header_font
        ws.cell(row=start_row, column=avg_start_col + 2, value="E-A & E-B").font = header_font

        ws.cell(row=start_row + 1, column=avg_start_col, value="Total Jumlah Key Lulus")
        ws.cell(row=start_row + 1, column=avg_start_col + 1, value=total_pass_ab)
        ws.cell(row=start_row + 1, column=avg_start_col + 2, value=total_pass_eve)

        ws.cell(row=start_row + 2, column=avg_start_col, value="Rata-rata p-value (ApEn)")
        ws.cell(row=start_row + 2, column=avg_start_col + 1, value=avg_pval_ab)
        ws.cell(row=start_row + 2, column=avg_start_col + 2, value=avg_pval_eve)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 2, min_col=avg_start_col, max_col=avg_start_col + 2):
            for cell in row:
                cell.alignment = center_align

        avg_row = start_row + 5

    for col in range(avg_start_col, avg_start_col + 3):
        ws.column_dimensions[get_column_letter(col)].width = 30

def build_kalman_sheet(wb, records):
    ws = wb.create_sheet(title="Kalman")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2

        start_row = current_row
        ws.cell(row=start_row, column=1, value="Parameter").font = header_font
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)

        ws.cell(row=start_row, column=2, value="Sebelum Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=5)

        ws.cell(row=start_row, column=6, value="Setelah Praproses").font = header_font
        ws.merge_cells(start_row=start_row, start_column=6, end_row=start_row, end_column=9)

        cols_names = ["Alice", "Bob", "Eve-Alice", "Eve-Bob", "Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, cname in enumerate(cols_names):
            ws.cell(row=start_row + 1, column=2 + idx, value=cname).font = header_font

        ws.cell(row=start_row + 2, column=1, value="Maksimum (dBm)")
        vals_max = [
            r['orig_max_alice'], r['orig_max_bob'], r['orig_max_evealice'], r['orig_max_evebob'],
            r['kalman_max_alice'], r['kalman_max_bob'], r['kalman_max_evealice'], r['kalman_max_evebob']
        ]
        for idx, val in enumerate(vals_max):
            ws.cell(row=start_row + 2, column=2 + idx, value=val)

        ws.cell(row=start_row + 3, column=1, value="Minimum (dBm)")
        vals_min = [
            r['orig_min_alice'], r['orig_min_bob'], r['orig_min_evealice'], r['orig_min_evebob'],
            r['kalman_min_alice'], r['kalman_min_bob'], r['kalman_min_evealice'], r['kalman_min_evebob']
        ]
        for idx, val in enumerate(vals_min):
            ws.cell(row=start_row + 3, column=2 + idx, value=val)

        ws.cell(row=start_row + 4, column=1, value="Koefisien Korelasi")
        ws.cell(row=start_row + 4, column=2, value=r['orig_corr_ab'])
        ws.merge_cells(start_row=start_row + 4, start_column=2, end_row=start_row + 4, end_column=3)
        ws.cell(row=start_row + 4, column=4, value=r['orig_corr_eve'])
        ws.merge_cells(start_row=start_row + 4, start_column=4, end_row=start_row + 4, end_column=5)
        ws.cell(row=start_row + 4, column=6, value=r['kalman_corr_ab'])
        ws.merge_cells(start_row=start_row + 4, start_column=6, end_row=start_row + 4, end_column=7)
        ws.cell(row=start_row + 4, column=8, value=r['kalman_corr_eve'])
        ws.merge_cells(start_row=start_row + 4, start_column=8, end_row=start_row + 4, end_column=9)

        ws.cell(row=start_row + 5, column=1, value="Waktu Komputasi (s)")
        ws.merge_cells(start_row=start_row + 5, start_column=1, end_row=start_row + 5, end_column=5)
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]):
            ws.cell(row=start_row + 5, column=6 + idx, value=val)

        ws.cell(row=start_row + 6, column=1, value="KGR (bit/s)")
        ws.merge_cells(start_row=start_row + 6, start_column=1, end_row=start_row + 6, end_column=5)
        for idx, val in enumerate([r['kgr_alice'], r['kgr_bob'], r['kgr_evealice'], r['kgr_evebob']]):
            ws.cell(row=start_row + 6, column=6 + idx, value=val)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 6, min_col=1, max_col=9):
            for cell in row:
                cell.alignment = center_align

        current_row = start_row + 9

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 17
    ws.column_dimensions['A'].width = 25

    # =================================================================
    # TABEL RATA-RATA KALMAN (diletakkan di kanan tabel existing)
    # =================================================================
    def _base_skenario_label(val):
        txt = str(val)
        if "(Part" in txt:
            return txt.split(" (Part")[0].strip()
        return txt

    def _mean_numeric(values):
        nums = []
        for v in values:
            try:
                nums.append(float(v))
            except:
                continue
        if not nums:
            return "N/A"
        return float(np.mean(nums))

    avg_start_col = 12  # Kolom L
    avg_row = 1
    seen = []
    grouped = {}

    for rec in records:
        key = (_base_skenario_label(rec['skenario']), rec['q'], rec['r'], rec['bb'])
        if key not in grouped:
            grouped[key] = []
            seen.append(key)
        grouped[key].append(rec)

    for skenario_label, q, r, bb in seen:
        recs = grouped[(skenario_label, q, r, bb)]

        title_val = f"Pengujian Skenario {skenario_label} - Saat Q = {q}, R = {r}, BB = {bb}"
        ws.cell(row=avg_row, column=avg_start_col, value=title_val).font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=avg_row, start_column=avg_start_col, end_row=avg_row, end_column=avg_start_col + 4)
        avg_row += 2

        start_row = avg_row
        ws.cell(row=start_row, column=avg_start_col, value="Parameter\nKalman Filter").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=avg_start_col + 1 + idx, value=val).font = header_font

        avg_corr_ab = _mean_numeric([r['kalman_corr_ab'] for r in recs])
        avg_corr_eve = _mean_numeric([r['kalman_corr_eve'] for r in recs])
        avg_kgr_alice = _mean_numeric([r['kgr_alice'] for r in recs])
        avg_kgr_bob = _mean_numeric([r['kgr_bob'] for r in recs])
        avg_kgr_ea = _mean_numeric([r['kgr_evealice'] for r in recs])
        avg_kgr_eb = _mean_numeric([r['kgr_evebob'] for r in recs])

        ws.cell(row=start_row + 1, column=avg_start_col, value="Koefisien Korelasi")
        ws.cell(row=start_row + 1, column=avg_start_col + 1, value=avg_corr_ab)
        ws.merge_cells(start_row=start_row + 1, start_column=avg_start_col + 1, end_row=start_row + 1, end_column=avg_start_col + 2)
        ws.cell(row=start_row + 1, column=avg_start_col + 3, value=avg_corr_eve)
        ws.merge_cells(start_row=start_row + 1, start_column=avg_start_col + 3, end_row=start_row + 1, end_column=avg_start_col + 4)

        ws.cell(row=start_row + 2, column=avg_start_col, value="KGR (bit/s)")
        ws.cell(row=start_row + 2, column=avg_start_col + 1, value=avg_kgr_alice)
        ws.cell(row=start_row + 2, column=avg_start_col + 2, value=avg_kgr_bob)
        ws.cell(row=start_row + 2, column=avg_start_col + 3, value=avg_kgr_ea)
        ws.cell(row=start_row + 2, column=avg_start_col + 4, value=avg_kgr_eb)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 2, min_col=avg_start_col, max_col=avg_start_col + 4):
            for cell in row:
                cell.alignment = center_align

        avg_row = start_row + 5

    ws.column_dimensions[get_column_letter(avg_start_col)].width = 25
    for col in range(avg_start_col + 1, avg_start_col + 5):
        ws.column_dimensions[get_column_letter(col)].width = 18

def build_kuantisasi_sheet(wb, records):
    ws = wb.create_sheet(title="Kuantisasi")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    current_row = 1
    for r in records:
        title_val = f"Pengujian Skenario {r['skenario']} - Saat Q = {r['q']}, R = {r['r']}, BB = {r['bb']}"
        ws.cell(row=current_row, column=1, value=title_val).font = Font(bold=True, italic=True)
        current_row += 2

        start_row = current_row
        ws.cell(row=start_row, column=1, value="Parameter Performansi").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=2 + idx, value=val).font = header_font

        ws.cell(row=start_row + 1, column=1, value="KDR (%)")
        ws.cell(row=start_row + 1, column=2, value=r['kdr_ab'])
        ws.merge_cells(start_row=start_row + 1, start_column=2, end_row=start_row + 1, end_column=3)
        ws.cell(row=start_row + 1, column=4, value=r['kdr_eve'])
        ws.merge_cells(start_row=start_row + 1, start_column=4, end_row=start_row + 1, end_column=5)

        ws.cell(row=start_row + 2, column=1, value="KGR (bit/s)")
        for idx, val in enumerate([r['kgr_alice'], r['kgr_bob'], r['kgr_evealice'], r['kgr_evebob']]):
            ws.cell(row=start_row + 2, column=2 + idx, value=val)

        ws.cell(row=start_row + 3, column=1, value="Waktu komputasi (s)")
        for idx, val in enumerate([r['time_alice'], r['time_bob'], r['time_evealice'], r['time_evebob']]):
            ws.cell(row=start_row + 3, column=2 + idx, value=val)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 3, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = center_align

        current_row = start_row + 6

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # =================================================================
    # TABEL RATA-RATA KUANTISASI (diletakkan di kanan tabel existing)
    # =================================================================
    def _base_skenario_label(val):
        txt = str(val)
        if "(Part" in txt:
            return txt.split(" (Part")[0].strip()
        return txt

    def _mean_numeric(values):
        nums = []
        for v in values:
            try:
                nums.append(float(v))
            except:
                continue
        if not nums:
            return "N/A"
        return float(np.mean(nums))

    avg_start_col = 8  # Kolom H
    avg_row = 1
    seen = []
    grouped = {}

    for rec in records:
        key = (_base_skenario_label(rec['skenario']), rec['q'], rec['r'], rec['bb'])
        if key not in grouped:
            grouped[key] = []
            seen.append(key)
        grouped[key].append(rec)

    for skenario_label, q, r, bb in seen:
        recs = grouped[(skenario_label, q, r, bb)]

        title_val = f"Pengujian Skenario {skenario_label} - Saat Q = {q}, R = {r}, BB = {bb}"
        ws.cell(row=avg_row, column=avg_start_col, value=title_val).font = Font(bold=True, italic=True)
        ws.merge_cells(start_row=avg_row, start_column=avg_start_col, end_row=avg_row, end_column=avg_start_col + 4)
        avg_row += 2

        start_row = avg_row
        ws.cell(row=start_row, column=avg_start_col, value="Parameter\nPerformansi").font = header_font
        cols = ["Alice", "Bob", "Eve-Alice", "Eve-Bob"]
        for idx, val in enumerate(cols):
            ws.cell(row=start_row, column=avg_start_col + 1 + idx, value=val).font = header_font

        avg_kdr_ab = _mean_numeric([r['kdr_ab'] for r in recs])
        avg_kdr_eve = _mean_numeric([r['kdr_eve'] for r in recs])
        avg_kgr_alice = _mean_numeric([r['kgr_alice'] for r in recs])
        avg_kgr_bob = _mean_numeric([r['kgr_bob'] for r in recs])
        avg_kgr_ea = _mean_numeric([r['kgr_evealice'] for r in recs])
        avg_kgr_eb = _mean_numeric([r['kgr_evebob'] for r in recs])

        ws.cell(row=start_row + 1, column=avg_start_col, value="KDR (%)")
        ws.cell(row=start_row + 1, column=avg_start_col + 1, value=avg_kdr_ab)
        ws.merge_cells(start_row=start_row + 1, start_column=avg_start_col + 1, end_row=start_row + 1, end_column=avg_start_col + 2)
        ws.cell(row=start_row + 1, column=avg_start_col + 3, value=avg_kdr_eve)
        ws.merge_cells(start_row=start_row + 1, start_column=avg_start_col + 3, end_row=start_row + 1, end_column=avg_start_col + 4)

        ws.cell(row=start_row + 2, column=avg_start_col, value="KGR (bit/s)")
        ws.cell(row=start_row + 2, column=avg_start_col + 1, value=avg_kgr_alice)
        ws.cell(row=start_row + 2, column=avg_start_col + 2, value=avg_kgr_bob)
        ws.cell(row=start_row + 2, column=avg_start_col + 3, value=avg_kgr_ea)
        ws.cell(row=start_row + 2, column=avg_start_col + 4, value=avg_kgr_eb)

        for row in ws.iter_rows(min_row=start_row, max_row=start_row + 2, min_col=avg_start_col, max_col=avg_start_col + 4):
            for cell in row:
                cell.alignment = center_align

        avg_row = start_row + 5

    ws.column_dimensions[get_column_letter(avg_start_col)].width = 25
    for col in range(avg_start_col + 1, avg_start_col + 5):
        ws.column_dimensions[get_column_letter(col)].width = 18

# =====================================================================
# SUMMARY EXCEL BUILDERS (Semua Skenario, Semua Part)
# =====================================================================
def _get_qr_groups(records):
    """Ambil daftar pasangan unik (q, r) dari records, urut."""
    seen = []
    for rec in records:
        key = (rec['q'], rec['r'])
        if key not in seen:
            seen.append(key)
    return seen

def _get_bb_list():
    return [p['bb'] for p in PARAM_VARIATIONS[:4]]  # [1, 5, 50, 100]

def _make_thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def build_summary_kalman_excel(output_path, all_records_by_scenario):
    """
    Buat Excel rangkuman Kalman untuk semua skenario dan semua part.
    - Setiap sheet = 1 skenario
    - Di tiap sheet, tabel dikelompokkan per pasangan Q & R
    - Setiap baris = 1 record (Part X, BB=Y)
    """
    wb = Workbook()
    wb.remove(wb.active)  # Hapus sheet default

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=11)
    qr_font = Font(bold=True, italic=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    qr_fill = PatternFill("solid", fgColor="366092")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    alt_fill = PatternFill("solid", fgColor="EEF2F8")
    thin_border = _make_thin_border()

    # Kolom: Part | Blok Data | [Alice dan Bob: Koef.Korelasi, KGR Alice, KGR Bob]
    #                          | [Eve-Alice dan Eve-Bob: Koef.Korelasi, KGR Eve-Alice, KGR Eve-Bob]
    # Total: 8 kolom
    COL_HEADERS = [
        "Part", "Blok Data",
        "Koefisien Korelasi", "KGR Alice (bit/s)", "KGR Bob (bit/s)",
        "Koefisien Korelasi", "KGR Eve-Alice (bit/s)", "KGR Eve-Bob (bit/s)"
    ]
    NUM_COLS = len(COL_HEADERS)  # 8

    for skenario, records in sorted(all_records_by_scenario.items()):
        ws = wb.create_sheet(title=f"Skenario {skenario}")

        # Judul Sheet
        ws.cell(row=1, column=1, value=f"Rekap Kalman - Skenario {skenario}").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        ws.cell(row=1, column=1).alignment = center_align

        current_row = 3
        qr_groups = _get_qr_groups(records)

        for (q, r) in qr_groups:
            # Filter records untuk pasangan Q, R ini
            qr_records = [rec for rec in records if rec['q'] == q and rec['r'] == r]

            # --- Baris Judul Q & R ---
            qr_title = f"Q = {q} ; R = {r}"
            ws.cell(row=current_row, column=1, value=qr_title)
            ws.cell(row=current_row, column=1).font = qr_font
            ws.cell(row=current_row, column=1).fill = qr_fill
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=NUM_COLS)
            current_row += 1

            # --- Baris Header Grup Atas ---
            h_row = current_row
            ws.cell(row=h_row, column=1, value="Part").font = header_font
            ws.cell(row=h_row, column=1).fill = header_fill
            ws.cell(row=h_row, column=1).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=1, end_row=h_row+1, end_column=1)

            ws.cell(row=h_row, column=2, value="Blok Data").font = header_font
            ws.cell(row=h_row, column=2).fill = header_fill
            ws.cell(row=h_row, column=2).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=2, end_row=h_row+1, end_column=2)

            ws.cell(row=h_row, column=3, value="Alice dan Bob").font = header_font
            ws.cell(row=h_row, column=3).fill = header_fill
            ws.cell(row=h_row, column=3).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=3, end_row=h_row, end_column=5)

            ws.cell(row=h_row, column=6, value="Eve-Alice dan Eve-Bob").font = header_font
            ws.cell(row=h_row, column=6).fill = header_fill
            ws.cell(row=h_row, column=6).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=6, end_row=h_row, end_column=8)

            # --- Baris Header Sub-Kolom ---
            sub_cols = ["Koefisien Korelasi", "KGR Alice (bit/s)", "KGR Bob (bit/s)",
                        "Koefisien Korelasi", "KGR Eve-Alice (bit/s)", "KGR Eve-Bob (bit/s)"]
            for idx, col_name in enumerate(sub_cols):
                c = ws.cell(row=h_row+1, column=3+idx, value=col_name)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center_align

            current_row = h_row + 2

            # --- Baris Data ---
            for i, rec in enumerate(qr_records):
                # Ambil nomor part dari string "X (Part Y)"
                skenario_str = str(rec['skenario'])
                part_str = skenario_str.split("Part ")[-1].replace(")", "").strip() if "Part" in skenario_str else skenario_str

                fill = alt_fill if i % 2 == 1 else PatternFill()

                row_vals = [
                    f"Part {part_str}",
                    rec['bb'],
                    rec['kalman_corr_ab'],
                    rec['kgr_alice'],
                    rec['kgr_bob'],
                    rec['kalman_corr_eve'],
                    rec['kgr_evealice'],
                    rec['kgr_evebob']
                ]
                for col_i, val in enumerate(row_vals):
                    cell = ws.cell(row=current_row, column=1+col_i, value=val)
                    cell.alignment = center_align
                    cell.border = thin_border
                    if fill.fgColor.value != '00000000':
                        cell.fill = fill

                current_row += 1

            current_row += 2  # Spasi antar grup Q&R

        # Lebar kolom
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        for col in range(3, NUM_COLS + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save(output_path)
    print(f"  [OK] Rekap Kalman disimpan: {output_path}")


def build_summary_kuantisasi_excel(output_path, all_records_by_scenario):
    """
    Buat Excel rangkuman Kuantisasi untuk semua skenario dan semua part.
    - Setiap sheet = 1 skenario
    - Di tiap sheet, tabel dikelompokkan per pasangan Q & R
    - Setiap baris = 1 record (Part X, BB=Y)
    """
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    qr_font = Font(bold=True, italic=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    qr_fill = PatternFill("solid", fgColor="375623")
    header_fill = PatternFill("solid", fgColor="E2EFDA")
    alt_fill = PatternFill("solid", fgColor="F2F9EE")
    thin_border = _make_thin_border()

    # Kolom: Part | Blok Data | [Alice dan Bob: KDR(%), KGR Alice, KGR Bob]
    #                          | [Eve-Alice dan Eve-Bob: KDR(%), KGR Eve-Alice, KGR Eve-Bob]
    # Total: 8 kolom
    NUM_COLS = 8

    for skenario, records in sorted(all_records_by_scenario.items()):
        ws = wb.create_sheet(title=f"Skenario {skenario}")

        ws.cell(row=1, column=1, value=f"Rekap Kuantisasi - Skenario {skenario}").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
        ws.cell(row=1, column=1).alignment = center_align

        current_row = 3
        qr_groups = _get_qr_groups(records)

        for (q, r) in qr_groups:
            qr_records = [rec for rec in records if rec['q'] == q and rec['r'] == r]

            # --- Judul Q & R ---
            qr_title = f"Q = {q} ; R = {r}"
            ws.cell(row=current_row, column=1, value=qr_title)
            ws.cell(row=current_row, column=1).font = qr_font
            ws.cell(row=current_row, column=1).fill = qr_fill
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=NUM_COLS)
            current_row += 1

            # --- Header Grup Atas ---
            h_row = current_row
            ws.cell(row=h_row, column=1, value="Part").font = header_font
            ws.cell(row=h_row, column=1).fill = header_fill
            ws.cell(row=h_row, column=1).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=1, end_row=h_row+1, end_column=1)

            ws.cell(row=h_row, column=2, value="Blok Data").font = header_font
            ws.cell(row=h_row, column=2).fill = header_fill
            ws.cell(row=h_row, column=2).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=2, end_row=h_row+1, end_column=2)

            ws.cell(row=h_row, column=3, value="Alice dan Bob").font = header_font
            ws.cell(row=h_row, column=3).fill = header_fill
            ws.cell(row=h_row, column=3).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=3, end_row=h_row, end_column=5)

            ws.cell(row=h_row, column=6, value="Eve-Alice dan Eve-Bob").font = header_font
            ws.cell(row=h_row, column=6).fill = header_fill
            ws.cell(row=h_row, column=6).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=6, end_row=h_row, end_column=8)

            # --- Header Sub-Kolom ---
            sub_cols = ["KDR (%)", "KGR Alice (bit/s)", "KGR Bob (bit/s)",
                        "KDR (%)", "KGR Eve-Alice (bit/s)", "KGR Eve-Bob (bit/s)"]
            for idx, col_name in enumerate(sub_cols):
                c = ws.cell(row=h_row+1, column=3+idx, value=col_name)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center_align

            current_row = h_row + 2

            # --- Baris Data ---
            for i, rec in enumerate(qr_records):
                skenario_str = str(rec['skenario'])
                part_str = skenario_str.split("Part ")[-1].replace(")", "").strip() if "Part" in skenario_str else skenario_str

                fill = alt_fill if i % 2 == 1 else PatternFill()

                row_vals = [
                    f"Part {part_str}",
                    rec['bb'],
                    rec['kdr_ab'],
                    rec['kgr_alice'],
                    rec['kgr_bob'],
                    rec['kdr_eve'],
                    rec['kgr_evealice'],
                    rec['kgr_evebob']
                ]
                for col_i, val in enumerate(row_vals):
                    cell = ws.cell(row=current_row, column=1+col_i, value=val)
                    cell.alignment = center_align
                    cell.border = thin_border
                    if fill.fgColor.value != '00000000':
                        cell.fill = fill

                current_row += 1

            current_row += 2

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        for col in range(3, NUM_COLS + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    wb.save(output_path)
    print(f"  [OK] Rekap Kuantisasi disimpan: {output_path}")


# =====================================================================
# MAIN ENTRY POINT
# =====================================================================
def main():
    print("=== FULL SECRET KEY GENERATION (SKG) AUTOMATION ===")
    
    # Sumber data utama tetap dari folder data, pemotongan dilakukan internal.
    base_data = "data"
    output_base = "Output100"
    
    all_kalman_records = {}
    all_kuan_records = {}
    all_bch_records = {}
    all_hash_records = {}
    all_nist_records = {}

    for skenario in SCENARIOS:
        print(f"\n>>>> Memproses Skenario {skenario} <<<<")
        
        skenario_out_dir = os.path.join(output_base, f"skenario_{skenario}")
        excel_kalman_dir = os.path.join(skenario_out_dir, "data_excel_kalman")
        excel_kuan_dir = os.path.join(skenario_out_dir, "data_excel_kuantisasi")
        os.makedirs(excel_kalman_dir, exist_ok=True)
        os.makedirs(excel_kuan_dir, exist_ok=True)
        
        kalman_records = []
        kuan_records = []
        bch_records = []
        hash_records = []
        nist_records = []
        all_kalman_records[skenario] = kalman_records
        all_kuan_records[skenario] = kuan_records
        all_bch_records[skenario] = bch_records
        all_hash_records[skenario] = hash_records
        all_nist_records[skenario] = nist_records
        
        path_alice = os.path.join(base_data, "alice", f"skenario{skenario}_mita_alice.csv")
        path_bob   = os.path.join(base_data, "bob", f"skenario{skenario}_mita_bob.csv")
        path_eve_a = os.path.join(base_data, "eve alice", f"skenario{skenario}_mita_evealice.csv")
        path_eve_b = os.path.join(base_data, "eve bob", f"skenario{skenario}_mita_evebob.csv")

        raw_alice_full = read_rssi_csv(path_alice)
        raw_bob_full = read_rssi_csv(path_bob)
        raw_eve_a_full = read_rssi_csv(path_eve_a)
        raw_eve_b_full = read_rssi_csv(path_eve_b)

        if not (raw_alice_full and raw_bob_full and raw_eve_a_full and raw_eve_b_full):
            print(f"Melewati skenario {skenario} karena data file tidak lengkap di direktori.")
            continue

        total_len = min(len(raw_alice_full), len(raw_bob_full), len(raw_eve_a_full), len(raw_eve_b_full))
        usable_len = (total_len // CHUNK_SIZE) * CHUNK_SIZE
        remainder = total_len - usable_len
        if usable_len == 0:
            print(f"Melewati skenario {skenario}: panjang data < {CHUNK_SIZE}.")
            continue
        if remainder > 0:
            print(f"Skenario {skenario}: sisa {remainder} data dibuang (chunk = {CHUNK_SIZE}).")

        num_parts = usable_len // CHUNK_SIZE
        for part in range(1, num_parts + 1):
            start_idx = (part - 1) * CHUNK_SIZE
            end_idx = start_idx + CHUNK_SIZE
            raw_alice = raw_alice_full[start_idx:end_idx]
            raw_bob = raw_bob_full[start_idx:end_idx]
            raw_eve_a = raw_eve_a_full[start_idx:end_idx]
            raw_eve_b = raw_eve_b_full[start_idx:end_idx]

            print(f"  --- Memproses Part {part} ---")

            for param in PARAM_VARIATIONS:
                q, r, bb = param['q'], param['r'], param['bb']
                
                total_len = min(len(raw_alice), len(raw_bob), len(raw_eve_a), len(raw_eve_b))
                cut_len = (total_len // bb) * bb
                
                ra_cut = raw_alice[:cut_len]
                rb_cut = raw_bob[:cut_len]
                rea_cut = raw_eve_a[:cut_len]
                reb_cut = raw_eve_b[:cut_len]
                
                # --- 1. Evaluasi Sebelum Praproses ---
                orig_max_alice = np.max(ra_cut) if cut_len > 0 else 0
                orig_max_bob = np.max(rb_cut) if cut_len > 0 else 0
                orig_max_evea = np.max(rea_cut) if cut_len > 0 else 0
                orig_max_eveb = np.max(reb_cut) if cut_len > 0 else 0
                
                orig_min_alice = np.min(ra_cut) if cut_len > 0 else 0
                orig_min_bob = np.min(rb_cut) if cut_len > 0 else 0
                orig_min_evea = np.min(rea_cut) if cut_len > 0 else 0
                orig_min_eveb = np.min(reb_cut) if cut_len > 0 else 0
                
                orig_corr_ab = calc_corr(ra_cut, rb_cut)
                orig_corr_eve = calc_corr(rea_cut, reb_cut)
                
                # --- 2. Filter Kalman Praproses (pindah ke module) ---
                kal_a, kgr_kal_a, time_kal_a = process_kalman(raw_alice, q, r, bb, benchmark_iterations=BENCHMARK_ITERATIONS)
                kal_b, kgr_kal_b, time_kal_b = process_kalman(raw_bob, q, r, bb, BENCHMARK_ITERATIONS)
                kal_ea, kgr_kal_ea, time_kal_ea = process_kalman(raw_eve_a, q, r, bb, BENCHMARK_ITERATIONS)
                kal_eb, kgr_kal_eb, time_kal_eb = process_kalman(raw_eve_b, q, r, bb, BENCHMARK_ITERATIONS)
                
                # Simpan Sinyal array ke excel (.xlsx) dengan akhiran part
                v_name = f"Q{q}_R{r}_BB{bb}_part{part}"
                save_data_list(excel_kalman_dir, f"{v_name}_kalman_alice.xlsx", kal_a, "alice_kalman")
                save_data_list(excel_kalman_dir, f"{v_name}_kalman_bob.xlsx", kal_b, "bob_kalman")
                save_data_list(excel_kalman_dir, f"{v_name}_kalman_evealice.xlsx", kal_ea, "evealice_kalman")
                save_data_list(excel_kalman_dir, f"{v_name}_kalman_evebob.xlsx", kal_eb, "evebob_kalman")
                
                kal_max_alice = np.max(kal_a) if len(kal_a)>0 else 0
                kal_max_bob = np.max(kal_b) if len(kal_b)>0 else 0
                kal_max_evea = np.max(kal_ea) if len(kal_ea)>0 else 0
                kal_max_eveb = np.max(kal_eb) if len(kal_eb)>0 else 0
                
                kal_min_alice = np.min(kal_a) if len(kal_a)>0 else 0
                kal_min_bob = np.min(kal_b) if len(kal_b)>0 else 0
                kal_min_evea = np.min(kal_ea) if len(kal_ea)>0 else 0
                kal_min_eveb = np.min(kal_eb) if len(kal_eb)>0 else 0
                
                kal_corr_ab = calc_corr(kal_a, kal_b)
                kal_corr_eve = calc_corr(kal_ea, kal_eb)
                
                kalman_records.append({
                    "skenario": f"{skenario} (Part {part})", "q": q, "r": r, "bb": bb,
                    "orig_max_alice": orig_max_alice, "orig_max_bob": orig_max_bob, "orig_max_evealice": orig_max_evea, "orig_max_evebob": orig_max_eveb,
                    "orig_min_alice": orig_min_alice, "orig_min_bob": orig_min_bob, "orig_min_evealice": orig_min_evea, "orig_min_evebob": orig_min_eveb,
                    "orig_corr_ab": orig_corr_ab, "orig_corr_eve": orig_corr_eve,
                    "kalman_max_alice": kal_max_alice, "kalman_max_bob": kal_max_bob, "kalman_max_evealice": kal_max_evea, "kalman_max_evebob": kal_max_eveb,
                    "kalman_min_alice": kal_min_alice, "kalman_min_bob": kal_min_bob, "kalman_min_evealice": kal_min_evea, "kalman_min_evebob": kal_min_eveb,
                    "kalman_corr_ab": kal_corr_ab, "kalman_corr_eve": kal_corr_eve,
                    "time_alice": time_kal_a, "time_bob": time_kal_b, "time_evealice": time_kal_ea, "time_evebob": time_kal_eb,
                    "kgr_alice": kgr_kal_a, "kgr_bob": kgr_kal_b, "kgr_evealice": kgr_kal_ea, "kgr_evebob": kgr_kal_eb
                })
                
                # --- 3. Kuantisasi Multibit 10x Iterasi (Pindah ke module) ---
                bs_a, kgr_kuan_a, time_kuan_a = process_kuantisasi(kal_a, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                bs_b, kgr_kuan_b, time_kuan_b = process_kuantisasi(kal_b, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                bs_ea, kgr_kuan_ea, time_kuan_ea = process_kuantisasi(kal_ea, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                bs_eb, kgr_kuan_eb, time_kuan_eb = process_kuantisasi(kal_eb, KUANTISASI_NUM_BITS, BENCHMARK_ITERATIONS)
                
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_alice.xlsx", [bs_a], "bitstream")
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_bob.xlsx", [bs_b], "bitstream")
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evealice.xlsx", [bs_ea], "bitstream")
                save_data_list(excel_kuan_dir, f"{v_name}_kuantisasi_evebob.xlsx", [bs_eb], "bitstream")
                
                kdr_ab = calculate_kdr(bs_a, bs_b)
                kdr_eve = calculate_kdr(bs_ea, bs_eb)
                
                kuan_records.append({
                    "skenario": f"{skenario} (Part {part})", "q": q, "r": r, "bb": bb,
                    "kdr_ab": kdr_ab, "kdr_eve": kdr_eve,
                    "time_alice": time_kuan_a, "time_bob": time_kuan_b, "time_evealice": time_kuan_ea, "time_evebob": time_kuan_eb,
                    "kgr_alice": kgr_kuan_a, "kgr_bob": kgr_kuan_b, "kgr_evealice": kgr_kuan_ea, "kgr_evebob": kgr_kuan_eb
                })

                # --- 4. BCH, Hash dan NIST Test (Simulasi modul) ---
                try:
                    from bch_module import process_bch
                    b_alice, b_bob, kdr_after_ab, kgr_bch_ab, time_bch_ab = process_bch(bs_a, bs_b, apply_correction=True)
                    b_ea, b_eb, kdr_after_eve, kgr_bch_eve, time_bch_eve = process_bch(bs_ea, bs_eb, apply_correction=False)

                    os.makedirs(os.path.join(skenario_out_dir, "data_excel_bch"), exist_ok=True)
                    save_data_list(os.path.join(skenario_out_dir, "data_excel_bch"), f"{v_name}_bch_alice.xlsx", b_alice, "alice_bch_bits")
                    save_data_list(os.path.join(skenario_out_dir, "data_excel_bch"), f"{v_name}_bch_bob.xlsx", b_bob, "bob_bch_bits")

                    bch_records.append({
                        "skenario": f"{skenario} (Part {part})", "q": q, "r": r, "bb": bb,
                        "kdr_after_ab": kdr_after_ab, "kdr_after_eve": kdr_after_eve,
                        "time_bch_ab": time_bch_ab, "time_bch_eve": time_bch_eve
                    })

                    try:
                        from hash_module import process_hash
                        h_alice, h_bob, aes_ab, time_hash_ab = process_hash(b_alice, b_bob)
                        h_ea, h_eb, aes_eve, time_hash_eve = process_hash(b_ea, b_eb)

                        os.makedirs(os.path.join(skenario_out_dir, "data_excel_hash"), exist_ok=True)
                        save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_alice.xlsx", h_alice, "AES_keys")
                        save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_bob.xlsx", h_bob, "AES_keys")
                        save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_evealice.xlsx", h_ea, "AES_keys")
                        save_data_list(os.path.join(skenario_out_dir, "data_excel_hash"), f"{v_name}_hash_evebob.xlsx", h_eb, "AES_keys")

                        hash_records.append({
                            "skenario": f"{skenario} (Part {part})", "q": q, "r": r, "bb": bb,
                            "aes_count_ab": len(aes_ab), "aes_count_eve": len(aes_eve),
                            "final_key_alice": h_alice[0] if len(h_alice) > 0 else "N/A",
                            "final_key_bob": h_bob[0] if len(h_bob) > 0 else "N/A",
                            "final_key_ea": h_ea[0] if len(h_ea) > 0 else "N/A",
                            "final_key_eb": h_eb[0] if len(h_eb) > 0 else "N/A",
                            "time_hash_ab": time_hash_ab, "time_hash_eve": time_hash_eve
                        })

                        try:
                            from nist_module import process_nist
                            pass_ab, pval_ab, time_nist_ab = process_nist(aes_ab)
                            pass_eve, pval_eve, time_nist_eve = process_nist(aes_eve)

                            nist_records.append({
                                "skenario": f"{skenario} (Part {part})", "q": q, "r": r, "bb": bb,
                                "passed_keys_ab": pass_ab, "passed_keys_eve": pass_eve,
                                "pval_ab": pval_ab, "pval_eve": pval_eve,
                                "time_nist_ab": time_nist_ab, "time_nist_eve": time_nist_eve
                            })
                        except Exception as e:
                            print("NIST Modul Error:", e)

                    except Exception as e:
                        print("Hash Modul Error:", e)
                except ImportError:
                    print("Modul BCH Belum ditaruh di root folder!")
            
        # Mengukir Tabel Excel untuk Skenario 
        print(f" == Menyusun File Table Laporan untuk Skenario {skenario} ==")
        build_kalman_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Kalman.xlsx"), kalman_records)
        build_kuantisasi_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Kuantisasi.xlsx"), kuan_records)
        if bch_records:
            build_bch_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_BCH.xlsx"), bch_records)
        if hash_records:
            build_hash_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_Hash.xlsx"), hash_records)
        if nist_records:
            build_nist_excel(os.path.join(skenario_out_dir, "Laporan_Tabel_NIST.xlsx"), nist_records)
        print("Selesai diproses untuk Skenario", skenario)

    # =====================================================================
    # Buat 2 File Excel Rangkuman Keseluruhan (Semua Skenario, Semua Part)
    # =====================================================================
    print("\n=== Menyusun Rekap Keseluruhan ===")
    summary_kalman_path = os.path.join(output_base, "Rekap_Keseluruhan_Kalman.xlsx")
    summary_kuan_path = os.path.join(output_base, "Rekap_Keseluruhan_Kuantisasi.xlsx")
    summary_bch_path = os.path.join(output_base, "Rekap_Keseluruhan_BCH.xlsx")
    summary_hash_path = os.path.join(output_base, "Rekap_Keseluruhan_Hash.xlsx")
    summary_nist_path = os.path.join(output_base, "Rekap_Keseluruhan_NIST.xlsx")

    # Hanya buat jika ada data
    if any(all_kalman_records.values()):
        build_summary_kalman_excel(summary_kalman_path, all_kalman_records)
    else:
        print("  [SKIP] Tidak ada data Kalman untuk direkap.")

    if any(all_kuan_records.values()):
        build_summary_kuantisasi_excel(summary_kuan_path, all_kuan_records)
    else:
        print("  [SKIP] Tidak ada data Kuantisasi untuk direkap.")

    print("\n==== MERANGKUM SELURUH SKENARIO KE DALAM SATU FILE EXCEL ====")
    rekap_excel_path = os.path.join(output_base, "Rekap_Evaluasi_SKG_Semua_Skenario.xlsx")
    rekap_wb = Workbook()
    if 'Sheet' in rekap_wb.sheetnames:
        rekap_wb.remove(rekap_wb['Sheet'])

    print("Menyusun sheet Kalman...")
    if any(all_kalman_records.values()):
        build_kalman_sheet(rekap_wb, [rec for records in all_kalman_records.values() for rec in records])

    print("Menyusun sheet Kuantisasi...")
    if any(all_kuan_records.values()):
        build_kuantisasi_sheet(rekap_wb, [rec for records in all_kuan_records.values() for rec in records])

    if any(all_bch_records.values()):
        print("Menyusun sheet BCH...")
        build_bch_sheet(rekap_wb, [rec for records in all_bch_records.values() for rec in records])
    else:
        print("  [SKIP] Tidak ada data BCH untuk direkap.")

    if any(all_hash_records.values()):
        print("Menyusun sheet Hash...")
        build_hash_sheet(rekap_wb, [rec for records in all_hash_records.values() for rec in records])
    else:
        print("  [SKIP] Tidak ada data Hash untuk direkap.")

    if any(all_nist_records.values()):
        print("Menyusun sheet NIST...")
        build_nist_sheet(rekap_wb, [rec for records in all_nist_records.values() for rec in records])
    else:
        print("  [SKIP] Tidak ada data NIST untuk direkap.")

    rekap_wb.save(rekap_excel_path)
    print(f"Selesai! File rekap global berhasil disimpan di: {rekap_excel_path}")

    print("\n=== SELESAI SEMUA PROSES ===")

if __name__ == "__main__":
    main()
