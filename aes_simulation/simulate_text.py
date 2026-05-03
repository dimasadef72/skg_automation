import os
import time
import pandas as pd
from openpyxl import Workbook
import random
import string
from openpyxl.styles import Alignment, Font
from Crypto.Cipher import AES

# Direktori dan Path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "data", "input")
PUBLIC_DIR = os.path.join(BASE_DIR, "data", "public")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "output")
EXCEL_SKG_PATH = os.path.join(BASE_DIR, "..", "Output", "Rekap_Evaluasi_SKG_Semua_Skenario.xlsx")
TEXT_PATH = os.path.join(INPUT_DIR, "secret_text.txt")
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, "Rekap_Simulasi_AES_Text.xlsx")

def ensure_dirs():
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(PUBLIC_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def generate_sample_text():
    if not os.path.exists(TEXT_PATH):
        with open(TEXT_PATH, "w") as f:
            f.write("RAHASIA NEGARA: Titik kumpul berada di koordinat 14A, pukul 03.00 WIB.")
        print(f"[INFO] Created dummy secret text at {TEXT_PATH}")

def get_noise_text(length=20):
    # Menghasilkan teks sampah (garbage/burek) untuk simulasi kegagalan dekripsi
    chars = string.ascii_letters + string.punctuation + "1234567890" + "¥§©®¶"
    return "".join(random.choice(chars) for _ in range(length))

def read_keys_from_skg():
    print(f"[INFO] Membaca file SKG Excel: {EXCEL_SKG_PATH}")
    if not os.path.exists(EXCEL_SKG_PATH):
        print(f"[ERROR] File SKG tidak ditemukan di {EXCEL_SKG_PATH}. Jalankan main.py terlebih dahulu.")
        return []

    try:
        df = pd.read_excel(EXCEL_SKG_PATH, sheet_name="Hash_SHA_AES", header=None)
    except Exception as e:
        print(f"[ERROR] Gagal membaca sheet Hash_SHA_AES: {e}")
        return []

    scenarios = []
    current_scenario = ""
    for index, row in df.iterrows():
        val = str(row[0]).strip()
        if val.startswith("Skenario"):
            current_scenario = val
        elif val == "Kunci Pertama (Hex)" or val == "Kunci Terbaik (Best Key by NIST)":
            alice_key = str(row[1]).strip() if pd.notna(row[1]) else ""
            bob_key = str(row[2]).strip() if pd.notna(row[2]) else ""
            ea_key = str(row[3]).strip() if pd.notna(row[3]) else ""
            eb_key = str(row[4]).strip() if pd.notna(row[4]) else ""
            
            if bob_key and bob_key != "N/A" and bob_key != "nan":
                scenarios.append({
                    "name": current_scenario,
                    "alice": alice_key,
                    "bob": bob_key,
                    "ea": ea_key,
                    "eb": eb_key
                })
    return scenarios

def encrypt_text_aes_gcm(text_path, hex_key):
    try:
        key_bytes = bytes.fromhex(hex_key)
        if len(key_bytes) < 32:
            key_bytes = key_bytes.ljust(32, b'\0')
        else:
            key_bytes = key_bytes[:32]
            
        with open(text_path, "rb") as f:
            data = f.read()
            
        start_time = time.time()
        cipher = AES.new(key_bytes, AES.MODE_GCM)
        ciphertext, tag = cipher.encrypt_and_digest(data)
        nonce = cipher.nonce
        enc_time = time.time() - start_time
        
        return {
            "success": True,
            "ciphertext": ciphertext,
            "tag": tag,
            "nonce": nonce,
            "time": enc_time
        }
    except Exception as e:
        return {"success": False, "error": str(e), "time": 0}

def decrypt_text_aes_gcm(ciphertext, tag, nonce, hex_key):
    try:
        key_bytes = bytes.fromhex(hex_key)
        if len(key_bytes) < 32:
            key_bytes = key_bytes.ljust(32, b'\0')
        else:
            key_bytes = key_bytes[:32]
            
        start_time = time.time()
        cipher = AES.new(key_bytes, AES.MODE_GCM, nonce=nonce)
        plaintext = cipher.decrypt_and_verify(ciphertext, tag)
        dec_time = time.time() - start_time
        
        return {"success": True, "time": dec_time, "text": plaintext.decode('utf-8', errors='replace')}
    except ValueError:
        return {"success": False, "error": "MAC Check Failed", "time": 0}
    except Exception as e:
        return {"success": False, "error": str(e), "time": 0}

def run_simulation():
    ensure_dirs()
    generate_sample_text()
    
    with open(TEXT_PATH, "r") as f:
        original_text = f.read()
        
    scenarios = read_keys_from_skg()
    if not scenarios:
        print("[WARN] Tidak ada skenario yang valid untuk disimulasikan.")
        return
        
    print(f"[INFO] Ditemukan {len(scenarios)} skenario. Memulai simulasi Text AES-GCM...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Simulasi Text AES"
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)
    
    # ROW 1 (Top Headers)
    ws.cell(row=1, column=1, value="Skenario")
    ws.merge_cells("A1:A2")
    
    ws.cell(row=1, column=2, value="Bob")
    ws.merge_cells("B1:C1")
    
    ws.cell(row=1, column=4, value="Chipher Text")
    ws.merge_cells("D1:D2")
    
    ws.cell(row=1, column=5, value="Alice")
    ws.merge_cells("E1:F1")
    
    ws.cell(row=1, column=7, value="Eve-Alice")
    ws.merge_cells("G1:H1")
    
    ws.cell(row=1, column=9, value="Eve-Bob")
    ws.merge_cells("I1:J1")
    
    # ROW 2 (Sub Headers)
    headers_row_2 = {
        2: "Kunci", 3: "Teks Asli",
        5: "Kunci", 6: "Hasil Decrypt",
        7: "Kunci", 8: "Hasil Decrypt",
        9: "Kunci", 10: "Hasil Decrypt"
    }
    for col_idx, text in headers_row_2.items():
        ws.cell(row=2, column=col_idx, value=text)
        
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=10):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_align
        
    # Lebar kolom
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 45 
    ws.column_dimensions['C'].width = 30 
    ws.column_dimensions['D'].width = 60 
    ws.column_dimensions['E'].width = 45 
    ws.column_dimensions['F'].width = 30 
    ws.column_dimensions['G'].width = 45 
    ws.column_dimensions['H'].width = 30 
    ws.column_dimensions['I'].width = 45 
    ws.column_dimensions['J'].width = 30 
    
    row_idx = 3
    for scen in scenarios:
        print(f"\n>> Menjalankan simulasi Text untuk: {scen['name']}")
        
        # 1. BOB ENCRYPTS
        enc_res = encrypt_text_aes_gcm(TEXT_PATH, scen["bob"])
        if not enc_res["success"]:
            print(f"   [ERROR] Bob gagal enkripsi: {enc_res['error']}")
            continue
            
        ciphertext_hex = enc_res["ciphertext"].hex().upper()
        
        safe_name = scen["name"].replace(" ", "_").replace("=", "").replace(",", "")
        public_filepath = os.path.join(PUBLIC_DIR, f"ciphertext_text_{safe_name}.enc")
        with open(public_filepath, "wb") as f:
            f.write(enc_res["nonce"] + enc_res["tag"] + enc_res["ciphertext"])
        
        # 2. ALICE DECRYPTS
        dec_alice = decrypt_text_aes_gcm(enc_res["ciphertext"], enc_res["tag"], enc_res["nonce"], scen["alice"])
        
        # 3. EVE-ALICE DECRYPTS
        dec_ea = decrypt_text_aes_gcm(enc_res["ciphertext"], enc_res["tag"], enc_res["nonce"], scen["ea"])
        
        # 4. EVE-BOB DECRYPTS
        dec_eb = decrypt_text_aes_gcm(enc_res["ciphertext"], enc_res["tag"], enc_res["nonce"], scen["eb"])
        
        # Save to Excel
        ws.row_dimensions[row_idx].height = 40 
        
        ws.cell(row=row_idx, column=1, value=scen["name"]).alignment = center_align
        ws.cell(row=row_idx, column=2, value=scen["bob"]).alignment = center_align
        ws.cell(row=row_idx, column=3, value=original_text).alignment = center_align
        ws.cell(row=row_idx, column=4, value=ciphertext_hex).alignment = center_align
        
        # Alice
        ws.cell(row=row_idx, column=5, value=scen["alice"]).alignment = center_align
        if dec_alice["success"]:
            ws.cell(row=row_idx, column=6, value=dec_alice["text"]).alignment = center_align
        else:
            ws.cell(row=row_idx, column=6, value=get_noise_text()).alignment = center_align
            
        # Eve-Alice
        ws.cell(row=row_idx, column=7, value=scen["ea"]).alignment = center_align
        if dec_ea["success"]:
            ws.cell(row=row_idx, column=8, value=dec_ea["text"]).alignment = center_align
        else:
            ws.cell(row=row_idx, column=8, value=get_noise_text()).alignment = center_align
            
        # Eve-Bob
        ws.cell(row=row_idx, column=9, value=scen["eb"]).alignment = center_align
        if dec_eb["success"]:
            ws.cell(row=row_idx, column=10, value=dec_eb["text"]).alignment = center_align
        else:
            ws.cell(row=row_idx, column=10, value=get_noise_text()).alignment = center_align
            
        row_idx += 1
        
    print(f"\n[INFO] Menyimpan laporan Excel ke {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    print("[INFO] Simulasi Selesai!")

if __name__ == "__main__":
    run_simulation()
