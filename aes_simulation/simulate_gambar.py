import os
import time
import pandas as pd
from openpyxl import Workbook
import numpy as np
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw
from Crypto.Cipher import AES
from Crypto.Random import get_random_bytes
from Crypto.Util.Padding import pad, unpad

# Direktori dan Path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "data", "input")
PUBLIC_DIR = os.path.join(BASE_DIR, "data", "public")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "output")
EXCEL_SKG_PATH = os.path.join(BASE_DIR, "..", "Output", "Rekap_Evaluasi_SKG_Semua_Skenario.xlsx")
IMAGE_PATH = os.path.join(INPUT_DIR, "secret_image.png")
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, "Rekap_Simulasi_AES_Visual.xlsx")

def ensure_dirs():
    os.makedirs(INPUT_DIR, exist_ok=True)
    os.makedirs(PUBLIC_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def generate_sample_image():
    if not os.path.exists(IMAGE_PATH):
        # Buat gambar 100x100 dengan warna gradien atau solid sederhana
        img = Image.new('RGB', (100, 100), color=(73, 109, 137))
        d = ImageDraw.Draw(img)
        d.text((10, 40), "SECRET\n DATA", fill=(255, 255, 0))
        img.save(IMAGE_PATH)
        print(f"[INFO] Created dummy secret image at {IMAGE_PATH}")

def get_noise_image_path():
    noise_path = os.path.join(OUTPUT_DIR, "noise_burek.png")
    if not os.path.exists(noise_path):
        noise_arr = np.random.randint(0, 256, (100, 100, 3), dtype=np.uint8)
        img = Image.fromarray(noise_arr, 'RGB')
        img.save(noise_path)
    return noise_path

def read_keys_from_skg():
    print(f"[INFO] Membaca file SKG Excel: {EXCEL_SKG_PATH}")
    if not os.path.exists(EXCEL_SKG_PATH):
        print(f"[ERROR] File SKG tidak ditemukan di {EXCEL_SKG_PATH}. Jalankan main.py terlebih dahulu.")
        return []

    try:
        df = pd.read_excel(EXCEL_SKG_PATH, sheet_name="Hash_SHA_AES", header=None)
    except Exception as e:
        print(f"[ERROR] Gagal membaca sheet Hash_SHA_AES: {e}")
        return []

    scenarios = []
    current_scenario = ""
    for index, row in df.iterrows():
        val = str(row[0]).strip()
        if val.startswith("Skenario"):
            current_scenario = val
        elif val == "Kunci Pertama (Hex)" or val == "Kunci Terbaik (Best Key by NIST)":
            alice_key = str(row[1]).strip() if pd.notna(row[1]) else ""
            bob_key = str(row[2]).strip() if pd.notna(row[2]) else ""
            ea_key = str(row[3]).strip() if pd.notna(row[3]) else ""
            eb_key = str(row[4]).strip() if pd.notna(row[4]) else ""
            
            if bob_key and bob_key != "N/A" and bob_key != "nan":
                scenarios.append({
                    "name": current_scenario,
                    "alice": alice_key,
                    "bob": bob_key,
                    "ea": ea_key,
                    "eb": eb_key
                })
    return scenarios

def encrypt_image_aes_gcm(image_path, hex_key):
    try:
        key_bytes = bytes.fromhex(hex_key)
        # Pad key if necessary or truncate to 32 bytes (256-bit)
        if len(key_bytes) < 32:
            key_bytes = key_bytes.ljust(32, b'\0')
        else:
            key_bytes = key_bytes[:32]
            
        with open(image_path, "rb") as f:
            data = f.read()
            
        start_time = time.time()
        cipher = AES.new(key_bytes, AES.MODE_GCM)
        ciphertext, tag = cipher.encrypt_and_digest(data)
        nonce = cipher.nonce
        enc_time = time.time() - start_time
        
        return {
            "success": True,
            "ciphertext": ciphertext,
            "tag": tag,
            "nonce": nonce,
            "time": enc_time
        }
    except Exception as e:
        return {"success": False, "error": str(e), "time": 0}

def decrypt_image_aes_gcm(ciphertext, tag, nonce, hex_key, output_temp_path):
    try:
        key_bytes = bytes.fromhex(hex_key)
        if len(key_bytes) < 32:
            key_bytes = key_bytes.ljust(32, b'\0')
        else:
            key_bytes = key_bytes[:32]
            
        start_time = time.time()
        cipher = AES.new(key_bytes, AES.MODE_GCM, nonce=nonce)
        plaintext = cipher.decrypt_and_verify(ciphertext, tag)
        dec_time = time.time() - start_time
        
        with open(output_temp_path, "wb") as f:
            f.write(plaintext)
            
        return {"success": True, "time": dec_time, "path": output_temp_path}
    except ValueError as ve:
        # MAC check failed
        return {"success": False, "error": "MAC Check Failed (Wrong Key)", "time": 0}
    except Exception as e:
        return {"success": False, "error": str(e), "time": 0}

def run_simulation():
    ensure_dirs()
    generate_sample_image()
    scenarios = read_keys_from_skg()
    if not scenarios:
        print("[WARN] Tidak ada skenario yang valid untuk disimulasikan.")
        return
        
    print(f"[INFO] Ditemukan {len(scenarios)} skenario. Memulai simulasi AES-GCM...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Simulasi AES"
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_font = Font(bold=True)
    
    # ROW 1 (Top Headers)
    ws.cell(row=1, column=1, value="Skenario")
    ws.merge_cells("A1:A2")
    
    ws.cell(row=1, column=2, value="Bob")
    ws.merge_cells("B1:C1")
    
    ws.cell(row=1, column=4, value="Chipher Text")
    ws.merge_cells("D1:D2")
    
    ws.cell(row=1, column=5, value="Alice")
    ws.merge_cells("E1:F1")
    
    ws.cell(row=1, column=7, value="Eve-Alice")
    ws.merge_cells("G1:H1")
    
    ws.cell(row=1, column=9, value="Eve-Bob")
    ws.merge_cells("I1:J1")
    
    # ROW 2 (Sub Headers)
    headers_row_2 = {
        2: "Kunci", 3: "Gambar Asli",
        5: "Kunci", 6: "Hasil Decrypt",
        7: "Kunci", 8: "Hasil Decrypt",
        9: "Kunci", 10: "Hasil Decrypt"
    }
    for col_idx, text in headers_row_2.items():
        ws.cell(row=2, column=col_idx, value=text)
        
    # Apply styling to headers
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=10):
        for cell in row:
            cell.font = header_font
            cell.alignment = center_align
        
    # Set lebar kolom
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 45 # Bob Key
    ws.column_dimensions['C'].width = 15 # Orig Image
    ws.column_dimensions['D'].width = 60 # Ciphertext
    ws.column_dimensions['E'].width = 45 # Alice Key
    ws.column_dimensions['F'].width = 15 # Alice Image
    ws.column_dimensions['G'].width = 45 # EA Key
    ws.column_dimensions['H'].width = 15 # EA Image
    ws.column_dimensions['I'].width = 45 # EB Key
    ws.column_dimensions['J'].width = 15 # EB Image
    
    row_idx = 3
    for scen in scenarios:
        print(f"\n>> Menjalankan simulasi untuk: {scen['name']}")
        
        # 1. BOB ENCRYPTS
        enc_res = encrypt_image_aes_gcm(IMAGE_PATH, scen["bob"])
        if not enc_res["success"]:
            print(f"   [ERROR] Bob gagal enkripsi: {enc_res['error']}")
            continue
            
        ciphertext_hex = enc_res["ciphertext"].hex().upper()
        
        # Save to public channel
        safe_name = scen["name"].replace(" ", "_").replace("=", "").replace(",", "")
        public_filepath = os.path.join(PUBLIC_DIR, f"ciphertext_{safe_name}.enc")
        with open(public_filepath, "wb") as f:
            f.write(enc_res["nonce"] + enc_res["tag"] + enc_res["ciphertext"])
        
        # 2. ALICE DECRYPTS
        temp_alice = os.path.join(OUTPUT_DIR, f"temp_alice_{row_idx}.png")
        dec_alice = decrypt_image_aes_gcm(
            enc_res["ciphertext"], enc_res["tag"], enc_res["nonce"], 
            scen["alice"], temp_alice
        )
        
        # 3. EVE-ALICE DECRYPTS
        temp_ea = os.path.join(OUTPUT_DIR, f"temp_ea_{row_idx}.png")
        dec_ea = decrypt_image_aes_gcm(
            enc_res["ciphertext"], enc_res["tag"], enc_res["nonce"], 
            scen["ea"], temp_ea
        )
        
        # 4. EVE-BOB DECRYPTS
        temp_eb = os.path.join(OUTPUT_DIR, f"temp_eb_{row_idx}.png")
        dec_eb = decrypt_image_aes_gcm(
            enc_res["ciphertext"], enc_res["tag"], enc_res["nonce"], 
            scen["eb"], temp_eb
        )
        
        # Save to Excel
        ws.row_dimensions[row_idx].height = 80 
        
        ws.cell(row=row_idx, column=1, value=scen["name"]).alignment = center_align
        ws.cell(row=row_idx, column=2, value=scen["bob"]).alignment = center_align
        
        # Insert Original Image (Bob)
        try:
            img_orig = OpenpyxlImage(IMAGE_PATH)
            img_orig.width, img_orig.height = 95, 95
            ws.add_image(img_orig, f"C{row_idx}")
        except Exception as e:
            ws.cell(row=row_idx, column=3, value=f"Error img: {e}")
            
        # Ciphertext
        ws.cell(row=row_idx, column=4, value=ciphertext_hex).alignment = center_align
        
        # Alice
        ws.cell(row=row_idx, column=5, value=scen["alice"]).alignment = center_align
        if dec_alice["success"]:
            try:
                img_alice = OpenpyxlImage(temp_alice)
                img_alice.width, img_alice.height = 95, 95
                ws.add_image(img_alice, f"F{row_idx}")
            except Exception:
                ws.cell(row=row_idx, column=6, value="Error load img")
        else:
            img_noise = OpenpyxlImage(get_noise_image_path())
            img_noise.width, img_noise.height = 95, 95
            ws.add_image(img_noise, f"F{row_idx}")
            
        # Eve-Alice
        ws.cell(row=row_idx, column=7, value=scen["ea"]).alignment = center_align
        if dec_ea["success"]:
            try:
                img_ea = OpenpyxlImage(temp_ea)
                img_ea.width, img_ea.height = 95, 95
                ws.add_image(img_ea, f"H{row_idx}")
            except Exception:
                pass
        else:
            img_noise = OpenpyxlImage(get_noise_image_path())
            img_noise.width, img_noise.height = 95, 95
            ws.add_image(img_noise, f"H{row_idx}")
            
        # Eve-Bob
        ws.cell(row=row_idx, column=9, value=scen["eb"]).alignment = center_align
        if dec_eb["success"]:
            try:
                img_eb = OpenpyxlImage(temp_eb)
                img_eb.width, img_eb.height = 95, 95
                ws.add_image(img_eb, f"J{row_idx}")
            except Exception:
                pass
        else:
            img_noise = OpenpyxlImage(get_noise_image_path())
            img_noise.width, img_noise.height = 95, 95
            ws.add_image(img_noise, f"J{row_idx}")
        
        # Cleanup temp Alice image if exists, wait maybe openpyxl needs it during save?
        # Openpyxl loads image data to memory when added, so we can delete temp file, but let's keep it safe and delete after loop if needed.
        # Actually it's safer to delete them after wb.save()
        
        row_idx += 1
        
    print(f"\n[INFO] Menyimpan laporan Excel ke {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    
    # Cleanup temp images
    for f in os.listdir(OUTPUT_DIR):
        if f.startswith("temp_") and f.endswith(".png"):
            try:
                os.remove(os.path.join(OUTPUT_DIR, f))
            except:
                pass
                
    print("[INFO] Simulasi Selesai!")

if __name__ == "__main__":
    run_simulation()
